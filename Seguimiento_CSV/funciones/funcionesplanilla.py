import os
import io
import time
import json
import string 
import pathlib

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from itertools import cycle
from openpyxl.utils import FORMULAE
from pdb import set_trace as bp

from . import clasesyvariables
from .clasesyvariables import usuario
from .clasesyvariables import pregunta
from .clasesyvariables import subpreguntaencuesta
from .clasesyvariables import ora


from flask_debugtoolbar_lineprofilerpanel.profile import line_profile

if(clasesyvariables.doSilent):
    time.sleep = lambda x: None 

    # def print(*args):
    #     pass


def mergeBackwards(sheet, firstCol, lastCol, row, st):
    borderCols = Border(bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
    if sheet.cell(row=row,column = firstCol).value == sheet.cell(row=row, column = lastCol).value:
        mergeBackwards(sheet, firstCol-1, lastCol, row, 1)
    else:
        if st == 1:
            sheet.merge_cells(start_row = row, start_column = firstCol+1, end_row = row, end_column= lastCol)
            for i in list(range(firstCol+1,lastCol,1)):
                sheet.cell(row=row,column=i).border = borderCols
            sheet.cell(row=row,column=firstCol+1).border = Border(left=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
            sheet.cell(row=row,column=lastCol).border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
            
            if row==2:
                sheet.cell(row=row,column=firstCol+1).alignment = Alignment(horizontal='left')
            else:
                sheet.cell(row=row,column=firstCol+1).alignment = Alignment(horizontal='center')
        sheet.cell(row=row,column=lastCol).border = Border(bottom=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'))

def LoadListaNega(ArchivoXLS):
    # print ("Cargar lista negra")
    if os.path.isfile(ArchivoXLS):
        wb_listaNegra = openpyxl.load_workbook(ArchivoXLS) #Abrir lista negra
        
        sheet =  wb_listaNegra.active
        iterator = 0

        for row in sheet['A' +str(sheet.min_row) + ":A" + str(sheet.max_row) ]:
            for cell in row:
                iterator += 1
                correo = sheet['B'+str(iterator)].value
                #print ("entoncre al usuario " + str(cell.value) + " " + str(correo) + " en lista negra")
                for usr in clasesyvariables.usuarios:
                    if(str(cell.value).find(usr.username) != -1):
                        # print("Borarre al usuario " + usr.username + " por USERNAME")
                        usrtoDelete = usr
                        clasesyvariables.usuarios.remove(usrtoDelete)
                        ##time.sleep(6)
                    elif(str(correo).find(usr.email) != -1):
                        # print("Borarre al usuario " + usr.username + " por EMAIL")
                        usrtoDelete = usr
                        clasesyvariables.usuarios.remove(usrtoDelete)
                        ##time.sleep(6)


def LoadDatosUsuario(ArchivoXLS,listaUsuarios):
    print ("Cargar lista de usuarios")
    if os.path.isfile(ArchivoXLS):
        wb_listaUsuarios = openpyxl.load_workbook(ArchivoXLS) #Abrir datos de los usuarios

        sheet = wb_listaUsuarios.active
        iterator = 0
        for row in sheet.iter_rows('F{}:F{}'.format(sheet.min_row,sheet.max_row)):
            for cell in row:
                iterator += 1
                # print ("buscar " + str(cell.value) + " en la lista")
                for usr in listaUsuarios:
                    # print(str(cell.value) + " Vs. " + usr.email)
                    if(str(cell.value).find(usr.email) != -1):
                        # print("lo encontre en "+ str(row) + " - " + str(cell))
                        usr.RUT = sheet["B" + str(iterator)].value
                        usr.ApellidoP = sheet["C" + str(iterator)].value
                        usr.ApellidoM = sheet["D" + str(iterator)].value
                        usr.nombre =  sheet["E" + str(iterator)].value
                        usr.comuna = sheet["G" + str(iterator)].value
                        usr.telefono = sheet["H" + str(iterator)].value 
                        usr.RBD = sheet["I" + str(iterator)].value
                        usr.establecimiento = sheet["J" + str(iterator)].value

                        if( len(str(usr.RUT)) < 6  ):
                            clasesyvariables.logErrores.append('El usuario ' + str(usr.nombre) + " No tiene rut o tiene un rut anomalo " + str(usr.RUT)) 

                        if( (usr.email).find("@invalid.invalid") != -1 ):
                            clasesyvariables.logErrores.append('El usuario ' + str(usr.nombre) + " tiene un correo invalido " + str(usr.email)) 

                        
        ##time.sleep(12)

    else:
        # print('No existe archivo de usuarios CPEIP para este curso.')
        clasesyvariables.logErrores.append('No existe archivo de usuarios CPEIP para este curso.')


@line_profile
def createXLS(listaUsuarios,nombre):
    global cantidadControles
    # print("Crear xls")
    merge_headers = False
    cantidadTalleres = 0
    #ReportePath = '/var/www/html/seguimiento/ReporteDescargado.xlsx'
    ReportePath = 'ReporteDescargado.xlsx'
    bookExists = os.path.isfile(ReportePath)
    if bookExists and clasesyvariables.Resetear == False:
        # print("Encontre el libro. lo abrire " + ReportePath )
        #time.sleep(4)
        book = openpyxl.load_workbook(ReportePath) 
    else:
        # print("Creare el libro")
        #time.sleep(4)
        book = openpyxl.Workbook()

    ##time.sleep(12)
    #Definición de bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    rightbottom_border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
    right_border = Border(right = Side(style = 'thin'))  
    bottom_border = Border(bottom = Side(style = 'thin'))

    ## Hay que checkear si existen previamente

    sheetCompleto = book["Reporte Completo"] if ('Reporte Completo' in book.sheetnames) else book.create_sheet("Reporte Completo",0)
    sheetControl = book["Reporte Controles"] if ('Reporte Controles' in book.sheetnames) else book.create_sheet("Reporte Controles",1)
    sheetRepAvance = book["Reporte Avance"] if ('Reporte Avance' in book.sheetnames) else book.create_sheet("Reporte Avance",2)
    if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada  > 0):
        sheetPreguntas = book["Reporte Preguntas Calificadas"] if ('Reporte Preguntas Calificadas' in book.sheetnames) else book.create_sheet("Reporte Preguntas Calificadas",3)
    sheetTest = book["Reporte Tests"] if ('Reporte Tests' in book.sheetnames) else book.create_sheet("Reporte Tests",2)
    sheetAvance = book["Datos Avance"]  if ('Datos Avance' in book.sheetnames) else book.create_sheet("Datos Avance",4)
    sheetFechas = book["Datos Fechas Avance"] if ('Datos Fechas Avance' in book.sheetnames) else book.create_sheet("Datos Fechas Avance",5)
    #sheetDatos = book["Datos Info Usuario"] if ('Datos Info Usuario' in book.sheetnames) else book.create_sheet("Datos Info Usuario",6)

    try:
        #Hoja fantasma
        #book.remove_sheet('Sheet') #funcion obsoleta
        #book.remove('Sheet')
        del book['Sheet']

    except:
        pass

    sheetCompleto.sheet_properties.tabColor="1072BA"  #Pintar pestaña azul
    sheetControl.sheet_properties.tabColor="38761D"  #Pintar pestaña verde
    sheetRepAvance.sheet_properties.tabColor="38761D"  #Pintar pestaña verde
    sheetTest.sheet_properties.tabColor="38761D"  #Pintar pestaña verde
    if(clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0):
        sheetPreguntas.sheet_properties.tabColor="38761D"
    sheetAvance.sheet_properties.tabColor="980000"  #Pintar pestaña roja
    sheetFechas.sheet_properties.tabColor="980000"  #Pintar pestaña roja
    #sheetDatos.sheet_properties.tabColor="980000"  #Pintar pestaña roja
    
    #Nominas
    sheetAvance["A1"] = "Nómina"
    sheetAvance["A4"] = "N°"
    sheetAvance["A4"].border = thin_border
    sheetAvance["B4"] = "RUT"
    sheetAvance["B4"].border = thin_border
    sheetAvance["C4"] = "Usuario edX"
    sheetAvance["C4"].border = thin_border
    sheetAvance["D4"] = "Correo"
    sheetAvance["D4"].border = thin_border
    sheetAvance["E1"] = "Taller"
    sheetAvance["E1"].border = thin_border
    sheetAvance["E2"] = "Actividad"
    sheetAvance["E2"].border = thin_border
    sheetAvance["E3"] = "Pagina"
    sheetAvance["E3"].border = thin_border
    sheetAvance["E4"] = "ID Bloque"
    sheetAvance["E4"].border = thin_border
    sheetAvance.merge_cells('A1:D3')
    sheetAvance["A1"].border = thin_border
    
    sheetFechas["A1"] = "Nómina"
    sheetFechas["A4"] = "N°"
    sheetFechas["A4"].border = thin_border
    sheetFechas["B4"] = "RUT"
    sheetFechas["B4"].border = thin_border
    sheetFechas["C4"] = "Usuario edX"
    sheetFechas["C4"].border = thin_border
    sheetFechas["D4"] = "Correo"
    sheetFechas["D4"].border = thin_border
    sheetFechas["E1"] = "Taller"
    sheetFechas["E1"].border = thin_border
    sheetFechas["E2"] = "Actividad"
    sheetFechas["E2"].border = thin_border
    sheetFechas["E3"] = "Pagina"
    sheetFechas["E3"].border = thin_border
    sheetFechas["E4"] = "ID Bloque"
    sheetFechas["E4"].border = thin_border
    sheetFechas.merge_cells('A1:D3')
    sheetFechas["A1"].border = thin_border
    
    #Casillas de reporte de avance
    sheetRepAvance["A1"] = "Nómina"
    sheetRepAvance["A2"] = "N°"
    sheetRepAvance["A2"].border = thin_border
    sheetRepAvance["B2"] = "RUT"
    sheetRepAvance["B2"].border = thin_border
    sheetRepAvance["C2"] = "Usuario edX"
    sheetRepAvance["C2"].border = thin_border
    sheetRepAvance["D2"] = "Correo"
    sheetRepAvance["D2"].border = thin_border
    sheetRepAvance.merge_cells('A1:D1')

    sheetCompleto["A1"] = "Nómina"
    sheetCompleto["A1"].border = thin_border
    sheetCompleto["A2"] = "N°"
    sheetCompleto["A2"].border = thin_border
    sheetCompleto["B2"] = "RUT"
    sheetCompleto["B2"].border = thin_border
    sheetCompleto["C2"] = "Usuario edX"
    sheetCompleto["C2"].border = thin_border
    sheetCompleto["D2"] = "Correo"
    sheetCompleto["D2"].border = thin_border
    sheetCompleto["E2"] = "Última conexión"
    sheetCompleto["E2"].border = thin_border
    sheetCompleto.merge_cells('A1:E1')

    sheetControl["A1"] = "Nómina"
    sheetControl["A1"].border = thin_border
    sheetControl["A2"] = "N°"
    sheetControl["A2"].border = thin_border
    sheetControl["B2"] = "RUT"
    sheetControl["B2"].border = thin_border
    sheetControl["C2"] = "Usuario edX"
    sheetControl["C2"].border = thin_border
    sheetControl["D2"] = "Correo"
    sheetControl["D2"].border = thin_border
    sheetControl.merge_cells('A1:D1')

    if(clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0):
        sheetPreguntas["A1"] = "Nómina"
        sheetPreguntas["A1"].border = thin_border
        sheetPreguntas["A3"] = "N°"
        sheetPreguntas["A3"].border = thin_border
        sheetPreguntas["B3"] = "RUT"
        sheetPreguntas["B3"].border = thin_border
        sheetPreguntas["C3"] = "Usuario edX"
        sheetPreguntas["C3"].border = thin_border
        sheetPreguntas["D3"] = "Correo"
        sheetPreguntas["D3"].border = thin_border
        sheetPreguntas.merge_cells('A1:D2')

    sheetTest["A1"] = "Nómina"
    sheetTest["A2"] = "N°"
    sheetTest["A2"].border = thin_border
    sheetTest["B2"] = "RUT"
    sheetTest["B2"].border = thin_border
    sheetTest["C2"] = "Usuario edX"
    sheetTest["C2"].border = thin_border
    sheetTest["D2"] = "correo"
    sheetTest["D2"].border = thin_border
    sheetTest.merge_cells('A1:D1')

    '''
    sheetDatos["A1"] = "N°"
    sheetDatos["A1"].border = thin_border
    sheetDatos["B1"] = "RUT"
    sheetDatos["B1"].border = thin_border
    sheetDatos["C1"] = "Ap. Paterno"
    sheetDatos["C1"].border = thin_border
    sheetDatos["D1"] = "Ap. Materno"
    sheetDatos["D1"].border = thin_border
    sheetDatos["E1"] = "Nombre"
    sheetDatos["E1"].border = thin_border
    sheetDatos["F1"] = "Nombre usuario (edX)"
    sheetDatos["F1"].border = thin_border
    sheetDatos["G1"] = "Mail"
    sheetDatos["G1"].border = thin_border
    sheetDatos["H1"] = "Teléfono"
    sheetDatos["H1"].border = thin_border
    sheetDatos["I1"] = "RBD"
    sheetDatos["I1"].border = thin_border
    sheetDatos["J1"] = "Establecimiento"
    sheetDatos["J1"].border = thin_border
    sheetDatos["K1"] = "Ultima Conexión"
    sheetDatos["K1"].border = thin_border
    '''

    #Llenar columnas
    columnas = []
    for char in string.ascii_uppercase:
        columnas.append(char)
    for char1 in string.ascii_uppercase:
        for char2 in string.ascii_uppercase:
            columnas.append(char1 + char2)
    for char01 in string.ascii_uppercase:
        for char02 in string.ascii_uppercase:
            for char03 in string.ascii_uppercase:
                columnas.append(char01 + char02 + char03)

    ###----------------ESTILO----------------####
    #Colores

    colorFijo = '999999'
    sheetAvance["A1"].fill = PatternFill(fgColor=colorFijo, patternType="solid")
    sheetAvance["A1"].alignment = Alignment(horizontal='center',vertical='center')
    sheetAvance["A1"].font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
    for row_range in range(4, 5+len(listaUsuarios)):
        sheetAvance["A"+str(row_range)].fill = PatternFill(fgColor=colorFijo, patternType="solid")
        sheetAvance["A"+str(row_range)].alignment = Alignment(horizontal='center',vertical='center')
        sheetAvance["A"+str(row_range)].font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
        sheetAvance["A"+str(row_range)].border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))
    for row_range in range(1, 5):
        sheetAvance["G"+str(row_range)].fill = PatternFill(fgColor=colorFijo, patternType="solid")
        sheetAvance["G"+str(row_range)].alignment = Alignment(horizontal='center',vertical='center')
        sheetAvance["G"+str(row_range)].font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
        sheetAvance["G"+str(row_range)].border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))
    
    for column_range in range(2,7):
        sheetAvance.cell(row=4,column=column_range).fill = PatternFill(fgColor=colorFijo, patternType="solid")
        sheetAvance.cell(row=4,column=column_range).alignment = Alignment(horizontal='center',vertical='center')
        sheetAvance.cell(row=4,column=column_range).font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
        sheetAvance.cell(row=4,column=column_range).border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))

    for column_range in range(2,8):
        for row_range in range(4, 5+len(listaUsuarios)):
            sheetAvance.cell(row=row_range,column=column_range).alignment = Alignment(horizontal='left',vertical='center')
            sheetAvance.cell(row=row_range,column=column_range).border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))

    sheetAvance.merge_cells("E5:E"+str(max(len(listaUsuarios)+4,5)))
    for row_range in range(4, 5+len(listaUsuarios)):
        sheetAvance["G"+str(row_range)].fill = PatternFill(fgColor=colorFijo, patternType="solid")
        sheetAvance["G"+str(row_range)].border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))
    
    iteratorColumnPreguntas = 5
    talllerTexto= ""
    #print(len(usuarioPrueba.preguntas))
    fontNegrita = Font(bold=True)
    colors1 = ['F7CAAC','FBE4D5']
    colors2 = ['BDD6EE','DEEAF6']
    pats = [cycle(colors1),cycle(colors2)]
    currPat = 0
    currentColor = next(pats[currPat])
    dontChange = 1
    #beforeLength = len(usuarioPrueba.preguntas)
    #print("usuario prueba.preguntas =" + str(usuarioPrueba.preguntas))
    usuarioDummy= clasesyvariables.usuarioPrueba.preguntas[len(clasesyvariables.usuarioPrueba.preguntas)-1]
    usuarioDummy.tallerNombre = ""
    usuarioDummy.actividadNombre = ""
    usuarioDummy.esDeControl = False
    #usuarioPrueba.preguntas.append(usuarioDummy)
    beforeLength = len( clasesyvariables.usuarioPrueba.preguntas)
    #print("antes->"+str(beforeLength)+" despues->"+str(len(usuarioPrueba.preguntas)))

    for pregunta in clasesyvariables.usuarioPrueba.preguntas: #Busqueda de preguntas en talleres   
        if (pregunta.deTaller and pregunta.tallerNumero > 0): 
            if(pregunta.tallerNombre != talllerTexto): 
                talllerTexto = pregunta.tallerNombre
                cantidadTalleres += 1
                sheetAvance[columnas[iteratorColumnPreguntas]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
                sheetFechas[columnas[iteratorColumnPreguntas]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
                #print("taller " + str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre)
                ##time.sleep(3)
                
            sheetAvance[columnas[iteratorColumnPreguntas]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre    
            sheetAvance[columnas[iteratorColumnPreguntas]+'2']= pregunta.actividadNombre
            sheetAvance[columnas[iteratorColumnPreguntas]+'3']= pregunta.paginaNumero
            sheetAvance[columnas[iteratorColumnPreguntas]+'4']= pregunta.codigo
           
            if pregunta.tallerNombre == "":
                sheetAvance[columnas[iteratorColumnPreguntas]+'1'] = ""
                sheetAvance[columnas[iteratorColumnPreguntas]+'2'] = ""
                sheetAvance[columnas[iteratorColumnPreguntas]+'3'] = ""
                sheetAvance[columnas[iteratorColumnPreguntas]+'4'] = ""


            # Fuente bold para el encabezado
            sheetAvance[columnas[iteratorColumnPreguntas]+'1'].font = fontNegrita
            sheetAvance[columnas[iteratorColumnPreguntas]+'2'].font = fontNegrita
            sheetAvance[columnas[iteratorColumnPreguntas]+'3'].font = fontNegrita
            sheetAvance[columnas[iteratorColumnPreguntas]+'4'].font = fontNegrita
            sheetAvance.column_dimensions[columnas[iteratorColumnPreguntas]].width = 3
            
            if iteratorColumnPreguntas > 5:
                shiftIterator = 6
                #Taller
                if sheetAvance.cell(row=1,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=1, column = iteratorColumnPreguntas).value or iteratorColumnPreguntas+shiftIterator == beforeLength:
                    # print("Cuando entre a talleres->"+str(iteratorColumnPreguntas)+"| 1a :"+str(sheetAvance.cell(row=2,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=2, column = iteratorColumnPreguntas).value) + " 2a : "+str(iteratorColumnPreguntas+shiftIterator == beforeLength))
                    if iteratorColumnPreguntas > 8:
                        try:
                            if merge_headers:
                                mergeBackwards(sheetAvance,iteratorColumnPreguntas-2,iteratorColumnPreguntas-1,1,0)
                        except Exception as e:
                            pass
                        currPat = 1 - currPat
                        currentColor = next(pats[currPat])
                        if (currentColor == colors1[1] or currentColor == colors2[1]):
                            currentColor = next(pats[currPat])
                        dontChange = 1
                
                #Actividad
                if sheetAvance.cell(row=2,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=2, column = iteratorColumnPreguntas).value or iteratorColumnPreguntas+shiftIterator == beforeLength:
                    # print("Cuando entre a actividades->"+str(iteratorColumnPreguntas)+"| 1a :"+str(sheetAvance.cell(row=2,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=2, column = iteratorColumnPreguntas).value) + " 2a : "+str(iteratorColumnPreguntas+shiftIterator == beforeLength))
                    if(iteratorColumnPreguntas > 2 and iteratorColumnPreguntas > 1 ):
                        try:
                            if merge_headers:
                                mergeBackwards(sheetAvance,iteratorColumnPreguntas-2,iteratorColumnPreguntas-1,2,0)
                        except:
                            pass
                        
                    if dontChange != 1:
                        currentColor = next(pats[currPat])
                    else:
                        dontChange = 0
                
                #Pagina 
                if sheetAvance.cell(row=3,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=3, column = iteratorColumnPreguntas).value or iteratorColumnPreguntas+shiftIterator == beforeLength:
                    try:
                        if merge_headers:
                            mergeBackwards(sheetAvance,iteratorColumnPreguntas-2,iteratorColumnPreguntas-1,3,0)
                    except:
                        pass

                for row_range in range(1, 5+len(listaUsuarios)):
                    cell_title = sheetAvance.cell(row_range,iteratorColumnPreguntas)                    
                    # cell_title.border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
                    # cell_title.fill = PatternFill(fgColor=currentColor, patternType="solid")
            # Hasta aca estilo

            sheetFechas[columnas[iteratorColumnPreguntas]+'2']= pregunta.actividadNombre
            sheetFechas[columnas[iteratorColumnPreguntas]+'3']= pregunta.paginaNumero
            sheetFechas[columnas[iteratorColumnPreguntas]+'4']= pregunta.codigo

            codigotocheck = pregunta.codigo

            iteratorUser = 5
            iteratorUserFechas = 5
            numeroUser = 1
            for usr in listaUsuarios:
                #iteratorColumn2 = 1

                a_it = "A" + str(iteratorUser)
                b_it = "B" + str(iteratorUser)
                c_it = "C" + str(iteratorUser)
                d_it = "D" + str(iteratorUser)

                sheetAvance[a_it] = numeroUser
                sheetAvance[b_it] = usr.RUT
                sheetAvance[c_it] = usr.username
                sheetAvance[d_it] = usr.email
                # sheetAvance[d_it].border = right_border # se saco por performance
                
                sheetFechas[a_it] = numeroUser
                sheetFechas[b_it] = usr.RUT
                sheetFechas[c_it] = usr.username
                sheetFechas[d_it] = usr.email
                # sheetFechas[d_it].border = right_border # se saco por performance

                preguntas_con_codigo_igual = [pregunta for pregunta in usr.preguntas if pregunta.codigo == codigotocheck]
                # for pregunta in usr.preguntas: # versión antigua
                for pregunta in preguntas_con_codigo_igual: # versión filtrada antes
                    # print(pregunta.codigo)
                    # if(pregunta.codigo == codigotocheck): # no es necesario en versión filtrada
                    # print('__hola__' + pregunta.codigo)
                    #sheet[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = "Listo "+ usr.nombre
                    if(pregunta.esCorrecta == True):
                        sheetAvance[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = 1
                    elif(pregunta.esCorrecta == False):     
                        sheetAvance[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = 0
                    else:
                        sheetAvance[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = 2
                    
                    if (pregunta.fechaRespuesta != None):
                        sheetFechas[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = pregunta.fechaRespuesta
                    else:
                        sheetFechas[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = "No hay fecha"

                iteratorUserFechas += 1     
                iteratorUser += 1
                numeroUser += 1

            iteratorColumnPreguntas += 1
    clasesyvariables.usuarioPrueba.preguntas.pop()   

    ##-----------------#Reporte de avance#-----------------## 

    escribirAvance = 4
    i = 0
    sheetRepAvance[columnas[escribirAvance]+"1"] = "Nº pregs:"
    
    talleres = clasesyvariables.usuarioPrueba.totalPreguntasPorTaller

    talleres = list( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller)
    # print ("pestaña de avance, Talleres: " +  str(talleres))

    espaciosAvanzadosAvance = 0
    #los diccionarios no se pueden ordenar :( asi que hay que hacer este fix
    losTalleresOrdenados = list( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller.keys() ) 

    losTalleresOrdenados.sort()
    
    sumastalleres = []
    for key in losTalleresOrdenados:
        sumataller = 0
        # print (clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[key])
        for subkey in clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[key]:
            sumataller += clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[key][subkey]
        sumastalleres.append(sumataller)

    for taller in losTalleresOrdenados:
        sheetRepAvance[columnas[escribirAvance + i ]+"2"] = "T"+str(i + 1) + " (" + str(sumastalleres[i]) + ")"
        # print("Escribir taller en pestañaa de avance [" + columnas[escribirAvance + i ] +"2]" )

        iteratorUser = 3
        #numeroUser = 1 #Revisar si es necesario sumar denuevo el numeroUser
        for usr in clasesyvariables.usuarios:

            sheetRepAvance["A" + str(iteratorUser)] = numeroUser
            sheetRepAvance["B" + str(iteratorUser)] = usr.RUT
            sheetRepAvance["C" + str(iteratorUser)] = usr.username
            sheetRepAvance["D" + str(iteratorUser)] = usr.email
            sheetRepAvance["D" + str(iteratorUser)].border = right_border


            sumapreguntas = 0
            #utototalpreguntas = []
            # print("total preguntas por taller del usuario " + usr.nombre + ": " + str(usr.totalPreguntasPorTaller))
            # print("voy en taller "+taller)
            if(len(usr.totalPreguntasPorTaller) > 0 and int(taller) in usr.totalPreguntasPorTaller):
                # print ( "Escribir "+ usr.username +" en TALLER " + str(taller) + ": " + str(usr.totalPreguntasPorTaller[int(taller)]) )
                for act in usr.totalPreguntasPorTaller[int(taller)]:
                    sumapreguntas += usr.totalPreguntasPorTaller[int(taller)][act]
                # print("Total preguntas en taller " + str(taller) + ": " + str(sumapreguntas))
                ##time.sleep(2)
            sheetRepAvance[columnas[escribirAvance + i ]+ str(iteratorUser)]  = sumapreguntas
            iteratorUser += 1
            #numeroUser+=1 #Revisar si es necesario sumar denuevo el numeroUser
            
        ##time.sleep(6)
        i+=1
        espaciosAvanzadosAvance += 1

    ## FIX TOTAL PREGUNTAS ##
    total_auxiliar = 0
    for i in [i for i in clasesyvariables.usuarioPrueba.totalPreguntasPorTaller.values()]:
        for d in i:
            if d != '-1':
                total_auxiliar += i[d]

    #usuarioPrueba.totalContestadas = float(total_auxiliar)
    if total_auxiliar == 0:
        total_auxiliar = 1
    preguntastotales = float(total_auxiliar)
    ## FIX TOTAL PREGUNTAS ##

    #Escribir total preguntas
    sheetRepAvance[columnas[escribirAvance + 1 ]+"1"] = clasesyvariables.usuarioPrueba.totalConTaller
    escribirAvance += espaciosAvanzadosAvance

    sheetRepAvance[columnas[escribirAvance]+"2"] = "Total curso (" + str( clasesyvariables.usuarioPrueba.totalConTaller) + ")"
    sheetRepAvance.column_dimensions[columnas[escribirAvance]].width = 16

    iteratorUser = 3
    for usr in clasesyvariables.usuarios:
        sheetRepAvance[columnas[escribirAvance] + str(iteratorUser)] = usr.totalConTaller #Escribo la suma
        iteratorUser += 1
    
    sheetRepAvance[columnas[escribirAvance +1 ]+"1"] = "Taller"
    sheetRepAvance[columnas[escribirAvance +1 ]+"2"] = "Actvidad"

    sheetRepAvance[columnas[escribirAvance +1 ]+"3"] = "Correctas | Contestadas"
    sheetRepAvance.column_dimensions[columnas[escribirAvance +1 ]].width = 20
    sheetRepAvance[columnas[escribirAvance +1 ]+"3"].alignment = Alignment(vertical='center')
    sheetRepAvance.merge_cells(columnas[escribirAvance +1] + "3:" + columnas[escribirAvance +1] + str(iteratorUser))



    iteratorColumnAvance = escribirAvance + 2
    talleres.sort()
    #talleresParaRecorrer = []
    
    for taller in talleres:
        #talleres = list(usuarioPrueba.totalPreguntasPorTaller)
        if(int(taller) > 0):
            actividadeslist = list( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[taller] )
            actividadeslist.sort()
            for actividad in actividadeslist:
                if(int(actividad) > 0):
                    sheetRepAvance[columnas[iteratorColumnAvance]+'1'] = taller
                    sheetRepAvance[columnas[iteratorColumnAvance]+'2'] = actividad
                    #talleresParaRecorrer.append(int(taller))
                    iteratorColumnAvance += 1

    iteratorUser = 3
    for usr in clasesyvariables.usuarios:
        #aqui debería recorrer el avance
        iteratorColumnAvance2 = escribirAvance + 2
        talleresusr = usr.totalPreguntasPorTaller
        tallerbuenasusr = usr.totalPreguntasBuenasEnviadasTaller

        # print("Preguntas totalaes " + usr.username + ": " + str(talleresusr) )
       
        # print( "Buenas del usuario  "  + usr.username + ": " + str(tallerbuenasusr) )
        #print( "Preguntas que pueden ser buenas o malas: " + str(totalPreguntasBuenasOMalas) )

        #tallerPosibles = usr.totalPreguntaBuenaOMala
        talleres = list( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller )
        #print("talleres " + str(talleres))
        talleres.sort()
        #tallerbuenasusr = list(usr.totalPreguntasCorrectasPorTaller)
        #tallerbuenasusr.sort()

        # print("talleres ordenado" + str(talleres))
        ##time.sleep(12)

        for taller in talleres: #Recorro una vez para asegurarme que el usuario de prueba contesto mas que los demas
            if taller in talleresusr:
                actividadeslist = list( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[taller] )
                actividadeslist.sort()
                for actividad in actividadeslist:
                    # print ("Taller " + str(taller) + " actividad " + str(actividad )) 
                    if(actividad in talleresusr[int(taller)]):
                        
                        conuser = talleresusr[int(taller)][int(actividad)]
                        conprueba = buprueba = clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[int(taller)][int(actividad)]
                        
                        # print ("Contestadas en T" + str(taller) + "A" + str(actividad) +" por " + usr.username +": " + str(conuser) + " Vs. " + str(conprueba) )
                        if(conuser > conprueba):
                            # print ("El usuario " + usr.username + " contesto mas que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            clasesyvariables.logErrores.append("El usuario " + usr.username + " contesto mas que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            #usuarioPrueba.totalPreguntasPorTaller[taller][actividad] = conuser
                            usr.totalPreguntasPorTaller[int(taller)][int(actividad)] = conprueba
                            #time.sleep(4)
                        
                        buuser = 0
                        if(taller in tallerbuenasusr):
                            if(actividad in tallerbuenasusr[int(taller)]):
                                buuser = tallerbuenasusr[int(taller)][int(actividad)]

                        #Por ahora no es necesarios saber las evaluables
                        '''
                        if( taller in clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalas ):
                            if( actividad in clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalas[taller] ):
                                buprueba = clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalas[taller][actividad]
                        print ("Evaluables (no PE) en T" + str(taller) + "A" + str(actividad) +" por " + usr.username +": " + str(buuser) + " Vs. " + str(buprueba) )
                        '''
                        if(buuser > buprueba):
                            # print ("El usuario " + usr.username + " contesto mas evaluables (no Preg. Calificadas) que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            clasesyvariables.logErrores.append("El usuario " + usr.username + " contesto mas evaluables (no Preg. Calificadas) que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            usr.totalPreguntasCorrectasPorTaller[taller][actividad] = buprueba
                            #totalPreguntasBuenasOMalas[taller][actividad] = buuser
                            #time.sleep(4)
                    

                        
        for taller in talleres: #Recorro una seguda vez ahora para sacar los %
            if (int(taller) in talleresusr) and int(taller) > 0:
                actividadeslist = list( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[taller] )
                actividadeslist.sort()
                for actividad in actividadeslist:
                    if int(actividad) > 0:
                        if int(actividad) in talleresusr[int(taller)]:
                            contestadas = 0
                            totalContestadas = 1
                            buenas = 0
                            totalBuenas = 1

                            # print("talleresusr " + str(talleresusr))
                            if(int(taller) in talleresusr ):
                                if(int(actividad) in talleresusr[int(taller)]):
                                    #print()
                                    contestadas = talleresusr[int(taller)][int(actividad)]

                            # print("ppreguntas totales uprueba " + str( clasesyvariables.usuarioPrueba.totalPreguntasPorTaller))
                            if( taller in clasesyvariables.usuarioPrueba.totalPreguntasPorTaller):
                                if( actividad in clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[taller] ):
                                    totalContestadas = clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[taller][actividad]
                            
                            # print("talleresusr " + str(tallerbuenasusr))
                            if(int(taller) in tallerbuenasusr):
                                if(int(actividad) in tallerbuenasusr[int(taller)]):
                                    buenas = tallerbuenasusr[int(taller)][int(actividad)]

                            '''
                            print("preguntas evaluables uprueba " + str( clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalas) )
                            if( taller in clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalas ):
                                if( actividad in clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalas[taller] ):
                                    totalBuenas = clasesyvariables.usuarioPrueba.totalPreguntasBuenasOMalasPorTaller[taller][actividad]
                            '''
                            ##time.sleep(6)
                            #print ("% Contestadas en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(contestadas) + "/" + str(totalContestadas) + "= " + str(contestadas/totalContestadas) )
                            #print ("% Buenas  en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(buenas) + "/" + str(totalBuenas) + "= " + str(buenas/totalBuenas))
                            ##time.sleep(5)

                            porContestadas = 0
                            porBuenas = 0

                            usr.sumaBuneasEnviadas += buenas
                            
                            if(totalContestadas >1  ):
                                porContestadas = int( (float(contestadas)/float(totalContestadas)) * 100.0 )
                                porBuenas =  int( (float(buenas)/float(totalContestadas))* 100.0 )

                            if(porContestadas > 100):
                                clasesyvariables.logErrores.append( "el usuario " + usr.nombre + "obtuvo un porcentaje de avance mayor a 100... [" + str(porContestadas) + "] % Contestadas en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(contestadas) + "/" + str(totalContestadas) + "= " + str(contestadas/totalContestadas) )
                                time.sleep(3)
                                porContestadas = 100
                            
                            if(porBuenas > 100):
                                clasesyvariables.logErrores.append("el usuario " + usr.nombre + "obtuvo un porcentaje de correctas mayor a 100... [" + str(porBuenas) + "] % Buenas  en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(buenas) + "/" + str(totalBuenas) + "= " + str(buenas/totalBuenas))
                                time.sleep(3)
                                porBuenas = 100

                            sheetRepAvance[columnas[iteratorColumnAvance2]+str(iteratorUser)].value  = str(porBuenas) + "% | " + str(porContestadas) + "%"
                            sheetRepAvance.column_dimensions[str(columnas[iteratorColumnAvance2])].width = 12

                        iteratorColumnAvance2 += 1
            else:
                #si no existe el taller, me muevo a la derecha segun cuantas actividades hayan 
                iteratorColumnAvance2 += len(clasesyvariables.usuarioPrueba.totalPreguntasPorTaller[taller])
        iteratorUser += 1


    ##-----------------##Tests##-----------------##

    escribirtesten = 3
    sheetTest[columnas[escribirtesten + 1]+"2"] = "Prueba Diagnóstico"
    #sheetTest[columnas[escribirtesten + 2]+"2"] = "Prueba Final"

    sheetTest[columnas[escribirtesten + 2]+"1"] = "Test"
    sheetTest[columnas[escribirtesten + 2]+"2"] = "Pregunta"
 
    totalPreguntasPret = clasesyvariables.usuarioPrueba.totalpretest
    totalPreguntasPost = clasesyvariables.usuarioPrueba.totalpostest

    if(totalPreguntasPret < 1):
        totalPreguntasPret = 10

    if(totalPreguntasPost < 1):
        totalPreguntasPost = 10
        

    puntajepretotal = totalPreguntasPret * 12
    puntajepostotal = totalPreguntasPost * 12

    #Calculo de notas
    iteratorUser = 3
    numeroUser = 1
    for usrtest in listaUsuarios:
        # print ("Usuario: " + usrtest.username + "Tests " )

        sheetTest["A" + str(iteratorUser)] = numeroUser
        sheetTest["B" + str(iteratorUser)] = usrtest.RUT
        sheetTest["C" + str(iteratorUser)] = usrtest.username
        sheetTest["D" + str(iteratorUser)] = usrtest.email
        sheetTest["D" + str(iteratorUser)].border = right_border

        puntajepreusr = 0
        puntajepostusr = 0

        if (len( usrtest.preguntas) > 0):

            preguntassumadas = 0

            for  pregu in usrtest.preguntas:
                if(pregu.pretest):
                    puntajepreusr += pregu.score
                    preguntassumadas += 1

                if(pregu.postest):
                    puntajepostusr += pregu.score

            # print("Pre-Test " + usrtest.username + ": " + str(puntajepreusr) + "/" + str(puntajepretotal) + "= " +str(puntajepreusr/puntajepretotal))
            # print("Post-Test " + usrtest.username + ": " + str(puntajepostusr) + "/" + str(puntajepostotal) + "= " +str(puntajepostusr/puntajepostotal))

            usrtest.porLogroTest = round(float(puntajepreusr/puntajepretotal),2)
            porPrePuntaje = float(puntajepreusr/puntajepretotal) * 100
            porPostPuntaje = float(puntajepostusr/puntajepostotal) * 100

            notaTestPre = 1
            
            if(porPrePuntaje < 60):
                notaTestPre = (0.05 * porPrePuntaje) + 1
            else:
                notaTestPre = (0.075 * porPrePuntaje) - 0.5 
            if(notaTestPre < 1):
                notaTestPre = 1

            notaTestPost = 1
            if(porPostPuntaje < 60):
                notaTestPost = (0.05 * porPostPuntaje) + 1
            else:
                notaTestPost = (0.075 * porPostPuntaje) - 0.5 
            if(notaTestPost < 1):
                notaTestPost = 1
            

            # print("Escribir nota de control [" + columnas[escribirtesten + 1]+str(iteratorUser)  + "] " + str(notaTestPre))
            # print("Escribir nota de control [" + columnas[escribirtesten + 2]+str(iteratorUser)  + "] " + str(notaTestPost))

            usrtest.notaTestPre = round(notaTestPre,2)
            #usrtest.notaTestPost = round(notaTestPost,2)

            if(preguntassumadas == 0):
                usrtest.porLogroTest = 999

            sheetTest[columnas[escribirtesten + 1]+str(iteratorUser)] = usrtest.porLogroTest
            #sheetTest[columnas[escribirtesten + 2]+str(iteratorUser)] = usrtest.notaTestPost

        else: #No contesto el pretest

            clasesyvariables.logErrores.append("Usuario " + usrtest.username + " No registra respuestas del pretest " )
            usrtest.porLogroTest = 999
            sheetTest[columnas[escribirtesten + 1]+str(iteratorUser)] = usrtest.porLogroTest
            usrtest.notaTestPre = 1


        iteratorUser += 1
        numeroUser += 1

    escribrpreguntatest = escribirtesten + 3 
    
    sheetTest[columnas[escribrpreguntatest]+"1"] = "Prueba de diagnóstico"

    #Preguntas del pre-test

    preguntasadded = 0
    listapreguntaspre = []

    ''' #Temporalemnte removido
    for preguntapre in clasesyvariables.usuarioPrueba.preguntas:
        if(preguntapre.pretest == True):
            preguntasadded +=1
            iteratorUser = 3

            codigotocheck = preguntapre.codigo
            sheetTest[columnas[escribrpreguntatest]+"2"] = codigotocheck

            listapreguntaspre.append(preguntapre.codigo)

            for usr in listaUsuarios:
                for preguntapreusr in usr.preguntas:
                    if(preguntapreusr.codigo == codigotocheck):
                        #print("Escribir en " + columnas[olum])
                        sheetTest[columnas[escribrpreguntatest]+str(iteratorUser)] = preguntapreusr.score
                iteratorUser += 1

            escribrpreguntatest += 1

    if preguntasadded < clasesyvariables.usuarioPrueba.totalpretest :
        clasesyvariables.logErrores.append("Esciribi menos preguntas de pretest " + str(preguntasadded) + " VS " + str(clasesyvariables.usuarioPrueba.totalpretest) )
        clasesyvariables.logErrores.append("Lista:")
        clasesyvariables.logErrores.append(str(listapreguntaspre))
    '''

    for codigopretes in clasesyvariables.listapreguntaspretest: #Por ahora este cuenta si o si las preguntas de pretest
        
        iteratorUser = 3
        sheetTest[columnas[escribrpreguntatest]+"2"] = codigopretes
        
        for usrpre in listaUsuarios:
            for pregtestusr in usrpre.preguntas:
                if(pregtestusr.codigo == codigopretes):
                    sheetTest[columnas[escribrpreguntatest]+str(iteratorUser)] = pregtestusr.score 
            iteratorUser += 1

        escribrpreguntatest += 1

    
    #Post-test Desactivado por ahora 

    """
    if(preguntasdiag > 0):
        sheetTest[columnas[escribrpreguntatest]+"1"] = "Prueba final"
    else:
        sheetTest[columnas[escribrpreguntatest+1]+"1"] = "Prueba final"

    #Preguntas del post-test
    for pregunta in usuarioPrueba.preguntas:

        if(pregunta.postest):
            iteratorUser= 3

            sheetTest[columnas[escribrpreguntatest]+"2"] = pregunta.codigo
            codigotocheck = pregunta.codigo

            for usr in listaUsuarios:
                for pregunta in usr.preguntas:
                    if(pregunta.codigo == codigotocheck):
                        #print("Escribir en " + columnas[olum])
                        sheetTest[columnas[escribrpreguntatest]+str(iteratorUser)] = pregunta.score
                        
                iteratorUser += 1

            escribrpreguntatest += 1
    """


    ##-----------------## Reporte Controles ##-----------------##

    #los diccionarios no se pueden ordenar :(
    losControlesOrdenados = list(clasesyvariables.usuarioPrueba.totalPreguntasPorControl.keys()) #revisar
    losControlesOrdenados.sort()

    iteratorColumnControl = 5 + clasesyvariables.usuarioPrueba.cantidadControles + 1
    
    i = 1
    escribircontrolen = 3
    while(i<= clasesyvariables.usuarioPrueba.cantidadControles):
        # print("Escribir en [" + columnas[escribircontrolen + i]+"1" + "]" )
        sheetControl[columnas[escribircontrolen + i]+"1"] = "C"+str(losControlesOrdenados[i-1])
        sheetControl.merge_cells(columnas[escribircontrolen + i]+"1:"+columnas[escribircontrolen + i]+"2")
        sheetControl[columnas[escribircontrolen + i]+"1"].border = thin_border
        i += 1

    #escribircontrolen =
    # print("Escribir en [" + columnas[escribircontrolen + i]+"1" + "]" )
    sheetControl[ columnas[escribircontrolen + i ] + "1"] = "Nota Control"
    sheetControl.merge_cells(columnas[escribircontrolen + i]+"1:"+columnas[escribircontrolen + i]+"2")
    sheetControl[columnas[escribircontrolen + i]+"1"].border = thin_border

    i+=1
    controlTitulo= ""

    sheetControl[columnas[escribircontrolen + i ] + "1"] = "Control"
    sheetControl[columnas[escribircontrolen + i]+"1"].border = thin_border
    sheetControl[columnas[escribircontrolen + i ] + "2"] = "Pregunta"
    sheetControl[columnas[escribircontrolen + i]+"2"].border = thin_border


    #Preguntasporcontrol = []
    numeropreguntas = 0

    for pregunta in clasesyvariables.usuarioPrueba.preguntas: #Busqueda de preguntas Control en talleres   
        if (pregunta.esDeControl):
            
            if( pregunta.actividadNombre != controlTitulo): 
                controlTitulo = pregunta.actividadNombre
                #cantidadControles += 1
                #Preguntasporcontrol.append(numeropreguntas)
                numeropreguntas = 0
                sheetControl[columnas[iteratorColumnControl]+'1'] = controlTitulo #str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
            
            #sheetControl[columnas[iteratorColumnControl]+'1']= pregunta.actividadNombre
            #sheetControl[columnas[iteratorColumnControl]+'2']= pregunta.paginaNumero #pregunta.tipo 
            sheetControl[columnas[iteratorColumnControl]+'2']= pregunta.codigo


            codigotocheck = pregunta.codigo

            iteratorUser = 3
            numeroUser = 1
            for usr in listaUsuarios:
                sheetControl["A" + str(iteratorUser)] = numeroUser
                sheetControl["B" + str(iteratorUser)] = usr.RUT
                sheetControl["C" + str(iteratorUser)] = usr.username
                sheetControl["D" + str(iteratorUser)] = usr.email
                # sheetControl["D" + str(iteratorUser)].border = right_border

                preguntas_con_codigo_igual = [pregunta for pregunta in usr.preguntas if pregunta.codigo == codigotocheck]

                # for pregunta in usr.preguntas: # versión antigua, se cambio por performance
                for pregunta in preguntas_con_codigo_igual: #versión nueva, se desindenta el if de abajo
                    # if(pregunta.codigo == codigotocheck): # no es necesario con version nueva
                    #print("Escribir en " + columnas[olum])
                    if(pregunta.score > -1):
                        sheetControl[columnas[iteratorColumnControl]+str(iteratorUser)] = pregunta.score
                        
                iteratorUser += 1
                numeroUser += 1
            iteratorColumnControl +=1
            numeropreguntas +=1

    iteratorControl = 1
    
    #Notas de control

    for control in losControlesOrdenados:
        iteratorUser = 3
        notas = 0
        for usr in listaUsuarios:
            # print ("Usuario: " + usr.username + " control " + str(control))
            totalpcontrolparaUsuario = clasesyvariables.usuarioPrueba.totalPreguntasPorControl[control]
            puntajetotal = totalpcontrolparaUsuario * 12
            puntajeusuario = 0
            preguntasContestadas = 0

            for preguntisima in usr.preguntas:
                if(preguntisima.tallerNumero == int(control) and preguntisima.esDeControl and preguntisima.latieneusuariodeprueba):
                    #print ("la pregunta " + preguntisima.codigo + " es del control " + str(control) )
                    # print(usr.username + " saco " + str(preguntisima.score) + " en pregunta "  + preguntisima.codigo)
                    ##time.sleep(1)
                    gettedScore =  preguntisima.score 
                    preguntasContestadas +=1
                    if( preguntisima.score  < 0): 
                        clasesyvariables.logErrores.append(usr.username + "tiene puntaje de "+ str(preguntisima.score) + " en control" + str(control) + " __blockkey: " + str(preguntisima.blockKeyCompleto ))
                        gettedScore = 0
                    puntajeusuario += gettedScore
            
            # print(usr.username + ": " + str(puntajeusuario) + "/" + str(puntajetotal) + "= " +str(puntajeusuario/puntajetotal))
            porPuntaje = float(puntajeusuario/puntajetotal) * 100
            ##time.sleep(4)

            if(porPuntaje < 60):
                notacontrol = (0.05 * porPuntaje) + 1
            else:
                notacontrol = (0.075 * porPuntaje) - 0.5 

            if(notacontrol < 1):
                clasesyvariables.logErrores.append(usr.username + "tiene nota de control "+ str(control) +" menor a 1: " + str(notacontrol) )
                notacontrol = 1

            # print("Escribir nota de control [" + columnas[escribircontrolen + iteratorControl]+str(iteratorUser)  + "]")

            if preguntasContestadas > 0:
                sheetControl[columnas[escribircontrolen + iteratorControl]+str(iteratorUser)] =  round(notacontrol,1)   #totalPreguntasPorControl[control]
                usr.notaControles.append(round(notacontrol,1))
            else:
                sheetControl[columnas[escribircontrolen + iteratorControl]+str(iteratorUser)] =  999  
                usr.notaControles.append(999)


            # print("cantidad de controles " + str(clasesyvariables.usuarioPrueba.cantidadControles))
            ##time.sleep(5)
            #cantidadControles = clasesyvariables.usuarioPrueba.cantidadControles

            if(iteratorControl == clasesyvariables.usuarioPrueba.cantidadControles):  
                indicenota = 0
                notas = 0
                nocontestados = 0
                while(indicenota < clasesyvariables.usuarioPrueba.cantidadControles):
                    indicenota += 1
                    notatomada = sheetControl[ columnas[escribircontrolen + indicenota] + str(iteratorUser)].value
                    if notatomada > 998:
                        notatomada = 1
                        nocontestados += 1
                    # print( "notatomada de "+  usr.username  +" [" + columnas[escribircontrolen + indicenota] +   str(iteratorUser) + "]: "+ str(notatomada))
                    ##time.sleep(3)
                    if(notatomada is None):
                        notatomada = 0
                    notas += float(notatomada)
                notafinal = notas / clasesyvariables.usuarioPrueba.cantidadControles
                if(notafinal < 1):
                    clasesyvariables.logErrores.append(usr.username + "tiene nota final de controles menor a 1: " + str(notafinal) )
                    notafinal = 1
                # print("Nota control de " + usr.username + ": " + str(notafinal))
                ##time.sleep(3)
                sheetControl[columnas[escribircontrolen + iteratorControl+1]+str(iteratorUser)] =  round(notafinal,1)   
                usr.promedioControles = round(notafinal,1)

                if nocontestados ==  clasesyvariables.usuarioPrueba.cantidadControles:
                    sheetControl[columnas[escribircontrolen + iteratorControl+1]+str(iteratorUser)] =  1.0   
                    usr.promedioControles = 1.0 # antes 999 para ninguno respondido
               

            iteratorUser += 1
            notas = 0
        iteratorControl += 1

    ##-----------------##Preguntas calificadas##-----------------##

    if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0):
        
        iteratorColumnPreva = 5 + clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada + 1

        i = 1
        escribirPrevaluen = 3
        while( i<= clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada ):
            # print("Escribir en [" + columnas[escribirPrevaluen + i]+"1" + "]" )
            sheetPreguntas[columnas[escribirPrevaluen + i]+"1"] = "PT"+str(i)
            sheetPreguntas.merge_cells(columnas[escribirPrevaluen + i]+"1:"+columnas[escribirPrevaluen + i]+"3")
            sheetPreguntas[columnas[escribirPrevaluen + i]+"1"].border = thin_border
            i += 1

        # print("Escribir en [" + columnas[escribirPrevaluen + i]+"1" + "]" )
        sheetPreguntas[ columnas[escribirPrevaluen + i ] + "1"] = "Nota PC"
        sheetPreguntas.merge_cells(columnas[escribirPrevaluen + i]+"1:"+columnas[escribirPrevaluen + i]+"3")
        sheetPreguntas[columnas[escribirPrevaluen + i]+"1"].border = thin_border
        talllerTexto= ""

        i+=1

        sheetPreguntas[columnas[escribirPrevaluen + i ] + "1"] = "Taller"
        sheetPreguntas[columnas[escribirPrevaluen + i]+ "1"].border = thin_border
        sheetPreguntas[columnas[escribirPrevaluen + i ] + "2"] = "Pregunta"
        sheetPreguntas[columnas[escribirPrevaluen + i]+ "2"].border = thin_border
        sheetPreguntas[columnas[escribirPrevaluen + i ] + "3"] = "Key"
        sheetPreguntas[columnas[escribirPrevaluen + i]+ "3"].border = thin_border


        numeropreguntas = 0

        for pregunta in clasesyvariables.usuarioPrueba.preguntas: #Busqueda de preguntas Control en talleres   
            if (pregunta.preguntaEvaluada):
                if(pregunta.tallerNombre != talllerTexto): 
                    talllerTexto = pregunta.tallerNombre
                    #cantidadControles += 1
                    #Preguntasporcontrol.append(numeropreguntas)
                    numeropreguntas = 0
                    sheetPreguntas[columnas[iteratorColumnPreva]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
                    
                sheetPreguntas[columnas[iteratorColumnPreva]+'2']= pregunta.actividadNombre
                sheetPreguntas[columnas[iteratorColumnPreva]+'3']= pregunta.codigo


                codigotocheck = pregunta.codigo

                iteratorUser = 4
                numeroUser = 1
                for usr in listaUsuarios:
                    sheetPreguntas["A" + str(iteratorUser)] = numeroUser
                    sheetPreguntas["B" + str(iteratorUser)] = usr.RUT
                    sheetPreguntas["C" + str(iteratorUser)] = usr.username
                    sheetPreguntas["D" + str(iteratorUser)] = usr.email
                    sheetPreguntas["D" + str(iteratorUser)].border = right_border

                    for pregunta in usr.preguntas:
                        if(pregunta.codigo == codigotocheck):
                            #print("Escribir en " + columnas[olum])
                            sheetPreguntas[columnas[iteratorColumnPreva]+str(iteratorUser)] = pregunta.score
                            
                    iteratorUser += 1
                    numeroUser += 1
                iteratorColumnPreva +=1
                numeropreguntas +=1

        #los diccionarios no se pueden ordenar :( asi que hay que hacer este fix        
        losTalleresPEordenados = list( clasesyvariables.usuarioPrueba.totalPreguntasEvaluadas.keys() ) #Revisar
        losTalleresPEordenados.sort()
        for preva in losTalleresPEordenados:
            iteratorUser = 4
            notas = 0
            for usr in listaUsuarios:
                # print ("Usuario: " + usr.username)
                totalprevaporUsuario = clasesyvariables.usuarioPrueba.totalPreguntasEvaluadas[preva]
                # print("Total puntaje usuario " + str(totalprevaporUsuario) +  " total pc " + str(clasesyvariables.usuarioPrueba.totalPreguntasEvaluadas) )
                puntajetotal = totalprevaporUsuario * 12
                puntajeusuario = 0
                pcrespondidas = 0

                ##time.sleep(3)

                for preguntisima in usr.preguntas:
                    if(preguntisima.tallerNumero == int(preva) and preguntisima.preguntaEvaluada):
                        #print ("la pregunta " + preguntisima.codigo + " es del control " + str(control) )
                        # print(usr.username + " saco " + str(preguntisima.score) + " en pregunta "  + preguntisima.codigo)
                        ##time.sleep(3)
                        gettedScore =  preguntisima.score 
                        pcrespondidas += 1
                        if( preguntisima.score  < 0):
                            clasesyvariables.logErrores.append(usr.username + "tiene puntaje de "+ str(preguntisima.score) + " pregunta calificada" + str(preva) + " __blockkey: " + str(preguntisima.blockKeyCompleto ))
                            gettedScore = 0
                        puntajeusuario += preguntisima.score 
                

                porPuntaje = float(puntajeusuario/puntajetotal)*100

                if(porPuntaje < 60):
                    notaPregunta = (0.05 * porPuntaje) + 1
                else:
                    notaPregunta = (0.075* porPuntaje) - 0.5 

                if(notaPregunta < 1):
                    notaPregunta = 1
                #print("preva" + preva)
                # print("Cantidad de preguntas calificadas " + str( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada ))
                # print("Escribir nota de pregunta calificada " + preva + " en " + columnas[escribirPrevaluen + int(preva)] +str(iteratorUser) + " Nota " + str(notaPregunta) )

                if pcrespondidas > 0 :
                    sheetPreguntas[columnas[escribirPrevaluen + int(preva)]+str(iteratorUser)] =  round(notaPregunta,1)   #totalPreguntasPorControl[control]
                    usr.notaPreguntasCalificadas.append(round(notaPregunta,1))
                else:
                    sheetPreguntas[columnas[escribirPrevaluen + int(preva)]+str(iteratorUser)] = 999
                    usr.notaPreguntasCalificadas.append(999)
                ##time.sleep(4)

                if( int(preva) == clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada ):  
                    indicenota = 0
                    notas = 0
                    pcNOrespondidas = 0
                    while( indicenota < clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada ):
                        indicenota += 1
                        notatomada = sheetPreguntas[ columnas[escribirPrevaluen + indicenota] + str(iteratorUser)].value
                        # print( "notatomada de "+  usr.username  +" [" + columnas[escribirPrevaluen + indicenota] +   str(iteratorUser) + "]: "+ str(notatomada))
                        ##time.sleep(3)
                        if notatomada is None :
                            notatomada = 0
                            pcNOrespondidas += 1
                        if notatomada > 998 :
                            notatomada = 1
                            pcNOrespondidas += 1

                        notas += float(notatomada)
                    notafinal = notas / clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada
                    if(notafinal < 1):
                        clasesyvariables.logErrores.append(usr.username + "tiene nota final de preguntas evaluadas menor a 1: " + str(notafinal) )
                        notafinal = 1
                    # print("Nota control de " + usr.username + ": " + str(notafinal))
                    ##time.sleep(3)

                    sheetPreguntas[columnas[escribirPrevaluen + int(preva)+1]+str(iteratorUser)] =  round(notafinal,1)    
                    usr.promedioPreCal = round(notafinal,1) 

                    if(pcNOrespondidas == clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada):
                        sheetPreguntas[columnas[escribirPrevaluen + int(preva)+1]+str(iteratorUser)] =  999    
                        usr.promedioPreCal = 1 # antes 999 para ninguna respondida

                iteratorUser += 1
                notas = 0
           
    ##-----------------##Reporte completo##-----------------##

    #-----------------#Talleres#-----------------#
    sheetCompleto["F1"] = "Avance talleres"
    sheetCompleto["F2"] = "% Correctas"
    sheetCompleto["G2"] = "% Contestadas"
    sheetCompleto["H2"] = "Nota"
    columnaincio = 7
       
    preguntastotales = 0
    preguntaEvaluables = 0

    '''
    for pregprueba in clasesyvariables.usuarioPrueba.preguntas:
        if(pregprueba.deTaller  and pregprueba.tallerNumero > 0):
            preguntastotales += 1 

        if(pregprueba.deTaller and pregprueba.tallerNumero > 0 and pregprueba.esCorrecta != None ):
            preguntaEvaluables += 1
    '''
    numeroUser = 1
    iteratorUser = 3

    #print(" Preguntas evaluables :" + str(preguntaEvaluables) + " preguntas totales " + str(preguntastotales) )
    #time.sleep(1)
    
    for estudiante in clasesyvariables.usuarios:

        sheetCompleto["A" + str(iteratorUser)] = numeroUser
        sheetCompleto["B" + str(iteratorUser)] = estudiante.RUT
        sheetCompleto["C" + str(iteratorUser)] = estudiante.username
        sheetCompleto["D" + str(iteratorUser)] = estudiante.email
        sheetCompleto["E" + str(iteratorUser)] = estudiante.ultimaconexion
        sheetCompleto["E" + str(iteratorUser)].border = right_border

        totalpreguntasestudiante = 0
        totalpreguntasbuenas = 0

        cantidaddetalleres =  len(clasesyvariables.usuarioPrueba.totalPreguntasPorTaller)

        '''
        for preg in estudiante.preguntas:
            if( preg.deTaller and preg.tallerNumero > 0):
                for preguprueba in clasesyvariables.usuarioPrueba.preguntas:
                    if preg.codigo == preguprueba.codigo:
                        totalpreguntasestudiante += 1
                        if(preg.esCorrecta):
                            totalpreguntasbuenas += 1
        
        # FIX THOMAS
        preguntastotales = float(total_auxiliar)
        # FIX THOMAS


        if float(preguntastotales) >0:
            procentajetotalcorrectas = round(float(float(estudiante.sumaBuneasEnviadas )/float(preguntastotales)) * 100,2)
        else:
            procentajetotalcorrectas = 0
        
        if float(preguntastotales) >0:
            procentajetotalpreguntas = round(float(float(estudiante.totalConTaller)/float(preguntastotales)) * 100,2)
        else:
           procentajetotalpreguntas = 0

        print ("calcular % Correctas" + str(totalpreguntasbuenas) + "/" + str(preguntaEvaluables))
        print ("calcular % Totales " + str(totalpreguntasestudiante) + "/" + str(preguntastotales))

        if(totalpreguntasestudiante > preguntastotales):
            print("el Usuario " + estudiante.nombre + " contesto mas que el usuario de prueba ")
            clasesyvariables.logErrores.append("el Usuario " + estudiante.nombre + " contesto mas preguntas de taller que el usuario de prueba " + str(totalpreguntasestudiante) + " VS. "  + str(preguntastotales) )
            time.sleep(3)
            totalpreguntasestudiante = preguntaEvaluables

        #if(preguntaEvaluables> 0):
        porNota = procentajetotalpreguntas
        #else:
        #    porNota = 0
        
        if(porNota < 60):
            nota = (0.05 * porNota) + 1
        else:
            nota = (0.075* porNota) - 0.5 

        if(nota < 1):
            nota = 1

        if(nota >= 7.0):
            clasesyvariables.logErrores.append("el Usuario " + estudiante.nombre + " tiene una nota mayor que 7 en preguntas de taller: " + str(nota) + " _%Nota " + str(porNota) +"% _%Correctas " + str(procentajetotalcorrectas) + "% _%Respondidas " + str(procentajetotalpreguntas) + "%.")
            nota = 7

        print ("Porcentajes para " + str(estudiante.username) + " %Correctas " + str(procentajetotalcorrectas) + " %Respondidas " + str(procentajetotalpreguntas) )
        print ("%Nota " + str(porNota) + " nota " + str(nota) )
        ##time.sleep(12)

        sheetCompleto["F"+str(iteratorUser)] = str(procentajetotalcorrectas) 
        sheetCompleto["G"+str(iteratorUser)] = str(procentajetotalpreguntas) 
        sheetCompleto["H"+str(iteratorUser)] = round(nota,1) 
        '''
        #Calculo de avances por el mismo excell

        #% de correctas
        topepreguntas = clasesyvariables.usuarioPrueba.totalConTaller + 5
        countif1 = "COUNTIF('Datos Avance'!F"+str(iteratorUser + 2)+":"+columnas[topepreguntas]+str(iteratorUser + 2)+",1)"
        countif2 = "COUNTIF('Datos Avance'!F"+str(iteratorUser + 2)+":"+columnas[topepreguntas]+str(iteratorUser + 2)+",2)"
        formulaCorrecta = "=ROUND(( (" + countif1 + "+" + countif2 + ")/'Reporte Avance'!F1 ) ,2)"
        sheetCompleto["F"+str(iteratorUser)] = formulaCorrecta 
        
        #% de contestadas 
        formulaContestadas = "=ROUND(('Reporte Avance'!"+ columnas[4+ cantidaddetalleres ] + str(iteratorUser) +"/'Reporte Avance'!F1),2)"
        sheetCompleto["G"+str(iteratorUser)] = formulaContestadas 
        
        #Nota 
        formulaNota = '=ROUND(IF(G'+str(iteratorUser)+'<0.6,(5*G'+str(iteratorUser)+'+1),(7.5*G'+str(iteratorUser)+'-0.5)),1)'
        sheetCompleto["H"+str(iteratorUser)] = formulaNota
        
        iteratorUser += 1
        numeroUser += 1

    #------------------------#Tests#--------------------------#

    EscribirTestCom = 7 
    sheetCompleto[columnas[EscribirTestCom + 1]+'1'] = "Tests"
    sheetCompleto[columnas[EscribirTestCom + 1]+'2'] = "Prueba de diagnóstico"

    iteratorUser = 3
    for usr in listaUsuarios:
        sheetCompleto[columnas[EscribirTestCom + 1]+ str(iteratorUser)] = usr.porLogroTest #Mostrar porcentaje de logro
        #sheetCompleto[columnas[EscribirTestCom + 2]+ str(iteratorUser)] = round(notaTestPost,2) # Por ahora no es necesario mostrar el post test

        iteratorUser += 1


    #-----------------#Preguntas calificadas#-----------------#

    if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0 ): #revisar
        i= 1
        EscribirPregEvaluada= EscribirTestCom + 1
        sheetCompleto[columnas[EscribirPregEvaluada + 1]+'1'] = "Preguntas Calificadas"

        while( i<= clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada ):
            # print ("Escribir en [" + columnas[EscribirPregEvaluada + i]+"1" + "]" )
            sheetCompleto[columnas[EscribirPregEvaluada + i]+"2"] = "PET"+str(i)
            i+=1

        sheetCompleto[columnas[EscribirPregEvaluada + i]+"2"] = "Nota preguntas calificadas"

        iteratorUser = 3
        for usr in clasesyvariables.usuarios:
            
            npregev = 0
            notaenarreglo = 0

            while npregev < clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada :
                
                npregev += 1
                # print("escribir nota de " + usr.username + " en [" + str(columnas[EscribirPregEvaluada + npregev]) + str(iteratorUser) +"] la nota " + str( usr.notaPreguntasCalificadas[notaenarreglo] ) )
                sheetCompleto[columnas[EscribirPregEvaluada + npregev]+str(iteratorUser)] =  usr.notaPreguntasCalificadas[notaenarreglo]
                notaenarreglo += 1
                
                #iteratorpregun = 0
                if( npregev == clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada ):
                    indicenota = 0
                    sheetCompleto[columnas[EscribirPregEvaluada + npregev +1 ]+str(iteratorUser)] =  usr.promedioPreCal

            iteratorUser += 1
            

    #-----------------#Controles#-----------------#

    pivote = EscribirTestCom + 1  
    if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0 ):       
        pivote = EscribirPregEvaluada
    
    i= 1
    if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0 ):
        EscribirControl =  pivote +  clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada + 1
    else:
        EscribirControl =  pivote
    
    sheetCompleto[columnas[EscribirControl + 1]+'1'] = "Controles"

    while(i<= clasesyvariables.usuarioPrueba.cantidadControles):
        # print ("Escribir en [" + columnas[EscribirControl + i]+"1" + "]" )
        sheetCompleto[columnas[EscribirControl + i]+"2"] = "C"+ str(losControlesOrdenados[i-1])
        i+=1

    sheetCompleto[columnas[EscribirControl + i]+"2"] = "Promedio Control"
    
    controlIterado = 0
    iteratorUser = 3 
    for usr in clasesyvariables.usuarios:
        
        ncontrol = 0
        notaenarreglo = 0
        
        while ncontrol < clasesyvariables.usuarioPrueba.cantidadControles:
            ncontrol += 1
            # print("escribir nota de " + usr.username + " en [" + str(columnas[EscribirControl + ncontrol]) + str(iteratorUser) +"] la nota de control " + str( usr.notaControles[notaenarreglo] ) )
            sheetCompleto[columnas[EscribirControl + ncontrol]+str(iteratorUser)] =  usr.notaControles[notaenarreglo]
            notaenarreglo += 1
            
            if( ncontrol == clasesyvariables.usuarioPrueba.cantidadControles):
                indicenota = 0
                sheetCompleto[columnas[EscribirControl + ncontrol + 1 ]+str(iteratorUser)] =   round(usr.promedioControles,2)
        iteratorUser += 1   

   

    #-----------------#Encuestas y otros#-----------------#
   
    i= 1
    # print("Encuestas " + str( clasesyvariables.usuarioPrueba.nombresEncuestas ))
    #time.sleep(2)
    cantidadEncuestas =  len( clasesyvariables.usuarioPrueba.nombresEncuestas )
    cantidadOtros =  cantidadEncuestas + 2
    Otros = ["Reglamento","Consentimiento"] + clasesyvariables.usuarioPrueba.nombresEncuestas
    espacioadicional = 0
    if clasesyvariables.usuarioPrueba.cantidadControles > 0 :
        espacioadicional += 1
    if clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0 :
        espacioadicional += 1
    EscribirOtros =  pivote +  clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada + clasesyvariables.usuarioPrueba.cantidadControles + espacioadicional
    sheetCompleto[columnas[EscribirOtros + 1]+'1'] = "Encuestas y otros"

    while(i<=cantidadOtros):
        # print ("Escribir en [" + columnas[EscribirOtros + i]+"1" + "]" )
        sheetCompleto[columnas[EscribirOtros + i]+"2"] = Otros[i-1]
        i+=1

    iteratorUser = 3
    for usren in clasesyvariables.usuarios:

        canEncuestas  = []
        j=0
        while(j<=cantidadEncuestas):
            canEncuestas.append(0)
            j+=1

        contestoreglamento = False
        contestoconsentimiento = False
        
        for encues in usren.preguntas:
            if(encues.reglamento == True):
                if(encues.respuesta.isdecimal()):
                    sheetCompleto[columnas[EscribirOtros + 1] + str(iteratorUser)] = float(encues.respuesta)
                else:
                     sheetCompleto[columnas[EscribirOtros + 1] + str(iteratorUser)] = str(encues.respuesta)
                contestoreglamento = True
                # print("Es un reglamento")

            if(encues.consentimiento == True):
                if(encues.respuesta.isdecimal()):
                    sheetCompleto[columnas[EscribirOtros + 2] + str(iteratorUser)] = float(encues.respuesta)
                else:
                    sheetCompleto[columnas[EscribirOtros + 2] + str(iteratorUser)] = str(encues.respuesta)
                contestoconsentimiento = True
                # print("Es un consentimiento")

            if encues.esDeEncuesta == True :
                # print("Es una encuesta")
                numEncuesta = encues.numeroEncuesta
                canEncuestas[numEncuesta] += 1
            

        if contestoreglamento == False:
            sheetCompleto[columnas[EscribirOtros + 1] + str(iteratorUser)] = 999

        if contestoconsentimiento == False:
            sheetCompleto[columnas[EscribirOtros + 2] + str(iteratorUser)] = 999
            
        # print( usren.username + " cantidad de encuestas " + str(cantidadEncuestas) +" "+ str(canEncuestas) + " VS "+ str( clasesyvariables.usuarioPrueba.totalEncuestas ))
        j=1
        ##time.sleep(32)
        while j <= cantidadEncuestas: #verificar si <=
            #print ("canEncuestas " + str(canEncuestas[j]) )
            #print ("totalEncuestas "  +  str(totalEncuestas[str(j)]) )
            
            if str(j) in clasesyvariables.usuarioPrueba.totalEncuestas:
                # print("j:" +str(j) + " canEncuestas " + str(canEncuestas[j]) + "/" + " totalEncuestas " +  str( clasesyvariables.usuarioPrueba.totalEncuestas[str(j)]) ) 
                sheetCompleto[columnas[EscribirOtros + 2 + j] + str(iteratorUser)] = str(canEncuestas[j]) + "/" + str( clasesyvariables.usuarioPrueba.totalEncuestas[str(j)] )
            else:
                sheetCompleto[columnas[EscribirOtros + 2 + j] + str(iteratorUser)] = 0
                clasesyvariables.logErrores.append("No se encontro encuesta [" + str(j) + "] en el usuario "+ str(usren.username) + " ,encuestas usr prueba " + str(clasesyvariables.usuarioPrueba.totalEncuestas)  + " ["+ str(clasesyvariables.usuarioPrueba.nombresEncuestas) +"]" )
            j+= 1

        iteratorUser += 1

    #-----------------#Asistencias #-----------------#

    # print("ASISTENCIAS")
    pivote = EscribirOtros + cantidadOtros

    escribirAsistencia = pivote + 1
    sheetCompleto[columnas[escribirAsistencia]+'1'] = "Asistencia"

    cantidadAsistencias = len( clasesyvariables.usuarioPrueba.nombresAsistencias )
    asistencias = clasesyvariables.usuarioPrueba.nombresAsistencias
    # print(str(asistencias))
    
    k= 1

    
    while(k<=cantidadAsistencias):
        # print ("Escribir asistencia nombre en [" + columnas[escribirAsistencia + k-1]+"2" + "]" )

        nombrecodigoasistencia = str(asistencias[k-1]).split("_:_")
        nombreasist = nombrecodigoasistencia[0]
        #codigoasist = nombrecodigoasistencia[1]

        sheetCompleto[columnas[escribirAsistencia + k -1]+"2"] = nombreasist

        iteratorUser = 3

        for usras in clasesyvariables.usuarios:
            for asist in usras.preguntas:
                asist_score = 999
                if asist.esAsistencia:
                    asist_nombre = str(asist.nombreAsistencia)
                    asist_score = int(asist.score)
                    print("asistencia encontrada " + asist_nombre +  " score obtenido " + str(asist_score))
                    print(asist_nombre + " VS " + nombreasist)
                    if asist_nombre == nombreasist :
                        print ("coinciden")
                        print("score a escribir " + str(asist_score))
                        sheetCompleto[columnas[escribirAsistencia + k -1 ] + str(iteratorUser)] = asist_score
                    else:
                        print("no coinciden")
            iteratorUser += 1


        k+=1

    ##-------------------##Situación Actual##-------------------##
           
    # print("Situacion Actual")
    escribirSituacion = escribirAsistencia + 1
    if cantidadAsistencias > 0:
        escribirSituacion = escribirAsistencia + cantidadAsistencias
    iteratorUser = 3

    sheetCompleto[columnas[escribirSituacion]+'1'] = "Situación"
    sheetCompleto[columnas[escribirSituacion]+'2'] = "Situación actual"
    for usrsi in clasesyvariables.usuarios:
        situsccore = 0
        situsccore = int(usrsi.situacionactual)
        sheetCompleto[columnas[escribirSituacion] + str(iteratorUser)] = situsccore
        iteratorUser += 1


    ##-------------------##Pestaña de datos##-------------------##
    ''' #No existe por ahora
    iteratorUser = 1

    for usurdatos in usuarios:
        iteratorUser += 1
        sheetDatos['A' + str(iteratorUser)] = iteratorUser-1
        sheetDatos['B' + str(iteratorUser)] = usurdatos.RUT
        sheetDatos['C' + str(iteratorUser)] = usurdatos.ApellidoP
        sheetDatos['D' + str(iteratorUser)] = usurdatos.ApellidoM
        sheetDatos['E' + str(iteratorUser)] = usurdatos.nombre 
        sheetDatos['F' + str(iteratorUser)] = usurdatos.username
        sheetDatos['G' + str(iteratorUser)] = usurdatos.email
        sheetDatos['H' + str(iteratorUser)] = usurdatos.telefono
        sheetDatos['I' + str(iteratorUser)] = usurdatos.RBD
        sheetDatos['J' + str(iteratorUser)] = usurdatos.establecimiento
        sheetDatos['K' + str(iteratorUser)] = usurdatos.ultimaconexion
    '''
    ##----------------##Limpieza de filas##-----------------------##

    if(clasesyvariables.Limpiar):
        iniciolimpieza = len(clasesyvariables.usuarios)

        inicioCompleto = inicioControles = inicioRepAvance = 2
        inicioPregEva = 3
        inicioAvance = inicioFechas =  4
        
        sheetCompleto.delete_rows(iniciolimpieza + inicioCompleto + 1,len(clasesyvariables.usuarios))
        sheetAvance.delete_rows(iniciolimpieza + inicioAvance + 1, len(clasesyvariables.usuarios))
        sheetControl.delete_rows(iniciolimpieza + inicioControles + 1, len(clasesyvariables.usuarios))
        if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada > 0):
            sheetPreguntas.delete_rows(iniciolimpieza + inicioPregEva + 1, len(clasesyvariables.usuarios))
        sheetRepAvance.delete_rows(iniciolimpieza + inicioRepAvance + 1, len(clasesyvariables.usuarios))
        sheetFechas.delete_rows(iniciolimpieza + inicioFechas + 1, len(clasesyvariables.usuarios))
    
    ##-----------------##Ajustes de espacio##-----------------##

    sheetAvance.column_dimensions['A'].width = 3
    sheetAvance.column_dimensions['B'].width = 12
    sheetAvance.column_dimensions['C'].width = 16
    sheetAvance.column_dimensions['D'].width = 32

    sheetFechas.column_dimensions['A'].width = 3
    sheetFechas.column_dimensions['B'].width = 12
    sheetFechas.column_dimensions['C'].width = 16
    sheetFechas.column_dimensions['D'].width = 32

    sheetCompleto.column_dimensions['A'].width = 6
    sheetCompleto.column_dimensions['B'].width = 12
    sheetCompleto.column_dimensions['C'].width = 16
    sheetCompleto.column_dimensions['D'].width = 32
    sheetCompleto.column_dimensions['E'].width = 24


    sheetControl.column_dimensions['A'].width = 3
    sheetControl.column_dimensions['B'].width = 12
    sheetControl.column_dimensions['C'].width = 16
    sheetControl.column_dimensions['D'].width = 32


    sheetTest.column_dimensions['A'].width = 3
    sheetTest.column_dimensions['B'].width = 12
    sheetTest.column_dimensions['C'].width = 16
    sheetTest.column_dimensions['D'].width = 32
    sheetTest.column_dimensions['E'].width = 18


    if( clasesyvariables.usuarioPrueba.cantidadTipoPreguntaEvaluada >0):
        sheetPreguntas.column_dimensions['A'].width = 3
        sheetPreguntas.column_dimensions['B'].width = 12
        sheetPreguntas.column_dimensions['C'].width = 16
        sheetPreguntas.column_dimensions['D'].width = 32


    #Casillas de reporte de avance
    sheetRepAvance.column_dimensions['A'].width = 3
    sheetRepAvance.column_dimensions['B'].width = 12
    sheetRepAvance.column_dimensions['C'].width = 16
    sheetRepAvance.column_dimensions['D'].width = 32

    '''
    sheetDatos.column_dimensions['A'].width = 3
    sheetDatos.column_dimensions['B'].width = 12
    sheetDatos.column_dimensions['C'].width = 16
    sheetDatos.column_dimensions['D'].width = 16
    sheetDatos.column_dimensions['E'].width = 20
    sheetDatos.column_dimensions['F'].width = 14
    sheetDatos.column_dimensions['G'].width = 30
    sheetDatos.column_dimensions['H'].width = 10
    sheetDatos.column_dimensions['I'].width = 10
    sheetDatos.column_dimensions['J'].width = 16
    sheetDatos.column_dimensions['k'].width = 20
    '''

    ##-----------------##Guardado del archivo##-----------------##

    nombrereporte = "Reporte_" + nombre + ".xlsx"

    if( clasesyvariables.location_to_save_report == "" ):
        # location = os.path.join(os.path.dirname(__file__),"Generado/Xls", nombrereporte)
        dir_carpeta_generado = str(sorted(pathlib.Path('.').glob('**/Generado'))[0].absolute())
        location = os.path.join(dir_carpeta_generado,"Xls", nombrereporte)
    else:
        location =  clasesyvariables.location_to_save_report
    # print("Location: " + str(location))
    if os.path.isfile(location):
        os.remove(location)
    # breakpoint()
    book.save(location)
    return location



def createDocumentoEncuesta(listaEncuesta,nombre):
    global cantidadControles
    # print("Crear xls encuesta")

    #cantidadTalleres = 0
    book2 = openpyxl.Workbook()

    #Definición de bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    rightbottom_border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
    right_border = Border(right = Side(style = 'thin'))  
    bottom_border = Border(bottom = Side(style = 'thin'))

    #sheet.merge_cells('A2:D4') 

    #Llenar columnas
    columnas = []
    for char in string.ascii_uppercase:
        columnas.append(char)
    for char1 in string.ascii_uppercase:
        for char2 in string.ascii_uppercase:
            columnas.append(char1 + char2)
    for char01 in string.ascii_uppercase:
        for char02 in string.ascii_uppercase:
            for char03 in string.ascii_uppercase:
                columnas.append(char01 + char02 + char03)

    sheet = book2.active
    sheet.title = "Encuesta"
    sheet.sheet_properties.tabColor="470405"  

    sheet["A1"] = "N°"
    sheet["A1"].border = thin_border
    sheet["B1"] = "RUT"
    sheet["B1"].border = thin_border
    sheet["C1"] = "Nombre"
    sheet["C1"].border = thin_border
    sheet["D1"] = "correo"
    sheet["D1"].border = thin_border
    sheet["E1"] = "Encuesta N°"
    sheet["E1"].border = thin_border
    sheet["F1"] = "Encuesta"
    sheet["F1"].border = thin_border
    sheet["G1"] = "Pag"
    sheet["G1"].border = thin_border
    sheet["H1"] = "pregunta"
    sheet["H1"].border = thin_border
    sheet["I1"] = "Identificador"
    sheet["I1"].border = thin_border
    sheet["J1"] = "Respuesta"
    sheet["J1"].border = thin_border
    sheet["k1"] = "Link"
    sheet["K1"].border = thin_border
    

    iterator = 2
    numero = 0
    for encu in listaEncuesta:
        if(encu.nombreusuario != clasesyvariables.nombreUsuarioPrueba and encu.nombrepregunta != '' and encu.respuesta != ''):
            numero += 1
            sheet["A" + str(iterator)] = numero
            sheet["B" + str(iterator)] = encu.usuariorut
            sheet["C" + str(iterator)] = encu.nombreusuario
            sheet["D" + str(iterator)] = encu.usuariocorreo
            sheet["E" + str(iterator)] = encu.numeroencuesta
            sheet["F" + str(iterator)] = encu.nombreencuesta
            sheet["G" + str(iterator)] = encu.pagina
            sheet["H" + str(iterator)] = encu.nombrepregunta
            sheet["I" + str(iterator)] = str(encu.numeroidentificatorio1) + "_" + str(encu.numeroidentificatorio2)
            respu = str(encu.respuesta)

            if respu.isdecimal():
                sheet["J" + str(iterator)] = float(respu)
            else:
                sheet["J" + str(iterator)] = str(respu)

            sheet["K" + str(iterator)] = "https://studio.cmmeduformacion.uchile.cl/container/" + encu.blockkey
            iterator += 1
    
    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 32
    sheet.column_dimensions['E'].width = 8
    sheet.column_dimensions['F'].width = 14
    sheet.column_dimensions['G'].width = 5
    sheet.column_dimensions['I'].width = 8
    sheet.column_dimensions['H'].width = 16   
    sheet.column_dimensions['J'].width = 40    
    sheet.column_dimensions['K'].width = 30
    sheet.column_dimensions['L'].width = 80
    
    filename = "Encuestas_" + nombre + ".xlsx"

    if( clasesyvariables.location_to_save_report == "") :
        # location = os.path.join(os.path.dirname(__file__),"Generado/Xls", nombrereporte)
        dir_carpeta_generado = str(sorted(pathlib.Path('.').glob('**/Generado'))[0].absolute())
        location = os.path.join(dir_carpeta_generado,"Xls", filename)
    else:
        location =  clasesyvariables.location_to_save_report
    # print("Location: " + str(location))
    book2.save(location)
    return location
