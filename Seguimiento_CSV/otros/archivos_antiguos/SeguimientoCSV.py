#!/usr/bin/env python
# -*- encoding: utf-8 -*-

import sys
from datetime import datetime
import zipfile
from tempfile import NamedTemporaryFile
import csv
import os
import io
import time
import json
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from itertools import cycle
import string  
#import pyexcel as p
import xml.etree.ElementTree as ET
import driveapi
from pdb import set_trace as bp

import subprocess #para probar el script sin necesidad de estar cerrando excel todo el tiempo. solo funciona en windows.
try:
	subprocess.call(["taskkill", "/f", "/im", "EXCEL.EXE"])
except:
	pass

sys.stdout = open(sys.stdout.fileno(), mode='w', encoding='utf8', buffering=1)

#SACAR TIME SLEEP
org_sleep = time.sleep # save the original #time.sleep

sleep =  False # set sleep to false when doesn’t want to sleep
saltarPrimeriaLienaListaUsuarios = False
subirDrive = True # Subir los documentos a drive
generarjson = False #Generar json de preguntas de este curso


if sleep != True: # if sleep is false
    # override #time.sleep to an empty function
    time.sleep = lambda x: None 

if(clasesyvariables.doSilent):
    time.sleep = lambda x: None 

    def print(*args):
        pass

#Clases
class usuario:
    def __init__(self):
        self.nombre ="usuario"
        self.username = "username"
        self.RUT = "0-X"
        self.nverificador = "X"
        self.email = "@cmm.com"
        self.ApellidoP = "Paterno"
        self.ApellidoM = "Materno"
        self.comuna = ""
        self.establecimiento =  ""
        self.telefono = ""
        self.RBD = ""
        self.nacimiento = ""
        self.ultimaconexion = None
        self.preguntas = []
        self.totalContestadas = 1
        self.totalConTaller = 0
        self.totalBuenas = 0
        self.totalMalas = 0
        self.totalOmitidas = 0
        self.totalPreguntasPBlock = 0
        self.totalPreguntasVF = 0
        self.totalPreguntasFreeResp = 0
        self.totalPreguntasDiaQues = 0
        self.totalPreguntasDragDrop = 0
        self.notaControles = []
        self.notaPreguntasCalificadas = []
        self.notaPretest = 1
        self.notaPostest = 1
        self.totalPreguntasPorTaller = {}
        self.totalPreguntasBuenasOMalasPorTaller ={}
        self.totalPreguntasCorrectasPorTaller = {}
        self.totalPreguntasBuenasEnviadasTaller = {}
        self.totalpretest = 0
        self.totalpostest = 0


class pregunta:
    def __init__(self):
        self.pagina = ""
        self.paginaNumero = 0
        self.curso = ""
        self.tallerNombre = ""
        self.tallerNumero = ""
        self.actividadNombre = ""
        self.actividadNumero = 0
        self.nombrePregunta = 0
        self.numeroPregunta = 0
        self.blockKeyCompleto =""
        self.blockKey =""
        self.codigo = ""
        self.numero = 0
        self.tipo = ""
        self.pretest = False
        self.postest = False
        self.esDeControl = False
        self.esDeEncuesta = False
        self.numeroEncuesta = 0
        self.preguntaEvaluada = False
        self.esunapreguntade = ""
        self.esCorrecta = None
        self.completa = True
        self.intentos = 0
        self.multipleRespuesta = False
        self.respuesta = ""
        self.respuestas = []
        self.respuestaCorrecta = ""
        self.idrespuesta = ""
        self.fechaRespuesta = None
        self.score = 0
        self.tipoOra = False
        self.oraCode = ""
        self.numerocontrol = 0
        self.consentimiento = False
        self.reglamento = False
        self.deTaller = False
        self.tieneBlockkey = False
        #Necesarias cuando el blockkey esta repetido
        self.cantidaddeveces = 1
        self.ncorrectas = 0
        self.nincorrectas =  0 

class subpreguntaencuesta:
    def __init__(self):
        self.usuariorut = "0-0"
        self.nombre = "Nombre Apellido"
        self.nombreusuario = "username"
        self.nombreencuesta = ""
        self.numeroencuesta = 0
        self.numeroidentificatorio1 = 0
        self.numeroidentificatorio2 = 0
        self.nombrepregunta  = ""
        self.tipo = ""
        self.respuesta = ""
        self.arrayres = ""
        self.pagina = 0
        self.blockkey = ""

class ora:
    def __init__(self):
        self.submissioid = ""
        self.itemid = ""
        self.studentid = ""
        self.fecharespuesta = None
        self.fechacalificacion = None
        self.respuesta = ""
        self.score = 0
        self.calificada = False

#Variables
usuarioPrueba = usuario()
totalPreguntas = 0
totalPreguntasPBlock = 0
totalPreguntasVF = 0
totalPreguntasFreeResp = 0
totalPreguntasDiaQues = 0
totalPreguntasDragDrop = 0
totalpreguntasportaller = {}
totalPreguntasPorControl = {}
totalPreguntasEvaluadas = {}
totalPregutnasEncuestas = {}
totalPreguntasBuenasOMalas = {}
totalEncuestas = {}
cantidadControles = 0
cantidadEncuestas = 0
cantidadTipoPreguntaEvaluada = 0
Limpiar = False
Resetear = True

def mergeBackwards(sheet, firstCol, lastCol, row, st):
	borderCols = Border(bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
	if sheet.cell(row=row,column = firstCol).value == sheet.cell(row=row, column = lastCol).value:
		mergeBackwards(sheet, firstCol-1, lastCol, row, 1)
	else:
		if st == 1:
			sheet.merge_cells(start_row = row, start_column = firstCol+1, end_row = row, end_column= lastCol)
			for i in list(range(firstCol+1,lastCol,1)):
				sheet.cell(row=row,column=i).border = borderCols
			sheet.cell(row=row,column=firstCol+1).border = Border(left=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
			sheet.cell(row=row,column=lastCol).border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'))
			
			if row==2:
				sheet.cell(row=row,column=firstCol+1).alignment = Alignment(horizontal='left')
			else:
				sheet.cell(row=row,column=firstCol+1).alignment = Alignment(horizontal='center')
		sheet.cell(row=row,column=lastCol).border = Border(bottom=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'))

def find_between( s, first, last ):
    try:
        start = s.index( first ) + len( first )
        end = s.index( last, start )
        return s[start:end]
    except ValueError:
        return ""

def find_between_r( s, first, last ):
    try:
        start = s.rindex( first ) + len( first )
        end = s.rindex( last, start )
        return s[start:end]
    except ValueError:
        return ""

def find_id(s, last):
    try:
        print("Buscar")
        start = s.rindex( last ) + len( last )
        end = len(s)
        print ("id " + s)
        return s[start:end]
    except ValueError:
        return "ERROR"

def quitarAcentos(s):
    print("string con acentos " + s )
    replacements = (
        ("á", "a"),
        ("é", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ú", "u"),
    )
    for a, b in replacements:
        s = s.replace(a, b).replace(a.upper(), b.upper())
    print ("string sin acentos " + s) 
    return s

def createJson(listaUsuarios,nombre):
    strings_tojson = []
    string_tojson = "[" 
    substring = ""
    substring2 = ""
    substring3 = ""

    strings_tojson.append("[")

    for usr2 in listaUsuarios:
        print ("\n____-------"+usr2.nombre + " numero de preguntas " + str(len(usr2.preguntas)) +"-------____")
        if(usr2.ultimaconexion == None ):
            fechaultimaco = ""
        else:
            fechaultimaco = str(usr2.ultimaconexion)

        ''' '","avance":['+json.dumps(usr2.totalPreguntasPorTaller) +']' '''#Mostrado de avance
        substring = '{"nombrecompleto":"'+ usr2.nombre + '","rut":"'+ usr2.RUT + '","username":"'+ usr2.username  + '","correo":"'+ usr2.email + '","ultimaconexion":"'+ fechaultimaco +'","preguntas":['
        for pre in usr2.preguntas:
            print("añadir pregunta ")

            if (pre.fechaRespuesta == None):
                feharespuesta = ""
            else:
                feharespuesta = str(pre.fechaRespuesta)

            escribirCorrecta = str(pre.esCorrecta).lower()
            if( escribirCorrecta == "none" ):
                escribirCorrecta = "null"


            respuesta = pre.respuesta.replace('"',"'")
            respuesta = respuesta.replace(",",".")
            respuesta = respuesta.rstrip()
            respuesta = respuesta.replace("\\","/")
            if not (respuesta.isprintable()):
                respuesta = "Contiene caracteres no imprimibles"

            respcorrecta = pre.respuestaCorrecta.replace('"',"'")
            respcorrecta = respcorrecta.replace(",",".")
            respcorrecta = respcorrecta.rstrip()
            respcorrecta = respcorrecta.replace("\\","/")
            if not (respcorrecta.isprintable()):
                respcorrecta = "Contiene caracteres no imprimibles"

            pre_substring2= '{"pagina":"'+pre.pagina + '","paginanumero":'+ str(pre.paginaNumero) + ',"tallernombre":"'+ pre.tallerNombre + '","tallernumero":'+ str(pre.tallerNumero) +',"actividadnombre":"'+ pre.actividadNombre + '","actividadnumero":'+ str(pre.actividadNumero) +',"blockkeycompleto":"'+pre.blockKeyCompleto +'","blockkey":"'+pre.blockKey +'","codigo":"'+pre.codigo + '","tipo":"'+ pre.tipo + '","preguntacalificada":'+ str(pre.preguntaEvaluada).lower() + ',"pretest":'+ str(pre.pretest).lower() + ',"postest":'+ str(pre.postest).lower() + ',"Es de control":'+ str(pre.esDeControl).lower() + ',"numerocontrol":'+ str(pre.numerocontrol) + ',"Es de encuesta":'+ str(pre.esDeEncuesta).lower() + ',"consentimiento":' + str(pre.consentimiento).lower()+ ',"reglamento":' + str(pre.reglamento).lower() + ',"detaller":'+ str(pre.deTaller).lower()+ ',"tieneblockkey":'+ str(pre.tieneBlockkey).lower() + ',"intentos":' + str(pre.intentos) +',"idrespuesta":"'+ pre.idrespuesta+'","correcta":'+ escribirCorrecta +',"tipoora":' + str(pre.tipoOra).lower() + ',"oraCode":"'+  pre.oraCode + '","score":'+ str(pre.score) +',"fecharespuesta":"'+ feharespuesta + '","multiplerespuesta":'+ str(pre.multipleRespuesta).lower() + ',"respuestacorrecta":"'+ respcorrecta +'","respuesta":"'+respuesta+'","respuestas":['
            substring3 = ""
            for res in pre.respuestas :
                pre_substring3 = '"' + res.replace('"',"'") + '"'
                pre_substring3 = pre_substring3.replace("\\","/")
                pre_substring3 = pre_substring3.replace("\n","")
                if not (pre_substring3.isprintable()):
                    pre_substring3 = '"Contiene caracteres no imprimibles"'
                substring3 += pre_substring3 + ","
                print("res: " + res + " __substring 3: " + substring3)
                ##time.sleep(3)
            if(len(substring3) > 1 and substring3[-1] == ","): substring3 = substring3[:-1] 
            pre_substring2 += substring3 +"]},"
            substring2 += pre_substring2

        if(len(substring2) > 1 and substring2[-1] == ","): substring2 = substring2[:-1]    
        substring += substring2 + "]}," 
        strings_tojson.append(substring)
        string_tojson += substring
        substring2 = ""
        ##time.sleep(3)
            
    string_tojson = string_tojson[:-1] +']'
    strings_tojson[-1] = strings_tojson[-1][:-1] + ']'
            
            
    print("\n")
    print("caracteres en el json: " + str(len(string_tojson)))
    #print (string_tojson)

    #jsondir = os.path.join(sys.path[0], "seguimiento.json")
    mame = "Seguimiento_ " + nombre + ".json"
    jsondir = os.path.join(sys.path[0], mame)
    print ("jsondir " + jsondir)
    text_file = open(jsondir, "w+", encoding='utf-8')
    
    for jsontext in strings_tojson:
        #print (jsontext)
        n = text_file.write(jsontext)
        ##time.sleep(2)

    print ("Guardar Json")
    text_file.close()

    return jsondir


def CreateErrorFile(listaErrores,nombre):
    print("Es hora deeee Anomaliacs ")

    nombreAnomalo = "Anomalias_" + nombre + ".txt"    
    logError_dir = os.path.join(sys.path[0], nombreAnomalo )
    if os.path.isfile(logError_dir):
        os.remove(logError_dir)
    print ("Anomalias " + logError_dir)
    text_file2 = open(logError_dir, "w+", encoding='utf-8')

    for anomalia in listaErrores:
        n = text_file2.write(anomalia + '\n' )

    print ("Guardar Txt")
    text_file2.close()

    return logError_dir




def LoadListaNega(ArchivoXLS):
    print ("Cargar lista negra")
    if os.path.isfile(ArchivoXLS):
        wb_listaNegra = openpyxl.load_workbook(ArchivoXLS) #Abrir lista negra
        
        sheet =  wb_listaNegra.active
        iterator = 0
        for row in sheet.iter_rows('A{}:A{}'.format(sheet.min_row,sheet.max_row)):
            for cell in row:
                iterator += 1
                correo = sheet['B'+str(iterator)].value
                #print ("entoncre al usuario " + str(cell.value) + " " + str(correo) + " en lista negra")
                for usr in usuarios:
                    if(str(cell.value).find(usr.username) != -1):
                        print("Borarre al usuario " + usr.username + " por USERNAME")
                        usrtoDelete = usr
                        usuarios.remove(usrtoDelete)
                        ##time.sleep(6)
                    elif(str(correo).find(usr.email) != -1):
                        print("Borarre al usuario " + usr.username + " por EMAIL")
                        usrtoDelete = usr
                        usuarios.remove(usrtoDelete)
                        ##time.sleep(6)


def LoadDatosUsuario(ArchivoXLS):
    print ("Cargar lista de usuarios")
    if os.path.isfile(ArchivoXLS):
        wb_listaUsuarios = openpyxl.load_workbook(ArchivoXLS) #Abrir datos de los usuarios

        sheet = wb_listaUsuarios.active
        iterator = 0
        for row in sheet.iter_rows('F{}:F{}'.format(sheet.min_row,sheet.max_row)):
            for cell in row:
                iterator += 1
                print ("buscar " + str(cell.value) + " en la lista")
                for usr in usuarios:
                    print(str(cell.value) + " Vs. " + usr.email)
                    if(str(cell.value).find(usr.email) != -1):
                        print("lo encontre en "+ str(row) + " - " + str(cell))
                        usr.RUT = sheet["B" + str(iterator)].value
                        usr.ApellidoP = sheet["C" + str(iterator)].value
                        usr.ApellidoM = sheet["D" + str(iterator)].value
                        usr.nombre =  sheet["E" + str(iterator)].value
                        usr.comuna = sheet["G" + str(iterator)].value
                        usr.telefono = sheet["H" + str(iterator)].value 
                        usr.RBD = sheet["I" + str(iterator)].value
                        usr.establecimiento = sheet["J" + str(iterator)].value
                        
        ##time.sleep(12)

    else:
        print('No existe archivo de usuarios CPEIP para este curso.')
        logErrores.append('No existe archivo de usuarios CPEIP para este curso.')


def quitarUsuarioPruebadelaLista(nombreUsrPrueba):
    print("Buscar al ususario de prueba para sacarlo "  + str(nombreUsrPrueba))
    ##time.sleep(5)
    for usrprueba in usuarios:
        print (str(usrprueba.nombre) + " VS " + nombreUsrPrueba)
        if(usrprueba.username.find(nombreUsrPrueba) != -1):
            print("Borarre al usuario de prueba: " + usrprueba.nombre + " por USERNAME")
            usrtoDelete = usrprueba
            usuarios.remove(usrtoDelete)
           # #time.sleep(12)

def creteXLS(listaUsuarios,nombre):
    global cantidadControles
    print("Crear xls")

    cantidadTalleres = 0
    #ReportePath = '/var/www/html/seguimiento/ReporteDescargado.xlsx'
    ReportePath = 'ReporteDescargado.xlsx'
    bookExists = os.path.isfile(ReportePath)
    if bookExists and Resetear == False:
        print("Encontre el libro. lo abrire " + ReportePath )
        #time.sleep(4)
        book = openpyxl.load_workbook(ReportePath) 
    else:
        print("Creare el libro")
        #time.sleep(4)
        book = openpyxl.Workbook()

    ##time.sleep(12)
    #Definición de bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    rightbottom_border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
    right_border = Border(right = Side(style = 'thin'))  
    bottom_border = Border(bottom = Side(style = 'thin'))

    ## Hay que checkear si existen previamente

    sheetCompleto = book["Reporte Completo"] if ('Reporte Completo' in book.sheetnames) else book.create_sheet("Reporte Completo",0)
    sheetControl = book["Reporte Controles"] if ('Reporte Controles' in book.sheetnames) else book.create_sheet("Reporte Controles",1)
    sheetRepAvance = book["Reporte Avance"] if ('Reporte Avance' in book.sheetnames) else book.create_sheet("Reporte Avance",2)
    if(cantidadTipoPreguntaEvaluada  > 0):
        sheetPreguntas = book["Reporte Preguntas Calificadas"] if ('Reporte Preguntas Calificadas' in book.sheetnames) else book.create_sheet("Reporte Preguntas Calificadas",3)
    sheetTest = book["Reporte Tests"] if ('Reporte Tests' in book.sheetnames) else book.create_sheet("Reporte Tests",2)
    sheetAvance = book["Datos Avance"]  if ('Datos Avance' in book.sheetnames) else book.create_sheet("Datos Avance",4)
    sheetFechas = book["Datos Fechas Avance"] if ('Datos Fechas Avance' in book.sheetnames) else book.create_sheet("Datos Fechas Avance",5)
    sheetDatos = book["Datos Info Usuario"] if ('Datos Info Usuario' in book.sheetnames) else book.create_sheet("Datos Info Usuario",6)

    try:
        book.remove_sheet(book['Sheet']) #Hoja fantasma
    except:
        pass

    sheetCompleto.sheet_properties.tabColor="1072BA"  #Pintar pestaña azul
    sheetControl.sheet_properties.tabColor="38761D"  #Pintar pestaña verde
    sheetRepAvance.sheet_properties.tabColor="38761D"  #Pintar pestaña verde
    sheetTest.sheet_properties.tabColor="38761D"  #Pintar pestaña verde
    if(cantidadTipoPreguntaEvaluada > 0):
        sheetPreguntas.sheet_properties.tabColor="38761D"
    sheetAvance.sheet_properties.tabColor="980000"  #Pintar pestaña roja
    sheetFechas.sheet_properties.tabColor="980000"	#Pintar pestaña roja
    sheetDatos.sheet_properties.tabColor="980000"	#Pintar pestaña roja
    
    #Nominas

    sheetAvance["A1"] = "Nómina"
    sheetAvance["B4"].border = thin_border
    sheetAvance["C4"] = "Apellido Paterno"
    sheetAvance["C4"].border = thin_border
    sheetAvance["D4"] = "Apellido Materno"
    sheetAvance["D4"].border = thin_border
    sheetAvance["E4"] = "Nombre"
    sheetAvance["E4"].border = thin_border
    sheetAvance["F4"] = "Usuario edX"
    sheetAvance["F4"].border = thin_border
    sheetAvance["A4"] = "N°"
    sheetAvance["A4"].border = thin_border
    sheetAvance["B4"] = "RUT"
    sheetAvance["G1"] = "Taller"
    sheetAvance["G1"].border = thin_border
    sheetAvance["G2"] = "Actividad"
    sheetAvance["G2"].border = thin_border
    sheetAvance["G3"] = "Pagina"
    sheetAvance["G3"].border = thin_border
    sheetAvance["G4"] = "ID Bloque"
    sheetAvance["G4"].border = thin_border
    sheetAvance.merge_cells('A1:F3')
    sheetAvance["A1"].border = thin_border
    
    sheetFechas["A1"] = "Nómina"
    sheetFechas["A4"] = "N°"
    sheetFechas["A4"].border = thin_border
    sheetFechas["B4"] = "RUT"
    sheetFechas["B4"].border = thin_border
    sheetFechas["C4"] = "Apellido Paterno"
    sheetFechas["C4"].border = thin_border
    sheetFechas["D4"] = "Apellido Materno"
    sheetFechas["D4"].border = thin_border
    sheetFechas["E4"] = "Nombre"
    sheetFechas["E4"].border = thin_border
    sheetFechas["F4"] = "Usuario edX"
    sheetFechas["F4"].border = thin_border
    sheetFechas["G1"] = "Taller"
    sheetFechas["G1"].border = thin_border
    sheetFechas["G2"] = "Actividad"
    sheetFechas["G2"].border = thin_border
    sheetFechas["G3"] = "Pagina"
    sheetFechas["G3"].border = thin_border
    sheetFechas["G4"] = "ID Bloque"
    sheetFechas["G4"].border = thin_border
    sheetFechas.merge_cells('A1:F3')
    sheetFechas["A1"].border = thin_border
    
    sheetRepAvance["A1"] = "Nómina"
    sheetRepAvance["A1"].border = thin_border
    sheetRepAvance["A2"] = "N°"
    sheetRepAvance["A2"].border = thin_border
    sheetRepAvance["B2"] = "RUT"
    sheetRepAvance["B2"].border = thin_border
    sheetRepAvance["C2"] = "Apellido Paterno"
    sheetRepAvance["C2"].border = thin_border
    sheetRepAvance["D2"] = "Apellido Materno"
    sheetRepAvance["D2"].border = thin_border
    sheetRepAvance["G1"] = "Taller"
    sheetRepAvance["G1"].border = thin_border
    sheetRepAvance["E2"] = "Nombre"
    sheetRepAvance["E2"].border = thin_border
    sheetRepAvance["F2"] = "Usuario edX"
    sheetRepAvance["F2"].border = thin_border
    sheetRepAvance.merge_cells('A1:F1')

    sheetCompleto["A1"] = "Nómina"
    sheetCompleto["A1"].border = thin_border
    sheetCompleto["A2"] = "N°"
    sheetCompleto["A2"].border = thin_border
    sheetCompleto["B2"] = "RUT"
    sheetCompleto["B2"].border = thin_border
    sheetCompleto["C2"] = "Apellido Paterno"
    sheetCompleto["C2"].border = thin_border
    sheetCompleto["D2"] = "Apellido Materno"
    sheetCompleto["D2"].border = thin_border
    sheetCompleto["E2"] = "Nombre"
    sheetCompleto["E2"].border = thin_border
    sheetCompleto["F2"] = "Usuario edX"
    sheetCompleto["F2"].border = thin_border
    sheetCompleto["G2"] = "Última Conexión"
    sheetCompleto["G2"].border = thin_border
    sheetCompleto.merge_cells('A1:G1')

    sheetControl["A1"] = "Nómina"
    sheetControl["A1"].border = thin_border
    sheetControl["A2"] = "N°"
    sheetControl["A2"].border = thin_border
    sheetControl["B2"] = "RUT"
    sheetControl["B2"].border = thin_border
    sheetControl["C2"] = "Apellido Paterno"
    sheetControl["C2"].border = thin_border
    sheetControl["D2"] = "Apellido Materno"
    sheetControl["D2"].border = thin_border
    sheetControl["E2"] = "Nombre"
    sheetControl["E2"].border = thin_border
    sheetControl["F2"] = "Usuario edX"
    sheetControl["F2"].border = thin_border
    sheetControl.merge_cells('A1:F1')

    if(cantidadTipoPreguntaEvaluada > 0):
        sheetPreguntas["A1"] = "Nómina"
        sheetPreguntas["A1"].border = thin_border
        sheetPreguntas["A3"] = "N°"
        sheetPreguntas["A3"].border = thin_border
        sheetPreguntas["B3"] = "RUT"
        sheetPreguntas["B3"].border = thin_border
        sheetPreguntas["C3"] = "Apellido Paterno"
        sheetPreguntas["C3"].border = thin_border
        sheetPreguntas["D3"] = "Apellido Materno"
        sheetPreguntas["D3"].border = thin_border
        sheetPreguntas["E3"] = "Nombre"
        sheetPreguntas["E3"].border = thin_border
        sheetPreguntas["F3"] = "Usuario edX"
        sheetPreguntas["F3"].border = thin_border
        sheetPreguntas.merge_cells('A1:F2')

    sheetTest["A1"] = "Nómina"
    sheetTest["A2"] = "N°"
    sheetTest["A2"].border = thin_border
    sheetTest["B2"] = "RUT"
    sheetTest["B2"].border = thin_border
    sheetTest["C2"] = "Ap. Paterno"
    sheetTest["C2"].border = thin_border
    sheetTest["D2"] = "Ap. Materno"
    sheetTest["D2"].border = thin_border
    sheetTest["E2"] = "Nombre"
    sheetTest["E2"].border = thin_border
    sheetTest["F2"] = "Nombre usuario (edX)"
    sheetTest["F2"].border = thin_border
    sheetTest.merge_cells('A1:F1')

    sheetDatos["A1"] = "N°"
    sheetDatos["A1"].border = thin_border
    sheetDatos["B1"] = "RUT"
    sheetDatos["B1"].border = thin_border
    sheetDatos["C1"] = "Ap. Paterno"
    sheetDatos["C1"].border = thin_border
    sheetDatos["D1"] = "Ap. Materno"
    sheetDatos["D1"].border = thin_border
    sheetDatos["E1"] = "Nombre"
    sheetDatos["E1"].border = thin_border
    sheetDatos["F1"] = "Nombre usuario (edX)"
    sheetDatos["F1"].border = thin_border
    sheetDatos["G1"] = "Mail"
    sheetDatos["G1"].border = thin_border
    sheetDatos["H1"] = "Teléfono"
    sheetDatos["H1"].border = thin_border
    sheetDatos["I1"] = "RBD"
    sheetDatos["I1"].border = thin_border
    sheetDatos["J1"] = "Establecimiento"
    sheetDatos["J1"].border = thin_border
    sheetDatos["K1"] = "Ultima Conexión"
    sheetDatos["K1"].border = thin_border

    
    #Colores

    colorFijo = '999999'
    sheetAvance["A1"].fill = PatternFill(fgColor=colorFijo, patternType="solid")
    sheetAvance["A1"].alignment = Alignment(horizontal='center',vertical='center')
    sheetAvance["A1"].font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
    for row_range in range(4, 5+len(listaUsuarios)):
    	sheetAvance["A"+str(row_range)].fill = PatternFill(fgColor=colorFijo, patternType="solid")
    	sheetAvance["A"+str(row_range)].alignment = Alignment(horizontal='center',vertical='center')
    	sheetAvance["A"+str(row_range)].font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
    	sheetAvance["A"+str(row_range)].border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))
    for row_range in range(1, 5):
    	sheetAvance["G"+str(row_range)].fill = PatternFill(fgColor=colorFijo, patternType="solid")
    	sheetAvance["G"+str(row_range)].alignment = Alignment(horizontal='center',vertical='center')
    	sheetAvance["G"+str(row_range)].font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
    	sheetAvance["G"+str(row_range)].border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))
	
    for column_range in range(2,7):
    	sheetAvance.cell(row=4,column=column_range).fill = PatternFill(fgColor=colorFijo, patternType="solid")
    	sheetAvance.cell(row=4,column=column_range).alignment = Alignment(horizontal='center',vertical='center')
    	sheetAvance.cell(row=4,column=column_range).font = Font(bold=True,color='FFFFFF',vertAlign='baseline')
    	sheetAvance.cell(row=4,column=column_range).border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))

    for column_range in range(2,8):
    	for row_range in range(4, 5+len(listaUsuarios)):
	    	sheetAvance.cell(row=row_range,column=column_range).alignment = Alignment(horizontal='left',vertical='center')
	    	sheetAvance.cell(row=row_range,column=column_range).border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))

    sheetAvance.merge_cells("G5:G"+str(len(listaUsuarios)+4))
    for row_range in range(4, 5+len(listaUsuarios)):
    	sheetAvance["G"+str(row_range)].fill = PatternFill(fgColor=colorFijo, patternType="solid")
    	sheetAvance["G"+str(row_range)].border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),left=Side(border_style='thin',color='000000'))


    #Llenar columnas
    columnas = []
    for char in string.ascii_uppercase:
        columnas.append(char)
    for char1 in string.ascii_uppercase:
        for char2 in string.ascii_uppercase:
            columnas.append(char1 + char2)
    for char01 in string.ascii_uppercase:
        for char02 in string.ascii_uppercase:
            for char03 in string.ascii_uppercase:
                columnas.append(char01 + char02 + char03)
    
    iteratorColumnPreguntas = 7
    talllerTexto= ""
    print(len(usuarioPrueba.preguntas))
    fontNegrita = Font(bold=True)
    colors1 = ['F7CAAC','FBE4D5']
    colors2 = ['BDD6EE','DEEAF6']
    pats = [cycle(colors1),cycle(colors2)]
    currPat = 0
    currentColor = next(pats[currPat])
    dontChange = 1
    #beforeLength = len(usuarioPrueba.preguntas)
    print("usuario prueba.preguntas =" + str(usuarioPrueba.preguntas))
    usuarioDummy=usuarioPrueba.preguntas[len(usuarioPrueba.preguntas)-1]
    usuarioDummy.tallerNombre = ""
    usuarioDummy.actividadNombre = ""
    usuarioDummy.esDeControl = False
    #usuarioPrueba.preguntas.append(usuarioDummy)
    beforeLength = len(usuarioPrueba.preguntas)
    print("antes->"+str(beforeLength)+" despues->"+str(len(usuarioPrueba.preguntas)))

    for pregunta in usuarioPrueba.preguntas: #Busqueda de preguntas en talleres   
        if (pregunta.esDeControl == False and pregunta.preguntaEvaluada == False and pregunta.esDeEncuesta == False and pregunta.tallerNumero > 0):
            if(pregunta.tallerNombre != talllerTexto): 
                talllerTexto = pregunta.tallerNombre
                cantidadTalleres += 1
                sheetAvance[columnas[iteratorColumnPreguntas]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
                sheetFechas[columnas[iteratorColumnPreguntas]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
                print("taller " + str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre)
                ##time.sleep(3)
                
            sheetAvance[columnas[iteratorColumnPreguntas]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre    
            sheetAvance[columnas[iteratorColumnPreguntas]+'2']= pregunta.actividadNombre
            sheetAvance[columnas[iteratorColumnPreguntas]+'3']= pregunta.paginaNumero
            sheetAvance[columnas[iteratorColumnPreguntas]+'4']= pregunta.codigo
           
            if pregunta.tallerNombre == "":
            	sheetAvance[columnas[iteratorColumnPreguntas]+'1'] = ""
            	sheetAvance[columnas[iteratorColumnPreguntas]+'2'] = ""
            	sheetAvance[columnas[iteratorColumnPreguntas]+'3'] = ""
            	sheetAvance[columnas[iteratorColumnPreguntas]+'4'] = ""


            # Fuente bold para el encabezado
            sheetAvance[columnas[iteratorColumnPreguntas]+'1'].font = fontNegrita
            sheetAvance[columnas[iteratorColumnPreguntas]+'2'].font = fontNegrita
            sheetAvance[columnas[iteratorColumnPreguntas]+'3'].font = fontNegrita
            sheetAvance[columnas[iteratorColumnPreguntas]+'4'].font = fontNegrita
            sheetAvance.column_dimensions[columnas[iteratorColumnPreguntas]].width = 3
            
            if iteratorColumnPreguntas > 7:
            	shiftIterator = 6
            	#Taller
            	if sheetAvance.cell(row=1,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=1, column = iteratorColumnPreguntas).value or iteratorColumnPreguntas+shiftIterator == beforeLength:
            		print("Cuando entre a talleres->"+str(iteratorColumnPreguntas)+"| 1a :"+str(sheetAvance.cell(row=2,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=2, column = iteratorColumnPreguntas).value) + " 2a : "+str(iteratorColumnPreguntas+shiftIterator == beforeLength))
            		if iteratorColumnPreguntas > 8:
            			mergeBackwards(sheetAvance,iteratorColumnPreguntas-2,iteratorColumnPreguntas-1,1,0)
            			currPat = 1 - currPat
            			currentColor = next(pats[currPat])
            			if (currentColor == colors1[1] or currentColor == colors2[1]):
            				currentColor = next(pats[currPat])
            			dontChange = 1
            	
            	#Actividad
            	if sheetAvance.cell(row=2,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=2, column = iteratorColumnPreguntas).value or iteratorColumnPreguntas+shiftIterator == beforeLength:
            		print("Cuando entre a actividades->"+str(iteratorColumnPreguntas)+"| 1a :"+str(sheetAvance.cell(row=2,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=2, column = iteratorColumnPreguntas).value) + " 2a : "+str(iteratorColumnPreguntas+shiftIterator == beforeLength))
            		mergeBackwards(sheetAvance,iteratorColumnPreguntas-2,iteratorColumnPreguntas-1,2,0)

            		if dontChange != 1:
            			currentColor = next(pats[currPat])
            		else:
            			dontChange = 0
            	
            	#Pagina	
            	if sheetAvance.cell(row=3,column = iteratorColumnPreguntas-1).value != sheetAvance.cell(row=3, column = iteratorColumnPreguntas).value or iteratorColumnPreguntas+shiftIterator == beforeLength:
            		mergeBackwards(sheetAvance,iteratorColumnPreguntas-2,iteratorColumnPreguntas-1,3,0)

            	for row_range in range(1, 5+len(listaUsuarios)):
            	    cell_title = sheetAvance.cell(row_range,iteratorColumnPreguntas)            	    
            	    cell_title.border = Border(right=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))
            	    cell_title.fill = PatternFill(fgColor=currentColor, patternType="solid")
           	# Hasta aca estilo

            sheetFechas[columnas[iteratorColumnPreguntas]+'2']= pregunta.actividadNombre
            sheetFechas[columnas[iteratorColumnPreguntas]+'3']= pregunta.paginaNumero
            sheetFechas[columnas[iteratorColumnPreguntas]+'4']= pregunta.codigo

            codigotocheck = pregunta.codigo

            iteratorUser = 5
            iteratorUserFechas = 5
            numeroUser = 1
            for usr in listaUsuarios:
                #iteratorColumn2 = 1
                sheetAvance["A" + str(iteratorUser)] = numeroUser
                sheetAvance["B" + str(iteratorUser)] = usr.RUT
                sheetAvance["C" + str(iteratorUser)] = usr.ApellidoP
                sheetAvance["D" + str(iteratorUser)] = usr.ApellidoM
                sheetAvance["E" + str(iteratorUser)] = usr.nombre
                sheetAvance["F" + str(iteratorUser)] = usr.username
                sheetAvance["F" + str(iteratorUser)].border = right_border
                #sheet["F" + str(iteratorUser)].border = right_border

                
                sheetFechas["A" + str(iteratorUser)] = numeroUser
                sheetFechas["B" + str(iteratorUser)] = usr.RUT
                sheetFechas["C" + str(iteratorUser)] = usr.ApellidoP
                sheetFechas["D" + str(iteratorUser)] = usr.ApellidoM
                sheetFechas["E" + str(iteratorUser)] = usr.nombre
                sheetFechas["F" + str(iteratorUser)] = usr.username
                sheetFechas["F" + str(iteratorUser)].border = right_border

                
                for pregunta in usr.preguntas:
                    if(pregunta.codigo == codigotocheck):
                        #sheet[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = "Listo "+ usr.nombre
                        if(pregunta.esCorrecta == True):
                            sheetAvance[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = 1
                        elif(pregunta.esCorrecta == False):     
                            sheetAvance[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = 0
                        else:
                            sheetAvance[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = 2
                        if (pregunta.fechaRespuesta != None):
                            sheetFechas[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = pregunta.fechaRespuesta
                        else:
                            sheetFechas[columnas[iteratorColumnPreguntas]+str(iteratorUser)] = "No hay fecha"

                iteratorUserFechas += 1     
                iteratorUser += 1
                numeroUser += 1

            iteratorColumnPreguntas += 1
    usuarioPrueba.preguntas.pop()   

    ##-----------------##Avance#-----------------## 

    escribirAvance = 6
    i = 0
    sheetRepAvance[columnas[escribirAvance]+"1"] = "Nº pregs:"
    
    talleres = usuarioPrueba.totalPreguntasPorTaller

    talleres = list(usuarioPrueba.totalPreguntasPorTaller)
    print ("pestaña de avance, Talleres: " +  str(talleres))

    espaciosAvanzadosAvance = 0
    #los diccionarios no se pueden ordenar :( asi que hay que hacer este fix
    losTalleresOrdenados = list(totalPreguntasPorTaller.keys())
    losTalleresOrdenados.sort()
    for taller in losTalleresOrdenados:
        sheetRepAvance[columnas[escribirAvance + i ]+"2"] = "T"+str(i + 1)
        print("Escribir taller en pestañaa de avance [" + columnas[escribirAvance + i ] +"2]" )

        iteratorUser = 3
        for usr in usuarios:

            sheetRepAvance["A" + str(iteratorUser)] = numeroUser
            sheetRepAvance["B" + str(iteratorUser)] = usr.RUT
            sheetRepAvance["C" + str(iteratorUser)] = usr.ApellidoP
            sheetRepAvance["D" + str(iteratorUser)] = usr.ApellidoM
            sheetRepAvance["E" + str(iteratorUser)] = usr.nombre
            sheetRepAvance["F" + str(iteratorUser)] = usr.username
            sheetRepAvance["F" + str(iteratorUser)].border = right_border


            sumapreguntas = 0
            #utototalpreguntas = []
            print("total preguntas por taller del usuario " + usr.nombre + ": " + str(usr.totalPreguntasPorTaller))
            print("voy en taller "+taller)
            if(len(usr.totalPreguntasPorTaller) > 0 and int(taller) in usr.totalPreguntasPorTaller):
                print ( "Escribir "+ usr.username +" en TALLER " + str(taller) + ": " + str(usr.totalPreguntasPorTaller[int(taller)]) )
                for act in usr.totalPreguntasPorTaller[int(taller)]:
                    sumapreguntas += usr.totalPreguntasPorTaller[int(taller)][act]
                print("Total preguntas en taller " + str(taller) + ": " + str(sumapreguntas))
                ##time.sleep(2)
            sheetRepAvance[columnas[escribirAvance + i ]+ str(iteratorUser)]  = sumapreguntas
            iteratorUser += 1
            
        ##time.sleep(6)
        i+=1
        espaciosAvanzadosAvance += 1

    ## FIX TOTAL PREGUNTAS ##
    total_auxiliar = 0
    for i in [i for i in usuarioPrueba.totalPreguntasPorTaller.values()]:
    	for d in i:
    		if d != '-1':
    			total_auxiliar += i[d]
    usuarioPrueba.totalContestadas = float(total_auxiliar)
    preguntastotales = float(total_auxiliar)
    ## FIX TOTAL PREGUNTAS ##


    sheetRepAvance[columnas[escribirAvance + 1 ]+"1"] = usuarioPrueba.totalContestadas

    escribirAvance += espaciosAvanzadosAvance

    sheetRepAvance[columnas[escribirAvance]+"2"] = "Total curso"

    iteratorUser = 3
    for usr in usuarios:
        sheetRepAvance[columnas[escribirAvance] + str(iteratorUser)] = usr.totalConTaller
        iteratorUser += 1
    
    sheetRepAvance[columnas[escribirAvance +1 ]+"1"] = "Taller"
    sheetRepAvance[columnas[escribirAvance +1 ]+"2"] = "Actvidad"

    sheetRepAvance[columnas[escribirAvance +1 ]+"3"] = "Correctas | Contestadas"
    sheetRepAvance[columnas[escribirAvance +1 ]+"3"].alignment = Alignment(vertical='center')
    sheetRepAvance.merge_cells(columnas[escribirAvance +1] + "3:" + columnas[escribirAvance +1] + str(iteratorUser))



    iteratorColumnAvance = escribirAvance + 2
    talleres.sort()
    #talleresParaRecorrer = []
    
    for taller in talleres:
        #talleres = list(usuarioPrueba.totalPreguntasPorTaller)
        if(int(taller) > 0):
            actividadeslist = list(usuarioPrueba.totalPreguntasPorTaller[taller])
            actividadeslist.sort()
            for actividad in actividadeslist:
                if(int(actividad) > 0):
                    sheetRepAvance[columnas[iteratorColumnAvance]+'1'] = taller
                    sheetRepAvance[columnas[iteratorColumnAvance]+'2'] = actividad
                    #talleresParaRecorrer.append(int(taller))
                    iteratorColumnAvance += 1

    iteratorUser = 3
    for usr in usuarios:
        #aqui debería recorrer el avance
        iteratorColumnAvance2 = escribirAvance + 2
        talleresusr = usr.totalPreguntasPorTaller
        tallerbuenasusr = usr.totalPreguntasBuenasEnviadasTaller

        print("Preguntas totalaes " + usr.username + ": " + str(talleresusr) )
        print("Preguntas totales del usuario de prueba : " + str(usuarioPrueba.totalPreguntasPorTaller) )


        print( "Buenas del usuario  "  + usr.username + ": " + str(tallerbuenasusr) )
        #print( "Preguntas que pueden ser buenas o malas: " + str(totalPreguntasBuenasOMalas) )

        #tallerPosibles = usr.totalPreguntaBuenaOMala
        talleres = list(usuarioPrueba.totalPreguntasPorTaller)
        #print("talleres " + str(talleres))
        talleres.sort()
        #tallerbuenasusr = list(usr.totalPreguntasCorrectasPorTaller)
        #tallerbuenasusr.sort()

        print("talleres ordenado" + str(talleres))
        print("talleres buenas ordenados " + str(tallerbuenasusr) )
        ##time.sleep(12)

        for taller in talleres: #Recorro una vez para asegurarme que el usuario de prueba contesto mas que los demas
            if taller in talleresusr:
                actividadeslist = list(usuarioPrueba.totalPreguntasPorTaller[taller])
                actividadeslist.sort()
                for actividad in actividadeslist:
                    print ("Taller " + str(taller) + " actividad " + str(actividad )) 
                    if(actividad in talleresusr[int(taller)]):
                        
                        conuser = talleresusr[int(taller)][int(actividad)]
                        conprueba = usuarioPrueba.totalPreguntasPorTaller[taller][actividad]
                        
                        print ("Contestadas en T" + str(taller) + "A" + str(actividad) +" por " + usr.username +": " + str(conuser) + " Vs. " + str(conprueba) )
                        if(conuser > conprueba):
                            print ("El usuario " + usr.username + " contesto mas que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            logErrores.append("El usuario " + usr.username + " contesto mas que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            #usuarioPrueba.totalPreguntasPorTaller[taller][actividad] = conuser
                            usr.totalPreguntasPorTaller[int(taller)][int(actividad)] = conprueba
                            time.sleep(4)
                        
                        buuser = 0
                        if(taller in tallerbuenasusr):
                            if(actividad in tallerbuenasusr[int(taller)]):
                                buuser = tallerbuenasusr[int(taller)][int(actividad)]
                                
                        buprueba = 1
                        if(taller in totalPreguntasBuenasOMalas):
                            if(actividad in totalPreguntasBuenasOMalas[taller]):
                                buprueba = totalPreguntasBuenasOMalas[taller][actividad]
                        print ("Evaluables (no PE) en T" + str(taller) + "A" + str(actividad) +" por " + usr.username +": " + str(buuser) + " Vs. " + str(buprueba) )
                        
                        if(buuser > buprueba):
                            print ("El usuario " + usr.username + " contesto mas evaluables (no Preg. Calificadas) que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            logErrores.append("El usuario " + usr.username + " contesto mas evaluables (no Preg. Calificadas) que el usuario de prueba en T" + str(taller) + "A" + str(actividad))
                            usr.totalPreguntasCorrectasPorTaller[taller][actividad] = buprueba
                            #totalPreguntasBuenasOMalas[taller][actividad] = buuser
                            time.sleep(4)

                        
        for taller in talleres: #Recorro una seguda vez ahora para sacar los %
            if (int(taller) in talleresusr) and int(taller) > 0:
                actividadeslist = list(usuarioPrueba.totalPreguntasPorTaller[taller])
                actividadeslist.sort()
                for actividad in actividadeslist:
                    if int(actividad) > 0:
                        if int(actividad) in talleresusr[int(taller)]:
                            contestadas = 0
                            totalContestadas = 1
                            buenas = 0
                            totalBuenas = 1

                            print("talleresusr " + str(talleresusr))
                            if(int(taller) in talleresusr ):
                                if(int(actividad) in talleresusr[int(taller)]):
                                    #print()
                                    contestadas = talleresusr[int(taller)][int(actividad)]

                            print("ppreguntas totales uprueba " + str(usuarioPrueba.totalPreguntasPorTaller))
                            if(taller in usuarioPrueba.totalPreguntasPorTaller):
                                if(actividad in usuarioPrueba.totalPreguntasPorTaller[taller]):
                                    totalContestadas = usuarioPrueba.totalPreguntasPorTaller[taller][actividad]
                            
                            print("talleresusr " + str(tallerbuenasusr))
                            if(int(taller) in tallerbuenasusr):
                                if(int(actividad) in tallerbuenasusr[int(taller)]):
                                    buenas = tallerbuenasusr[int(taller)][int(actividad)]

                            #Por ahora no es necesarios saber las evaluables
                            print("ppreguntas evaluables uprueba " + str(totalPreguntasBuenasOMalas))
                            if(taller in totalPreguntasBuenasOMalas):
                                if(actividad in totalPreguntasBuenasOMalas[taller]):
                                    totalBuenas = usuarioPrueba.totalPreguntasBuenasOMalasPorTaller[taller][actividad]

                            ##time.sleep(6)

                            #print ("% Contestadas en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(contestadas) + "/" + str(totalContestadas) + "= " + str(contestadas/totalContestadas) )
                            #print ("% Buenas  en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(buenas) + "/" + str(totalBuenas) + "= " + str(buenas/totalBuenas))
                            ##time.sleep(5)

                            porContestadas = 0
                            if(totalContestadas >1  ):
                                porContestadas = int((float(contestadas)/float(totalContestadas)) * 100.0)

                            porBuenas = 0
                            if(totalBuenas > 1):
                                porBuenas =  int((float(buenas)/float(totalContestadas))* 100.0)

                            if(porContestadas > 100):
                                print ("[" + str(porContestadas) + "] % Contestadas en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(contestadas) + "/" + str(totalContestadas) + "= " + str(contestadas/totalContestadas) )
                                time.sleep(3)
                                porContestadas = 100
                            
                            if(porBuenas > 100):
                                print ("[" + str(porBuenas) + "] % Buenas  en taller " + str(taller) + " ,actividad " + str(actividad) + ": " + str(buenas) + "/" + str(totalBuenas) + "= " + str(buenas/totalBuenas))
                                time.sleep(3)
                                porBuenas = 100

                            sheetRepAvance[columnas[iteratorColumnAvance2]+str(iteratorUser)].value  = str(porBuenas) + "% | " + str(porContestadas) + "%"

                        iteratorColumnAvance2 += 1
            else:
                #si no existe el taller, me muevo a la derecha segun cuantas actividades hayan 
                iteratorColumnAvance2 += len(usuarioPrueba.totalPreguntasPorTaller[taller])
        iteratorUser += 1


    ##-----------------##Tests##-----------------##

    escribirtesten = 5
    sheetTest[columnas[escribirtesten + 1]+"2"] = "Prueba Diagnóstico"
    sheetTest[columnas[escribirtesten + 2]+"2"] = "Prueba Final"

    sheetTest[columnas[escribirtesten + 3]+"1"] = "Test"
    sheetTest[columnas[escribirtesten + 3]+"2"] = "Pregunta"
 
    totalPreguntasPret = usuarioPrueba.totalpretest
    totalPreguntasPost = usuarioPrueba.totalpostest

    #for usr in usuarios:

    if(totalPreguntasPret < 1):
        totalPreguntasPret = 10

    if(totalPreguntasPost < 1):
        totalPreguntasPost = 10
        

    puntajepretotal = totalPreguntasPret * 12
    puntajepostotal = totalPreguntasPost * 12

    #Calculo de notas
    iteratorUser = 3
    numeroUser = 1
    for usrtest in listaUsuarios:
        print ("Usuario: " + usrtest.username + "Tests " )

        sheetTest["A" + str(iteratorUser)] = numeroUser
        sheetTest["B" + str(iteratorUser)] = usrtest.RUT
        sheetTest["C" + str(iteratorUser)] = usrtest.ApellidoP
        sheetTest["D" + str(iteratorUser)] = usrtest.ApellidoM
        sheetTest["E" + str(iteratorUser)] = usrtest.nombre
        sheetTest["F" + str(iteratorUser)] = usrtest.username
        sheetTest["F" + str(iteratorUser)].border = right_border

        puntajepreusr = 0
        puntajepostusr = 0

        for  pregu in usrtest.preguntas:
            if(pregu.pretest):
                puntajepreusr += pregu.score

            if(pregu.postest):
                puntajepostusr += pregu.score

        print("Pre-Test " + usrtest.username + ": " + str(puntajepreusr) + "/" + str(puntajepretotal) + "= " +str(puntajepreusr/puntajepretotal))
        print("Post-Test " + usrtest.username + ": " + str(puntajepostusr) + "/" + str(puntajepostotal) + "= " +str(puntajepostusr/puntajepostotal))
        porPrePuntaje = float(puntajepreusr/puntajepretotal) * 100
        porPostPuntaje = float(puntajepostusr/puntajepostotal) * 100

        notaTestPre = 1
        if(porPrePuntaje < 60):
            notaTestPre = (0.05 * porPrePuntaje) + 1
        else:
            notaTestPre = (0.075 * porPrePuntaje) - 0.5 
        if(notaTestPre < 1):
            notaTestPre = 1

        notaTestPost = 1
        if(porPostPuntaje < 60):
            notaTestPost = (0.05 * porPostPuntaje) + 1
        else:
            notaTestPost = (0.075 * porPostPuntaje) - 0.5 
        if(notaTestPost < 1):
            notaTestPost = 1

        print("Escribir nota de control [" + columnas[escribirtesten + 1]+str(iteratorUser)  + "] " + str(notaTestPre))
        print("Escribir nota de control [" + columnas[escribirtesten + 2]+str(iteratorUser)  + "] " + str(notaTestPost))

        usrtest.notaTestPre = round(notaTestPre,2)
        usrtest.notaTestPost = round(notaTestPost,2)

        sheetTest[columnas[escribirtesten + 1]+str(iteratorUser)] = usrtest.notaTestPre
        sheetTest[columnas[escribirtesten + 2]+str(iteratorUser)] = usrtest.notaTestPost



        iteratorUser += 1
        numeroUser += 1

    escribrpreguntatest = escribirtesten + 4
    
    sheetTest[columnas[escribrpreguntatest]+"1"] = "Prueba de diagnóstico"

    preguntasdiag =0

    #Preguntas del pre-test
    for pregunta in usuarioPrueba.preguntas:
        if(pregunta.pretest):
            iteratorUser = 3

            sheetTest[columnas[escribrpreguntatest]+"2"] = pregunta.codigo
            codigotocheck = pregunta.codigo
            preguntasdiag += 1

            for usr in listaUsuarios:

                for pregunta in usr.preguntas:
                    if(pregunta.codigo == codigotocheck):
                        #print("Escribir en " + columnas[olum])
                        sheetTest[columnas[escribrpreguntatest]+str(iteratorUser)] = pregunta.score
                iteratorUser += 1

            escribrpreguntatest += 1

    #Cambio
    if(preguntasdiag > 0):
        sheetTest[columnas[escribrpreguntatest]+"1"] = "Prueba final"
    else:
        sheetTest[columnas[escribrpreguntatest+1]+"1"] = "Prueba final"

    #Preguntas del post-test
    for pregunta in usuarioPrueba.preguntas:

        if(pregunta.postest):
            iteratorUser= 3

            sheetTest[columnas[escribrpreguntatest]+"2"] = pregunta.codigo
            codigotocheck = pregunta.codigo

            for usr in listaUsuarios:
                for pregunta in usr.preguntas:
                    if(pregunta.codigo == codigotocheck):
                        #print("Escribir en " + columnas[olum])
                        sheetTest[columnas[escribrpreguntatest]+str(iteratorUser)] = pregunta.score
                        
                iteratorUser += 1

            escribrpreguntatest += 1



    ##-----------------##Controles##-----------------##

    #los diccionarios no se pueden ordenar :(
    losControlesOrdenados = list(totalPreguntasPorControl.keys())
    losControlesOrdenados.sort()

    iteratorColumnControl = 6 + cantidadControles + 2
    
    i = 1
    escribircontrolen = 5
    while(i<=cantidadControles):
        print("Escribir en [" + columnas[escribircontrolen + i]+"1" + "]" )
        sheetControl[columnas[escribircontrolen + i]+"1"] = "C"+str(losControlesOrdenados[i-1])
        sheetControl.merge_cells(columnas[escribircontrolen + i]+"1:"+columnas[escribircontrolen + i]+"2")
        sheetControl[columnas[escribircontrolen + i]+"1"].border = thin_border
        i += 1

    #escribircontrolen =
    print("Escribir en [" + columnas[escribircontrolen + i]+"1" + "]" )
    sheetControl[ columnas[escribircontrolen + i ] + "1"] = "Nota Control"
    sheetControl.merge_cells(columnas[escribircontrolen + i]+"1:"+columnas[escribircontrolen + i]+"2")
    sheetControl[columnas[escribircontrolen + i]+"1"].border = thin_border

    i+=1
    controlTitulo= ""

    sheetControl[columnas[escribircontrolen + i ] + "1"] = "Control"
    sheetControl[columnas[escribircontrolen + i]+"1"].border = thin_border
    sheetControl[columnas[escribircontrolen + i ] + "2"] = "Pregunta"
    sheetControl[columnas[escribircontrolen + i]+"2"].border = thin_border


    #Preguntasporcontrol = []
    numeropreguntas = 0

    for pregunta in usuarioPrueba.preguntas: #Busqueda de preguntas Control en talleres   
        if (pregunta.esDeControl):
            
            if( pregunta.actividadNombre != controlTitulo): 
                controlTitulo = pregunta.actividadNombre
                #cantidadControles += 1
                #Preguntasporcontrol.append(numeropreguntas)
                numeropreguntas = 0
                sheetControl[columnas[iteratorColumnControl]+'1'] = controlTitulo #str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
            
            #sheetControl[columnas[iteratorColumnControl]+'1']= pregunta.actividadNombre
            #sheetControl[columnas[iteratorColumnControl]+'2']= pregunta.paginaNumero #pregunta.tipo 
            sheetControl[columnas[iteratorColumnControl]+'2']= pregunta.codigo


            codigotocheck = pregunta.codigo

            iteratorUser = 3
            numeroUser = 1
            for usr in listaUsuarios:
                sheetControl["A" + str(iteratorUser)] = numeroUser
                sheetControl["B" + str(iteratorUser)] = usr.RUT
                sheetControl["C" + str(iteratorUser)] = usr.ApellidoP
                sheetControl["D" + str(iteratorUser)] = usr.ApellidoM
                sheetControl["E" + str(iteratorUser)] = usr.nombre
                sheetControl["F" + str(iteratorUser)] = usr.username
                sheetControl["F" + str(iteratorUser)].border = right_border

                for pregunta in usr.preguntas:
                    if(pregunta.codigo == codigotocheck):
                        #print("Escribir en " + columnas[olum])
                        sheetControl[columnas[iteratorColumnControl]+str(iteratorUser)] = pregunta.score
                        
                iteratorUser += 1
                numeroUser += 1
            iteratorColumnControl +=1
            numeropreguntas +=1

    iteratorControl = 1
    

    for control in losControlesOrdenados:
        iteratorUser = 3
        notas = 0
        for usr in listaUsuarios:
            print ("Usuario: " + usr.username + " control " + str(control))
            totalpcontrolparaUsuario = totalPreguntasPorControl[control]
            puntajetotal = totalpcontrolparaUsuario * 12
            puntajeusuario = 0

            for preguntisima in usr.preguntas:
                if(preguntisima.tallerNumero == int(control) and preguntisima.esDeControl):
                    #print ("la pregunta " + preguntisima.codigo + " es del control " + str(control) )
                    print(usr.username + " saco " + str(preguntisima.score) + " en pregunta "  + preguntisima.codigo)
                    ##time.sleep(1)
                    gettedScore =  preguntisima.score 
                    if( preguntisima.score  < 0): 
                        logErrores.append(usr.username + "tiene puntaje de "+ str(preguntisima.score) + " en control" + str(control) + " __blockkey: " + str(preguntisima.blockKeyCompleto ))
                        gettedScore = 0
                    puntajeusuario += gettedScore
            
            print(usr.username + ": " + str(puntajeusuario) + "/" + str(puntajetotal) + "= " +str(puntajeusuario/puntajetotal))
            porPuntaje = float(puntajeusuario/puntajetotal) * 100
            ##time.sleep(4)

            if(porPuntaje < 60):
                notacontrol = (0.05 * porPuntaje) + 1
            else:
                notacontrol = (0.075* porPuntaje) - 0.5 

            if(notacontrol < 1):
                logErrores.append(usr.username + "tiene nota de control "+ str(control) +" menor a 1: " + str(notacontrol) )
                notacontrol = 1

            print("Escribir nota de control [" + columnas[escribircontrolen + iteratorControl]+str(iteratorUser)  + "]")
            sheetControl[columnas[escribircontrolen + iteratorControl]+str(iteratorUser)] =  round(notacontrol,2)   #totalPreguntasPorControl[control]
            usr.notaControles.append(round(notacontrol,2))

            print("cantidad de controles " + str(cantidadControles))
            ##time.sleep(5)

            if(iteratorControl == cantidadControles):  
                indicenota = 0
                notas = 0
                while(indicenota < cantidadControles):
                    indicenota += 1
                    notatomada = sheetControl[ columnas[escribircontrolen + indicenota] + str(iteratorUser)].value
                    print( "notatomada de "+  usr.username  +" [" + columnas[escribircontrolen + indicenota] +   str(iteratorUser) + "]: "+ str(notatomada))
                    ##time.sleep(3)
                    if(notatomada is None):
                        notatomada = 0
                    notas += float(notatomada)
                notafinal = notas / cantidadControles
                if(notafinal < 1):
                    logErrores.append(usr.username + "tiene nota final de controles menor a 1: " + str(notafinal) )
                    notafinal = 1
                print("Nota control de " + usr.username + ": " + str(notafinal))
                ##time.sleep(3)
                sheetControl[columnas[escribircontrolen + iteratorControl+1]+str(iteratorUser)] =  round(notafinal,2)    

            iteratorUser += 1
            notas = 0
        iteratorControl += 1

    ##-----------------##Preguntas calificadas##-----------------##

    if(cantidadTipoPreguntaEvaluada > 0):
        
        iteratorColumnPreva = 7 + cantidadTipoPreguntaEvaluada + 1

        i = 1
        escribirPrevaluen = 5
        while(i<=cantidadTipoPreguntaEvaluada):
            print("Escribir en [" + columnas[escribirPrevaluen + i]+"1" + "]" )
            sheetPreguntas[columnas[escribirPrevaluen + i]+"1"] = "PT"+str(i)
            sheetPreguntas.merge_cells(columnas[escribirPrevaluen + i]+"1:"+columnas[escribirPrevaluen + i]+"3")
            sheetPreguntas[columnas[escribirPrevaluen + i]+"1"].border = thin_border
            i += 1

        print("Escribir en [" + columnas[escribirPrevaluen + i]+"1" + "]" )
        sheetPreguntas[ columnas[escribirPrevaluen + i ] + "1"] = "Nota PC"
        sheetPreguntas.merge_cells(columnas[escribirPrevaluen + i]+"1:"+columnas[escribirPrevaluen + i]+"3")
        sheetPreguntas[columnas[escribirPrevaluen + i]+"1"].border = thin_border
        talllerTexto= ""

        i+=1

        sheetPreguntas[columnas[escribirPrevaluen + i ] + "1"] = "Taller"
        sheetPreguntas[columnas[escribirPrevaluen + i]+ "1"].border = thin_border
        sheetPreguntas[columnas[escribirPrevaluen + i ] + "2"] = "Pregunta"
        sheetPreguntas[columnas[escribirPrevaluen + i]+ "2"].border = thin_border
        sheetPreguntas[columnas[escribirPrevaluen + i ] + "3"] = "Key"
        sheetPreguntas[columnas[escribirPrevaluen + i]+ "3"].border = thin_border


        numeropreguntas = 0

        for pregunta in usuarioPrueba.preguntas: #Busqueda de preguntas Control en talleres   
            if (pregunta.preguntaEvaluada):
                if(pregunta.tallerNombre != talllerTexto): 
                    talllerTexto = pregunta.tallerNombre
                    #cantidadControles += 1
                    #Preguntasporcontrol.append(numeropreguntas)
                    numeropreguntas = 0
                    sheetPreguntas[columnas[iteratorColumnPreva]+'1'] = str(pregunta.tallerNumero) + ": " + pregunta.tallerNombre
                    
                sheetPreguntas[columnas[iteratorColumnPreva]+'2']= pregunta.actividadNombre
                sheetPreguntas[columnas[iteratorColumnPreva]+'3']= pregunta.codigo


                codigotocheck = pregunta.codigo

                iteratorUser = 4
                numeroUser = 1
                for usr in listaUsuarios:
                    sheetPreguntas["A" + str(iteratorUser)] = numeroUser
                    sheetPreguntas["B" + str(iteratorUser)] = usr.RUT
                    sheetPreguntas["C" + str(iteratorUser)] = usr.ApellidoP
                    sheetPreguntas["D" + str(iteratorUser)] = usr.ApellidoM
                    sheetPreguntas["E" + str(iteratorUser)] = usr.nombre
                    sheetPreguntas["F" + str(iteratorUser)] = usr.username
                    sheetPreguntas["F" + str(iteratorUser)].border = right_border

                    for pregunta in usr.preguntas:
                        if(pregunta.codigo == codigotocheck):
                            #print("Escribir en " + columnas[olum])
                            sheetPreguntas[columnas[iteratorColumnPreva]+str(iteratorUser)] = pregunta.score
                            
                    iteratorUser += 1
                    numeroUser += 1
                iteratorColumnPreva +=1
                numeropreguntas +=1

        #los diccionarios no se pueden ordenar :( asi que hay que hacer este fix        
        losTalleresPEordenados = list(totalPreguntasEvaluadas.keys())
        losTalleresPEordenados.sort()
        for preva in losTalleresPEordenados:
            iteratorUser = 4
            notas = 0
            for usr in listaUsuarios:
                print ("Usuario: " + usr.username)
                totalprevaporUsuario = totalPreguntasEvaluadas[preva]
                print("Total puntaje usuario " + str(totalprevaporUsuario) +  " total pc " + str(totalPreguntasEvaluadas) )
                puntajetotal = totalprevaporUsuario * 12
                puntajeusuario = 0

                ##time.sleep(3)

                for preguntisima in usr.preguntas:
                    if(preguntisima.tallerNumero == int(preva) and preguntisima.preguntaEvaluada):
                        #print ("la pregunta " + preguntisima.codigo + " es del control " + str(control) )
                        print(usr.username + " saco " + str(preguntisima.score) + " en pregunta "  + preguntisima.codigo)
                        ##time.sleep(3)
                        gettedScore =  preguntisima.score 
                        if( preguntisima.score  < 0):
                            logErrores.append(usr.username + "tiene puntaje de "+ str(preguntisima.score) + " pregunta calificada" + str(preva) + " __blockkey: " + str(preguntisima.blockKeyCompleto ))
                            gettedScore = 0
                        puntajeusuario += preguntisima.score 
                

                porPuntaje = float(puntajeusuario/puntajetotal)*100

                if(porPuntaje < 60):
                    notaPregunta = (0.05 * porPuntaje) + 1
                else:
                    notaPregunta = (0.075* porPuntaje) - 0.5 

                if(notaPregunta < 1):
                    notaPregunta = 1
                #print("preva" + preva)
                print("Cantidad de preguntas calificadas " + str(cantidadTipoPreguntaEvaluada))
                print("Escribir nota de pregunta calificada " + preva + " en " + columnas[escribirPrevaluen + int(preva)] +str(iteratorUser) + " Nota " + str(notaPregunta) )
                sheetPreguntas[columnas[escribirPrevaluen + int(preva)]+str(iteratorUser)] =  round(notaPregunta,2)   #totalPreguntasPorControl[control]
                usr.notaPreguntasCalificadas.append(round(notaPregunta,2))
                ##time.sleep(4)

                if(int(preva) == cantidadTipoPreguntaEvaluada):  
                    indicenota = 0
                    notas = 0
                    while(indicenota < cantidadTipoPreguntaEvaluada):
                        indicenota += 1
                        notatomada = sheetPreguntas[ columnas[escribirPrevaluen + indicenota] + str(iteratorUser)].value
                        print( "notatomada de "+  usr.username  +" [" + columnas[escribirPrevaluen + indicenota] +   str(iteratorUser) + "]: "+ str(notatomada))
                        ##time.sleep(3)
                        if(notatomada is None):
                            notatomada = 0
                        notas += float(notatomada)
                    notafinal = notas / cantidadTipoPreguntaEvaluada
                    if(notafinal < 1):
                        logErrores.append(usr.username + "tiene nota final de preguntas evaluadas menor a 1: " + str(notafinal) )
                        notafinal = 1
                    print("Nota control de " + usr.username + ": " + str(notafinal))
                    ##time.sleep(3)
                    sheetPreguntas[columnas[escribirPrevaluen + int(preva)+1]+str(iteratorUser)] =  round(notafinal,2)    

                iteratorUser += 1
                notas = 0
           
    ##-----------------##Reporte completo##-----------------##

    #-----------------#Talleres#-----------------#
    sheetCompleto["H1"] = "Avance talleres"
    sheetCompleto["H2"] = "% Correctas"
    sheetCompleto["I2"] = "% Contestadas"
    sheetCompleto["J2"] = "Nota"
    columnaincio = 9
       
    preguntastotales = 0
    preguntaEvaluables = 0

    for pregprueba in usuarioPrueba.preguntas:
        if(pregprueba.esDeControl == False and pregprueba.esDeEncuesta == False and pregprueba.preguntaEvaluada == False and pregprueba.consentimiento == False and pregprueba.reglamento == False  and pregprueba.tallerNumero > 0):
            preguntastotales += 1 

        if(pregprueba.esDeControl == False and pregprueba.esDeEncuesta == False and pregprueba.preguntaEvaluada == False and pregprueba.consentimiento == False  and pregprueba.reglamento == False and pregprueba.tallerNumero > 0 and pregprueba.esCorrecta != None ):
            preguntaEvaluables += 1

    numeroUser = 1
    iteratorUser = 3

    print(" Preguntas evaluables :" + str(preguntaEvaluables) + " preguntas totales " + str(preguntastotales) )
    #time.sleep(1)
    
    for estudiante in usuarios:

        sheetCompleto["A" + str(iteratorUser)] = numeroUser
        sheetCompleto["B" + str(iteratorUser)] = estudiante.RUT
        sheetCompleto["C" + str(iteratorUser)] = estudiante.ApellidoP
        sheetCompleto["D" + str(iteratorUser)] = estudiante.ApellidoM
        sheetCompleto["E" + str(iteratorUser)] = estudiante.nombre
        sheetCompleto["F" + str(iteratorUser)] = estudiante.username
        sheetCompleto["G" + str(iteratorUser)] = estudiante.ultimaconexion
        #sheetCompleto["F" + str(iteratorUser)].border = right_border

        totalpreguntasestudiante = 0
        totalpreguntasbuenas = 0

        for preg in estudiante.preguntas:
            if(preg.esDeControl == False and preg.esDeEncuesta == False and preg.preguntaEvaluada == False and preg.consentimiento == False and preg.reglamento == False and preg.tallerNumero > 0 and preg.esunapreguntade != 'pre-test'):
                for preguprueba in usuarioPrueba.preguntas:
                    if preg.codigo == preguprueba.codigo:
                        totalpreguntasestudiante += 1
                        if(preg.esCorrecta):
                            totalpreguntasbuenas += 1
        
        # FIX THOMAS
        preguntastotales = float(total_auxiliar)
        # FIX THOMAS


        if(preguntaEvaluables >0):
            procentajetotalcorrectas = round(float(float(totalpreguntasbuenas)/float(preguntaEvaluables)) * 100 )
        else:
            procentajetotalcorrectas = 0
        
        if(preguntastotales >0):
            procentajetotalpreguntas = round(float(float(totalpreguntasestudiante)/float(preguntastotales)) * 100)
        else:
           procentajetotalpreguntas = 0

        print ("calcular % Correctas" + str(totalpreguntasbuenas) + "/" + str(preguntaEvaluables))
        print ("calcular % Totales " + str(totalpreguntasestudiante) + "/" + str(preguntastotales))

        if(totalpreguntasestudiante > preguntastotales):
            print("el Usuario " + estudiante.nombre + " contesto mas que el usuario de prueba ")
            logErrores.append("el Usuario " + estudiante.nombre + " contesto mas preguntas de taller que el usuario de prueba " + str(totalpreguntasestudiante) + " VS. "  + str(preguntastotales) )
            time.sleep(3)
            totalpreguntasestudiante = preguntaEvaluables

        if(preguntaEvaluables> 0):
            porNota = procentajetotalpreguntas
        else:
            porNota = 0
        
        if(porNota < 60):
            nota = (0.05 * porNota) + 1
        else:
            nota = (0.075* porNota) - 0.5 

        if(nota < 1):
            nota = 1

        if(nota >= 7.0):
            logErrores.append("el Usuario " + estudiante.nombre + " tiene una nota mayor que 7 en preguntas de taller: " + str(nota) + " _%Nota " + str(porNota) +"% _%Correctas " + str(procentajetotalcorrectas) + "% _%Respondidas " + str(procentajetotalpreguntas) + "%.")
            nota = 7

        print ("Porcentajes para " + str(estudiante.username) + " %Correctas " + str(procentajetotalcorrectas) + " %Respondidas " + str(procentajetotalpreguntas) )
        print ("%Nota " + str(porNota) + " nota " + str(nota) )
        ##time.sleep(12)

        sheetCompleto["H"+str(iteratorUser)] = str(procentajetotalcorrectas) + "%"
        sheetCompleto["I"+str(iteratorUser)] = str(procentajetotalpreguntas) + "%"
        sheetCompleto["J"+str(iteratorUser)] = round(nota,1) 
        
        iteratorUser += 1
        numeroUser += 1

    #------------------------#Tests#--------------------------#

    EscribirTestCom = 9 
    sheetCompleto[columnas[EscribirTestCom + 1]+'1'] = "Tests"
    sheetCompleto[columnas[EscribirTestCom + 1]+'2'] = "Prueba de diagnóstico"
    sheetCompleto[columnas[EscribirTestCom + 2]+'2'] = "Prueba final"

    iteratorUser = 3
    for usr in listaUsuarios:

        puntajepreusr = 0
        puntajepostusr = 0

        for  pregu in usr.preguntas:
            if(pregu.pretest):
                puntajepreusr += pregu.score

            if(pregu.postest):
                puntajepostusr += pregu.score

        print("Pre-Test " + usrtest.username + ": " + str(puntajepreusr) + "/" + str(puntajepretotal) + "= " +str(puntajepreusr/puntajepretotal))
        print("Post-Test " + usrtest.username + ": " + str(puntajepostusr) + "/" + str(puntajepostotal) + "= " +str(puntajepostusr/puntajepostotal))
        porPrePuntaje = float(puntajepreusr/puntajepretotal) * 100
        porPostPuntaje = float(puntajepostusr/puntajepostotal) * 100

        notaTestPre = 1
        if(porPrePuntaje < 60):
            notaTestPre = (0.05 * porPrePuntaje) + 1
        else:
            notaTestPre = (0.075 * porPrePuntaje) - 0.5 
        if(notaTestPre < 1):
            notaTestPre = 1

        notaTestPost = 1
        if(porPostPuntaje < 60):
            notaTestPost = (0.05 * porPostPuntaje) + 1
        else:
            notaTestPost = (0.075 * porPostPuntaje) - 0.5 
        if(notaTestPost < 1):
            notaTestPost = 1

        sheetCompleto[columnas[EscribirTestCom + 1]+ str(iteratorUser)] = round(notaTestPre,2)
        sheetCompleto[columnas[EscribirTestCom + 2]+ str(iteratorUser)] = round(notaTestPost,2)

        iteratorUser += 1


    #-----------------#Preguntas calificadas#-----------------#

    if(cantidadTipoPreguntaEvaluada > 0):
        i= 1
        EscribirPregEvaluada= EscribirTestCom + 2
        sheetCompleto[columnas[EscribirPregEvaluada + 1]+'1'] = "Preguntas Calificadas"

        while(i<=cantidadTipoPreguntaEvaluada):
            print ("Escribir en [" + columnas[EscribirPregEvaluada + i]+"1" + "]" )
            sheetCompleto[columnas[EscribirPregEvaluada + i]+"2"] = "PET"+str(i)
            i+=1

        sheetCompleto[columnas[EscribirPregEvaluada + i]+"2"] = "Nota preguntas calificadas"

        iteratorUser = 3
        for usr in usuarios:
            
            npregev = 0
            notaenarreglo = 0

            while npregev < cantidadTipoPreguntaEvaluada:
                
                npregev += 1
                print("escribir nota de " + usr.username + " en [" + str(columnas[EscribirPregEvaluada + npregev]) + str(iteratorUser) +"] la nota " + str( usr.notaPreguntasCalificadas[notaenarreglo] ) )
                sheetCompleto[columnas[EscribirPregEvaluada + npregev]+str(iteratorUser)] =  usr.notaPreguntasCalificadas[notaenarreglo]
                notaenarreglo += 1
                
                #iteratorpregun = 0
                if(npregev == cantidadTipoPreguntaEvaluada):
                    indicenota = 0
                    notas = 0
                    while(indicenota < cantidadTipoPreguntaEvaluada):
                        indicenota += 1
                        notatomada = sheetCompleto[ columnas[EscribirPregEvaluada + indicenota] + str(iteratorUser)].value
                        print( "notatomada de "+  usr.username  +" [" + columnas[escribircontrolen + indicenota] +   str(iteratorUser) + "]: "+ str(notatomada))
                        ##time.sleep(3)
                        notas += float(notatomada)
                    notafinal = notas / cantidadTipoPreguntaEvaluada
                    if(notafinal < 1):
                        logErrores.append("el Usuario " + usr.nombre + " tiene nota de preguntas calificadas inferior a 1")
                        notafinal = 1
                    sheetCompleto[columnas[EscribirPregEvaluada + npregev +1 ]+str(iteratorUser)] =  round(notafinal,2)

            iteratorUser += 1
            

    #-----------------#Controles#-----------------#

    pivote = EscribirTestCom + 2  
    if(cantidadTipoPreguntaEvaluada > 0):       
        pivote = EscribirPregEvaluada
    
    i= 1
    if(cantidadTipoPreguntaEvaluada > 0):
        EscribirControl =  pivote +  cantidadTipoPreguntaEvaluada + 1
    else:
        EscribirControl =  pivote
    
    sheetCompleto[columnas[EscribirControl + 1]+'1'] = "Controles"

    while(i<=cantidadControles):
        print ("Escribir en [" + columnas[EscribirControl + i]+"1" + "]" )
        sheetCompleto[columnas[EscribirControl + i]+"2"] = "C"+ str(losControlesOrdenados[i-1])
        i+=1

    sheetCompleto[columnas[EscribirControl + i]+"2"] = "Promedio Control"
    
    controlIterado = 0
    iteratorUser = 3 
    for usr in usuarios:
        
        ncontrol = 0
        notaenarreglo = 0
        
        while ncontrol < cantidadControles:
            ncontrol += 1
            print("escribir nota de " + usr.username + " en [" + str(columnas[EscribirControl + ncontrol]) + str(iteratorUser) +"] la nota de control " + str( usr.notaControles[notaenarreglo] ) )
            sheetCompleto[columnas[EscribirControl + ncontrol]+str(iteratorUser)] =  usr.notaControles[notaenarreglo]
            notaenarreglo += 1
            
            if( ncontrol == cantidadControles):
                indicenota = 0
                notas = 0
                while(indicenota < cantidadControles):
                    indicenota += 1
                    notatomada = sheetCompleto[ columnas[EscribirControl + indicenota] + str(iteratorUser)].value
                    print( "notatomada de "+  usr.username  +" [" + columnas[escribircontrolen + indicenota] +   str(iteratorUser) + "]: "+ str(notatomada))
                    ##time.sleep(3)
                    notas += float(notatomada)
                notafinal = notas / cantidadControles
                if(notafinal < 1):
                    logErrores.append("el Usuario " + usr.nombre + " tiene nota de preguntas controles a 1")
                    notafinal = 1
                sheetCompleto[columnas[EscribirControl + ncontrol + 1 ]+str(iteratorUser)] =   round(notafinal,2)
        iteratorUser += 1   

   

    #-----------------#Encuestas y otros#-----------------#
   
    i= 1
    print("Encuestas " + str(nombresEncuestas))
    #time.sleep(2)
    cantidadEncuestas =  len(nombresEncuestas)
    cantidadOtros =  cantidadEncuestas + 2
    Otros = ["Reglamento","Consentimiento"] + nombresEncuestas
    espacioadicional = 0
    if cantidadControles > 0:
        espacioadicional += 1
    if cantidadTipoPreguntaEvaluada > 0:
        espacioadicional += 1
    EscribirOtros =  pivote +  cantidadTipoPreguntaEvaluada + cantidadControles + espacioadicional
    sheetCompleto[columnas[EscribirOtros + 1]+'1'] = "Encuestas y otros"

    while(i<=cantidadOtros):
        print ("Escribir en [" + columnas[EscribirOtros + i]+"1" + "]" )
        sheetCompleto[columnas[EscribirOtros + i]+"2"] = Otros[i-1]
        i+=1

    iteratorUser = 3
    for usren in usuarios:

        canEncuestas  = []
        j=0
        while(j<=cantidadEncuestas):
            canEncuestas.append(0)
            j+=1
        
        for encues in usren.preguntas:
            if(encues.reglamento == True):
                sheetCompleto[columnas[EscribirOtros + 1] + str(iteratorUser)] = encues.respuesta
                print("Es un reglamento")

            if(encues.consentimiento == True):
                sheetCompleto[columnas[EscribirOtros + 2] + str(iteratorUser)] = encues.respuesta
                print("Es un consentimiento")

            if(encues.esDeEncuesta == True):
                print("Es una encuesta")
                numEncuesta = encues.numeroEncuesta
                canEncuestas[numEncuesta] += 1
                #sheetCompleto[columnas[EscribirOtros + 2 + numEncuesta] + str(iteratorUser)] = "Respondi"

        print("cantidad de encuestas " + str(cantidadEncuestas) +" "+ str(canEncuestas) + " VS "+ str(totalEncuestas))
        j=1
        ##time.sleep(32)
        while j <= cantidadEncuestas: #verificar si <=
            #print ("canEncuestas " + str(canEncuestas[j]) )
            #print ("totalEncuestas "  +  str(totalEncuestas[str(j)]) )
            
            print("j:" +str(j) + " canEncuestas " + str(canEncuestas[j]) + "/" + " totalEncuestas " +  str(totalEncuestas[str(j)]) ) 
            sheetCompleto[columnas[EscribirOtros + 2 + j] + str(iteratorUser)] = str(canEncuestas[j]) + "/" + str(totalEncuestas[str(j)])
            
            j+= 1

        iteratorUser += 1

    ##-------------------##Pestaña de datos##-------------------##

    iteratorUser = 1

    for usurdatos in usuarios:
        iteratorUser += 1
        sheetDatos['A' + str(iteratorUser)] = iteratorUser-1
        sheetDatos['B' + str(iteratorUser)] = usurdatos.RUT
        sheetDatos['C' + str(iteratorUser)] = usurdatos.ApellidoP
        sheetDatos['D' + str(iteratorUser)] = usurdatos.ApellidoM
        sheetDatos['E' + str(iteratorUser)] = usurdatos.nombre 
        sheetDatos['F' + str(iteratorUser)] = usurdatos.username
        sheetDatos['G' + str(iteratorUser)] = usurdatos.email
        sheetDatos['H' + str(iteratorUser)] = usurdatos.telefono
        sheetDatos['I' + str(iteratorUser)] = usurdatos.RBD
        sheetDatos['J' + str(iteratorUser)] = usurdatos.establecimiento
        sheetDatos['K' + str(iteratorUser)] = usurdatos.ultimaconexion

    ##----------------##Limpieza de filas##-----------------------##

    if(Limpiar):
        iniciolimpieza = len(usuarios)

        inicioCompleto = inicioControles = inicioRepAvance = 2
        inicioPregEva = 3
        inicioAvance = inicioFechas =  4
        
        sheetCompleto.delete_rows(iniciolimpieza + inicioCompleto + 1,len(usuarios))
        sheetAvance.delete_rows(iniciolimpieza + inicioAvance + 1, len(usuarios))
        sheetControl.delete_rows(iniciolimpieza + inicioControles + 1, len(usuarios))
        if(cantidadTipoPreguntaEvaluada > 0):
            sheetPreguntas.delete_rows(iniciolimpieza + inicioPregEva + 1, len(usuarios))
        sheetRepAvance.delete_rows(iniciolimpieza + inicioRepAvance + 1, len(usuarios))
        sheetFechas.delete_rows(iniciolimpieza + inicioFechas + 1, len(usuarios))
    
    ##-----------------##Ajustes de espacio##-----------------##

    sheetAvance.column_dimensions['A'].width = 3
    sheetAvance.column_dimensions['B'].width = 12
    sheetAvance.column_dimensions['C'].width = 16
    sheetAvance.column_dimensions['D'].width = 16
    sheetAvance.column_dimensions['E'].width = 16
    sheetAvance.column_dimensions['F'].width = 14
    sheetAvance.column_dimensions['G'].width = 10

    sheetCompleto.column_dimensions['A'].width = 3
    sheetCompleto.column_dimensions['B'].width = 12
    sheetCompleto.column_dimensions['C'].width = 16
    sheetCompleto.column_dimensions['D'].width = 16
    sheetCompleto.column_dimensions['E'].width = 16
    sheetCompleto.column_dimensions['F'].width = 14
    sheetCompleto.column_dimensions['G'].width = 10

    sheetControl.column_dimensions['A'].width = 3
    sheetControl.column_dimensions['B'].width = 12
    sheetControl.column_dimensions['C'].width = 16
    sheetControl.column_dimensions['D'].width = 16
    sheetControl.column_dimensions['E'].width = 16
    sheetControl.column_dimensions['F'].width = 14
    sheetControl.column_dimensions['G'].width = 10

    sheetTest.column_dimensions['A'].width = 3
    sheetTest.column_dimensions['B'].width = 12
    sheetTest.column_dimensions['C'].width = 16
    sheetTest.column_dimensions['D'].width = 16
    sheetTest.column_dimensions['E'].width = 16
    sheetTest.column_dimensions['F'].width = 14
    sheetTest.column_dimensions['G'].width = 16
    sheetTest.column_dimensions['H'].width = 16


    if(cantidadTipoPreguntaEvaluada >0):
        sheetPreguntas.column_dimensions['A'].width = 3
        sheetPreguntas.column_dimensions['B'].width = 12
        sheetPreguntas.column_dimensions['C'].width = 16
        sheetPreguntas.column_dimensions['D'].width = 16
        sheetPreguntas.column_dimensions['E'].width = 16
        sheetPreguntas.column_dimensions['F'].width = 14
        sheetPreguntas.column_dimensions['G'].width = 10

    sheetFechas.column_dimensions['A'].width = 3
    sheetFechas.column_dimensions['B'].width = 12
    sheetFechas.column_dimensions['C'].width = 16
    sheetFechas.column_dimensions['D'].width = 16
    sheetFechas.column_dimensions['E'].width = 16
    sheetFechas.column_dimensions['F'].width = 14
    sheetFechas.column_dimensions['G'].width = 10

    sheetRepAvance.column_dimensions['A'].width = 3
    sheetRepAvance.column_dimensions['B'].width = 12
    sheetRepAvance.column_dimensions['C'].width = 16
    sheetRepAvance.column_dimensions['D'].width = 16
    sheetRepAvance.column_dimensions['E'].width = 16
    sheetRepAvance.column_dimensions['F'].width = 14
    sheetRepAvance.column_dimensions['G'].width = 10

    sheetDatos.column_dimensions['A'].width = 3
    sheetDatos.column_dimensions['B'].width = 12
    sheetDatos.column_dimensions['C'].width = 16
    sheetDatos.column_dimensions['D'].width = 16
    sheetDatos.column_dimensions['E'].width = 20
    sheetDatos.column_dimensions['F'].width = 14
    sheetDatos.column_dimensions['G'].width = 30
    sheetDatos.column_dimensions['H'].width = 10
    sheetDatos.column_dimensions['I'].width = 10
    sheetDatos.column_dimensions['J'].width = 16
    sheetDatos.column_dimensions['k'].width = 20


    ##-----------------##Guardado del archivo##-----------------##

    nombrereporte = "Reporte_" + nombre + ".xlsx"

    if(location_to_save_report == ""):
        location = os.path.join(os.path.dirname(__file__), nombrereporte)
    else:
        location =  location_to_save_report
    print("Location: " + str(location))
    if os.path.isfile(location):
        os.remove(location)
    book.save(location)
    return location

def creteDocumentoEncuesta(listaUsuarios):
    global cantidadControles
    print("Crear xls encuesta")

    #cantidadTalleres = 0
    book2 = openpyxl.Workbook()

    #Definición de bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    rightbottom_border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
    right_border = Border(right = Side(style = 'thin'))  
    bottom_border = Border(bottom = Side(style = 'thin'))

    #sheet.merge_cells('A2:D4') 

    #Llenar columnas
    columnas = []
    for char in string.ascii_uppercase:
        columnas.append(char)
    for char1 in string.ascii_uppercase:
        for char2 in string.ascii_uppercase:
            columnas.append(char1 + char2)
    for char01 in string.ascii_uppercase:
        for char02 in string.ascii_uppercase:
            for char03 in string.ascii_uppercase:
                columnas.append(char01 + char02 + char03)

    sheet = book2.active
    sheet.title = "Encuesta"
    sheet.sheet_properties.tabColor="470405"  

    sheet["A1"] = "N°"
    sheet["A1"].border = thin_border
    sheet["B1"] = "RUT"
    sheet["B1"].border = thin_border
    sheet["C1"] = "Nombre"
    sheet["C1"].border = thin_border
    sheet["D1"] = "Usuario EDX"
    sheet["D1"].border = thin_border
    sheet["E1"] = "Encuesta"
    sheet["E1"].border = thin_border
    sheet["F1"] = "Respuesta"
    sheet["F1"].border = thin_border
    sheet["G1"] = "Blockey"
    sheet["G1"].border = thin_border

    iterator = 2
    numero = 0
    for usr in listaUsuarios:
        for preg in usr.preguntas:
            if preg.esDeEncuesta:
                numero += 1
                sheet["A" + str(iterator)] = numero
                sheet["B" + str(iterator)] = usr.RUT
                sheet["C" + str(iterator)] = usr.ApellidoP + " " + usr.nombre
                sheet["D" + str(iterator)] = usr.username
                sheet["E" + str(iterator)] = preg.numeroEncuesta
                sheet["F" + str(iterator)] = preg.respuesta 
                sheet["G" + str(iterator)] = preg.blockKeyCompleto
                iterator += 1

    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 64
    sheet.column_dimensions['G'].width = 125

    if(location_to_save_report == ""):
        location = os.path.join(os.path.dirname(__file__), "Encuesta.xlsx")
    else:
        location =  location_to_save_report
    print("Location: " + str(location))
    book2.save(location)


def creteDocumentoEncuesta2(listaEncuesta,nombre):
    global cantidadControles
    print("Crear xls encuesta")

    #cantidadTalleres = 0
    book2 = openpyxl.Workbook()

    #Definición de bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    rightbottom_border = Border(right=Side(style='thin'),bottom=Side(style='thin'))
    right_border = Border(right = Side(style = 'thin'))  
    bottom_border = Border(bottom = Side(style = 'thin'))

    #sheet.merge_cells('A2:D4') 

    #Llenar columnas
    columnas = []
    for char in string.ascii_uppercase:
        columnas.append(char)
    for char1 in string.ascii_uppercase:
        for char2 in string.ascii_uppercase:
            columnas.append(char1 + char2)
    for char01 in string.ascii_uppercase:
        for char02 in string.ascii_uppercase:
            for char03 in string.ascii_uppercase:
                columnas.append(char01 + char02 + char03)

    sheet = book2.active
    sheet.title = "Encuesta"
    sheet.sheet_properties.tabColor="470405"  

    sheet["A1"] = "N°"
    sheet["A1"].border = thin_border
    sheet["B1"] = "RUT"
    sheet["B1"].border = thin_border
    sheet["C1"] = "Nombre"
    sheet["C1"].border = thin_border
    sheet["D1"] = "Usuario EDX"
    sheet["D1"].border = thin_border
    sheet["E1"] = "Encuesta N°"
    sheet["E1"].border = thin_border
    sheet["F1"] = "Encuesta"
    sheet["F1"].border = thin_border
    sheet["G1"] = "Pag"
    sheet["G1"].border = thin_border
    sheet["H1"] = "pregunta"
    sheet["H1"].border = thin_border
    sheet["I1"] = "Identificador"
    sheet["I1"].border = thin_border
    sheet["J1"] = "Respuesta"
    sheet["J1"].border = thin_border
    sheet["k1"] = "Link"
    sheet["K1"].border = thin_border
    

    iterator = 2
    numero = 0
    for encu in listaEncuesta:
        if(encu.nombreusuario != nombreUsuarioPrueba and encu.nombrepregunta != '' and encu.respuesta != ''):
            numero += 1
            sheet["A" + str(iterator)] = numero
            sheet["B" + str(iterator)] = encu.usuariorut
            sheet["C" + str(iterator)] = encu.nombre
            sheet["D" + str(iterator)] = encu.nombreusuario
            sheet["E" + str(iterator)] = encu.numeroencuesta
            sheet["F" + str(iterator)] = encu.nombreencuesta
            sheet["G" + str(iterator)] = encu.pagina
            sheet["H" + str(iterator)] = encu.nombrepregunta
            sheet["I" + str(iterator)] = str(encu.numeroidentificatorio1) + "_" + str(encu.numeroidentificatorio2)
            sheet["J" + str(iterator)] = encu.respuesta
            sheet["K" + str(iterator)] = "https://studio.cmmeduformacion.uchile.cl/container/" + encu.blockkey
            iterator += 1
    
    sheet.column_dimensions['A'].width = 3
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 22
    sheet.column_dimensions['D'].width = 14
    sheet.column_dimensions['E'].width = 8
    sheet.column_dimensions['F'].width = 14
    sheet.column_dimensions['G'].width = 8
    sheet.column_dimensions['I'].width = 8
    sheet.column_dimensions['H'].width = 20    
    sheet.column_dimensions['J'].width = 40    
    sheet.column_dimensions['K'].width = 30
    sheet.column_dimensions['L'].width = 80
    
    filename = "Encuestas_" + nombre + ".xlsx"

    if(location_to_save_report == ""):
        location = os.path.join(os.path.dirname(__file__), filename)
    else:
        location =  location_to_save_report
    print("Location: " + str(location))
    book2.save(location)
    return location

def crearListaXLS(filelocation):
    print("Crear la lista de usuarios de " + filelocation )
    #time.sleep(5)
    
    File = filelocation

    nombres = []

    print("Arvchivo: " + File)
    usuarioActual = ""

    usuarioPrueba.username = nombreUsuarioPrueba

    with open(File, newline='', encoding="utf-8") as f: #recolectar nombres
        first = True
        reader = csv.reader(f)
        for row in reader:
            #print ("["+str(i) + "]: ")f
            #Estos index se movieron en +1 porque antes no habia rut en la lista de usuarios
            user  = usuario()
            username = row[2]
            email = row[4]
            nacimiento = row[7]
            #print ("NoMbRe:___" + nombre + "__")

            if ( (username  not in nombres) and (username != "username") ):
                print (username  + " Vs. " + "username")
                if(saltarPrimeriaLienaListaUsuarios==True and first == True):
                    first = False
                    print("es la primera linea")
                    ##time.sleep(4)
                else:
                    nombres.append(username)
                    print("agregar " + username)
                    user.username = username
                    user.email = email
                    user.nacimiento = nacimiento
                    usuarios.append(user)
                    #time.sleep(1)
    print("Usuarios: " + str(usuarios) )

def crearListaOra(oralocation):
    print("Leer CVS")
    first = True

    File = oralocation
    with open(File,newline='',encoding="utf-8") as f:
        reader= csv.reader(f)
        for row in reader:
            oratoadd = ora()
            oratoadd.submissioid = row[0]
            oratoadd.itemid = row[1]
            oratoadd.studentid = row[2]
            oratoadd.fecharespuesta = row[3]

            respuestatoadd = find_between(row[4],"{u'parts': [{u'text': u'","'}]}")
            oratoadd.respuesta = respuestatoadd

            oratoadd.fechacalificacion = row[7]
            
            if(first == True):
                first = False
            else:
                if(str(row[8]) != ""):
                    oratoadd.score = int(row[8])
                    oratoadd.calificada = True
                else:
                    oratoadd.score = 0
                    oratoadd.calificada = False
                listaoras.append(oratoadd)

def LoadUsuarioDePrueba(fileLocation):
    global totalPreguntasPorTaller
    global totalPreguntasPorControl
    global totalPreguntasEvaluadas
    global totalPregutnasEncuestas
    global totalPreguntasBuenasOMalas
    global totalEncuestas
    global cantidadControles
    global cantidadEncuestas
    global cantidadTipoPreguntaEvaluada
    global nombresEncuestas

    print ("Cargar usuario de prueba")

    with open(fileLocation, 'r', encoding='utf-8') as loaded_json_file:
        data = loaded_json_file.read()

    JsonUPrueba = json.loads(data)

    usuarioPrueba.username = JsonUPrueba["username"]
    
    for preg in JsonUPrueba["preguntas"]:
        print(" - " + str(preg) )
        
        pregtoadd = pregunta()
        pregtoadd.pagina = preg["pagina"]
        pregtoadd.paginaNumero = preg["paginanumero"]
        #pregtoadd.curso =  preg[""]
        pregtoadd.tallerNombre = preg["tallernombre"]
        pregtoadd.tallerNumero = preg["tallernumero"]
        pregtoadd.actividadNombre = preg["actividadnombre"]
        pregtoadd.actividadNumero = preg["actividadnumero"]
        #pregtoadd.blockKey = preg[""]
        pregtoadd.codigo = preg["codigo"]
        #pregtoadd.numero = preg[""]
        pregtoadd.tipo = preg["tipo"]
        pregtoadd.pretest = preg["pretest"]
        pregtoadd.postest = preg["postest"]
        pregtoadd.esDeControl = preg["Es de control"]
        pregtoadd.esDeEncuesta = preg["Es de encuesta"]
        pregtoadd.preguntaEvaluada = preg["preguntacalificada"]
        #pregtoadd.completa = preg[""]
        pregtoadd.multipleRespuesta = preg["multiplerespuesta"]
        pregtoadd.esCorrecta = preg["correcta"]
        pregtoadd.respuesta = preg["respuesta"]
        pregtoadd.respuestas = preg["respuestas"]
        pregtoadd.respuestaCorrecta = preg["respuestacorrecta"]
        pregtoadd.idrespuesta = preg["idrespuesta"]
        pregtoadd.tipoOra = preg["tipoora"]
        pregtoadd.oraCode = preg["oraCode"]
        pregtoadd.numerocontrol = preg["numerocontrol"]
        pregtoadd.consentimiento =preg["consentimiento"]
        pregtoadd.reglamento = preg["reglamento"]
        pregtoadd.deTaller = preg["detaller"]
        pregtoadd.tieneBlockkey = preg["tieneblockkey"]
        #pregtoadd.cantidaddeveces = preg[""]
        #pregtoadd.ncorrectas = preg[""]
        #pregtoadd.nincorrectas = preg[""]
        
        usuarioPrueba.preguntas.append(pregtoadd)
        ##time.sleep(5)
    
    #Variables del usuario de prueba
    usuarioPrueba.totalContestadas = JsonUPrueba["totalcontestadas"]
    usuarioPrueba.totalPreguntasPorTaller = JsonUPrueba["totalpreguntasportaller"]
    usuarioPrueba.totalPreguntasBuenasOMalasPorTaller = JsonUPrueba["totalpreguntasevaportaller"]
    usuarioPrueba.totalpretest = JsonUPrueba["totalpretest"]
    usuarioPrueba.totalpostest = JsonUPrueba["totalpostest"]

    #Variables genarles
    totalPreguntasPorTaller = JsonUPrueba["totalpreguntasportaller"]      
    totalPreguntasPorControl = JsonUPrueba["totalpreguntascontrol"] 
    totalPreguntasEvaluadas = JsonUPrueba["totalpreguntascalificadas"]
    totalPregutnasEncuestas = JsonUPrueba["totalpreguntasporencuesta"]
    totalPreguntasBuenasOMalas = JsonUPrueba["totalpreguntasevaportaller"]
    totalEncuestas = JsonUPrueba["totalencuestas"]
    cantidadControles = JsonUPrueba["cantidadcontroles"]
    cantidadEncuestas = JsonUPrueba["cantidadencuestas"]
    cantidadTipoPreguntaEvaluada = JsonUPrueba["cantidadtipopc"]
    nombresEncuestas = JsonUPrueba["nombresdeencuestas"]
    #usuarioPrueba.totalPreguntasCorrectas = JsonUPrueba["totalpreguntascorrectastaller"]


def RecorrerXLS(fileLocation,usuarios,listadeoras): 
    global cantidadControles   
    global cantidadEncuestas
    global cantidadTipoPreguntaEvaluada
    global totalPreguntasBuenasOMalas
    #global ListaPregutnasEncuestas

    print("Recorrer CVS")
    ##time.sleep(3)

    File = fileLocation
    with open(File, newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        completo = False
        for row in reader:
            nombre = row[0]
            print ("Nombre tomado del csv: " + nombre)
            ##time.sleep(3)
            for usr in usuarios:
                print ("Nombre en la lista de usuarios " + usr.username )
                ##time.sleep(1)
                preg = pregunta()
                #if usr.username.lower().find(nombre.lower()) != -1: 
                #el nombre de usuario tiene que ser identico o puede encontrar un usuario dos veces
                if usr.username.lower() == nombre.lower(): 
                    usuarioActual = nombre
                    print ("Lo encontre " + nombre + " VS " + usr.username)
                    ##time.sleep(1)
                    titulos= row[2].split(">")
                    preg.curso = titulos[0]
                    enunciadopreg = row[4]

                    if( len(titulos) > 1 ):

                        #ubicaciones 
                        loc1 = 1
                        loc2 = 2
                        loc3 = 3
                        loc4 = 4

                        print("tamaño del arreglo titulos " + str(len(titulos)) )
                        if(len(titulos) == 5):
                            completo == True
                        elif(len(titulos) == 4):
                            loc1 = 0
                            loc2 = 1 
                            loc3 = 2
                            loc4 = 3
                            completo == False
                        elif(len(titulos) == 3):
                            loc1 = 0
                            loc2 = 0 
                            loc3 = 1
                            loc4 = 4

                        
                        subtitulo = str(titulos[loc2].lower())
                        if(subtitulo[0] == ' '):
                            print("Tiene un espaciio en blanco")
                            subtitulo = subtitulo[1:]
                            ##time.sleep(1)

                        #subtitulo = quitarAcentos(subtitulo)

                        if( subtitulo.startswith("encuesta") ):
                            print("Encontre la encuesta " + str(titulos[loc2]))
                            nombreencuesta = subtitulo
                            preg.esDeEncuesta = True
                            preg.tallerNumero = -1
                            preg.numeroEncuesta = 0
                            preg.esunapreguntade = "Encuesta"

                            if(nombreencuesta not in nombresEncuestas):
                                print("añadir nombre de encuesta " + nombreencuesta + " a la lista")
                                nombresEncuestas.append(nombreencuesta)
                            print("Encuestas " + str(nombresEncuestas))
                            indexEncuesta = nombresEncuestas.index(nombreencuesta)

                            preg.numeroEncuesta = indexEncuesta + 1
                            preg.deTaller = False

                        elif( subtitulo.find("prueba de diagnóstico") != -1):
                            print("titulo: " + str(titulos[loc2]) )
                            print ("PRUEBA DE DIAGNÓSTICO")
                            preg.pretest = True
                            preg.tallerNumero = 1
                            #time.sleep(5)
                            preg.deTaller = False
                            preg.esunapreguntade = "pre-Test"
                            print("Tenemos una pregunta de diagnóstico para" + usr.username ) 

                        elif( subtitulo.find("post-test") != -1):
                            print("titulo: " + str(titulos[loc2]) )
                            print ("PRUEBA DE DIAGNÓSTICO")
                            preg.postest = True
                            preg.tallerNumero = 1
                            #time.sleep(5)
                            preg.deTaller = False
                            preg.esunapreguntade = "post-Test"
                            print("Tenemos una pregunta final para" + usr.username ) 

                        elif( subtitulo.find("consentimiento") != -1 ):
                            print("titulo: " + str(titulos[loc2]) )
                            preg.consentimiento = True
                            #preg.tallerNumero = 1
                            ##time.sleep(5)
                            preg.deTaller = False
                            preg.esunapreguntade = "consentimiento"
                            print("Tenemos un consentimiento para" + usr.username ) 

                        elif( subtitulo.find("reglamento") != -1):
                            print("titulo: " + str(titulos[2]) )
                            preg.reglamento = True
                            #preg.tallerNumero = 1
                            ##time.sleep(5)
                            preg.deTaller = False
                            preg.esunapreguntade = "reglamento"
                            print("Tenemos un reglamento para" + usr.username )
                        else:
                            print("titulo: " + str(titulos[loc2]))
                            print (" NO Tenemos una encuesta")
                            ##time.sleep(1)
                            #preg.esDeEncuesta = False

                        tallerNombre = titulos[loc1]
                        numerosEnTaller = [int(i) for i in tallerNombre if i.isdigit()]
                        if(len(numerosEnTaller)>0):
                            print(tallerNombre +" numeros en taller " + str(numerosEnTaller))
                            
                            preg.tallerNumero = numerosEnTaller[0]
                            if(preg.esDeControl):
                                preg.numerocontrol = preg.tallerNumero
                            
                            preg.deTaller = True
                            preg.tallerNombre = tallerNombre.replace("Taller " + str(numerosEnTaller[0]) + ": " , "")
                        else:
                            print("No tiene numero en el taller")
                            preg.tallerNumero = -1
                            preg.deTaller = False
                            preg.tallerNombre = tallerNombre
                            ##time.sleep(1)

                            
                        actividadNombre = titulos[loc2]
                        numerosEnActividad =  [int(i) for i in actividadNombre if i.isdigit()]
                        if(len(numerosEnActividad)>0):
                            preg.actividadNumero = numerosEnActividad[0]
                            preg.actividadNombre = actividadNombre.replace("Actividad "+ str(numerosEnActividad[0]) + ": ", "")
                        else:
                            print("No tiene numero en el taller")
                            preg.actividadNumero = -1
                            preg.actvidadNombre = actividadNombre
                            ##time.sleep(1)


                        preg.pagina = titulos[loc3]

                        recap = False
                        
                        if (preg.pagina.lower()).find("control") != -1 :
                            print (preg.pagina.lower() + " contiene la palabra control")
                            preg.esDeControl = True
                            preg.paginaNumero = 0
                            preg.deTaller = False
                            preg.esunapreguntade = "control"
                            ##time.sleep(3)

                        if (preg.pagina.lower()).find("recapitulemos") != -1 :
                            print (preg.pagina.lower() + " contiene la palabra recapitulemos")
                            #preg.esDeControl = True
                            preg.deTaller = False
                            recap = True
                            preg.esunapreguntade = "recapitulemos"
                            ##time.sleep(3)

                        if (preg.pagina.lower()).find("prueba de diagnóstico") != -1 :
                            print (preg.pagina.lower() + " contiene la palabra prueba de diagnóstico")
                            preg.pretest = True
                            preg.tallerNumero = 1
                            #time.sleep(5)
                            preg.deTaller = False
                            preg.esunapreguntade = "pre-test"
                            print("Tenemos una pregunta pretest para" + usr.username ) 

                        if (preg.pagina.lower()).find("post-test") != -1 :
                            print (preg.pagina.lower() + " contiene la palabra prueba de diagnóstico")
                            preg.postest = True
                            preg.tallerNumero = 1
                            #time.sleep(5)
                            preg.deTaller = False
                            preg.esunapreguntade = "post-test"
                            print("Tenemos una pregunta final para" + usr.username ) 


                        if (preg.pagina.lower()).find("pregunta calificada") != -1 or (preg.pagina.lower()).find("pregunta evaluada") != -1:
                            print (preg.pagina.lower() + " contiene la palabra pregunta calificada")
                            #preg.esDeControl = True
                            #preg.esDeControl = True
                            preg.deTaller = False
                            preg.preguntaEvaluada = True
                            preg.esunapreguntade = "preguntas calificadas"
                            ##time.sleep(3)


                        if(preg.pagina.lower()).find("consentimiento") != -1:
                            print (preg.pagina.lower() + " contiene la palabra consentimiento")
                            preg.consentimiento = True
                            preg.deTaller = False
                            preg.esunapreguntade = "consentimiento"
                            #print ( (preg.nombre).lower() + " contiene la palabra condentimiento" )
                            ##time.sleep(3)

                        if(preg.pagina.lower()).find("reglamento") != -1:
                            print (preg.pagina.lower() + " contiene la palabra reglamento")
                            preg.reglamento = True
                            preg.deTaller = False
                            preg.esunapreguntade = "reglamento"
                            ##time.sleep(3)

                        if(preg.pagina.lower()).find("encuesta pagina") != -1 or (preg.pagina.lower()).find("encuesta taller") != -1 :
                            print (preg.pagina.lower() + " contiene la palabra encuesta")
                            preg.esDeEncuesta = True
                            preg.deTaller = False
                            preg.esunapreguntade = "enucesta"
                            ##time.sleep(3)

                        
                        if(recap == False and preg.esDeControl == False and preg.esDeEncuesta == False and preg.preguntaEvaluada == False and preg.consentimiento == False and preg.reglamento == False):
                            paginatext = titulos[loc3].split('#')
                            print("Títulos "+ str(titulos) +" paginatext: " + str(paginatext) )
                            preg.deTaller = True
                            if(len(paginatext) > 1):
                                preg.paginaNumero = int(paginatext[1])
                            #else:
                                ##time.sleep(1)


                        preg.blockKeyCompleto = row[7]

                        blockkeyArray = row[7].split('@')
                        

                        print ("Blockkey Array: " + str(blockkeyArray))

                        if(len(blockkeyArray) > 1 ):
                            preg.tieneBlockkey = True
                        else:
                            preg.tieneBlockkey = False
                            print("ERROR en el blockkey REVISE LOS PUNTO Y COMA")
                            ##time.sleep(6)


                        idRespuesta = ""
                        contestado = False

                        if(preg.tieneBlockkey):

                            preg.blockKey = blockkeyArray[0]
                            preg.tipo = blockkeyArray[1]
                            preg.codigo = blockkeyArray[2]

                            #Asignar tipo si le falta
                            tipo = preg.tipo

                            #print ("ususario " + usr.nombre + " tiene " + str(len(usr.preguntas)) )
                            ##time.sleep(1)
                            if(is_json(row[8])):
                                loaded_json = json.loads(row[8])
                            else:
                                #No se por que pasa esto pero si no se salta esta fila da error
                                continue
                            #print("Json:")
                            #print(loaded_json)
                            print("Values: ")
                            for key, value in loaded_json.items():
                                print (key +" : "+ str(value))

                            if("correct_map" in loaded_json):
                                contestado = True


                            print("\n revisar si tiene intentos...") 
                            if(loaded_json.get("attempts")):
                                print("si tiene intentos")
                                preg.intentos = loaded_json["attempts"]
                                contestado = True

                            print("\n revisar si tiene Score...")
                            if(loaded_json.get("score")):
                                print ("esta pregunta tiene score")
                                if( str(loaded_json["score"])[0] != '{' ):
                                    preg.score = loaded_json["score"]

                            else:
                                if (preg.pretest or preg.postest or preg.esDeControl or preg.preguntaEvaluada):
                                    preg.score = -2
                                    logErrores.append("la pregunta " + preg.blockKeyCompleto+ " contestada por " + usuarioActual + " no tiene score, es una " + preg.esunapreguntade )
                                

                            if( tipo.lower().find("problem+block") != -1):
                                print("\n es un problem block")

                                preg.multipleRespuesta = False
                                


                                interactivo = False
                                if(row[3] != ""):
                                    print("row " + str(row) )
                                    print(" ")
                                    print("json " + str(loaded_json) )
                                    if(preg.deTaller == False):
                                        print("Sí es un caso especial")
                                        ##time.sleep(3)
                                    preg.idrespuesta = row[3]
                                    idRespuesta = row[3]
                                    preg.respuestaCorrecta = row[6]
                                    if "last_submission_time" in loaded_json:
                                        preg.fechaRespuesta = loaded_json["last_submission_time"]

                                    resopu = row[5]

                                    addfromList= False

                                    print("resopu " + str(resopu))

                                    if( len(resopu) > 1 and resopu[0:2] != '{"' ):
                                        preg.respuesta = row[5]
                                        print ("Se guardo respuesta rapida")
                                    else:
                                        print("No hay respuesta rapida")
                                        preg.respuesta = "Presumiblemente RECURSO INTERACTIVO"
                                        addfromList= True
                                        interactivo = True
                                    

                                    if ("correct_map" not in loaded_json):
                                        print ("NO TIENE CORRECT MAP")

                                    if("student_answers" not in loaded_json):
                                        print ("NO TIENE STUDENT ANSWERS")

                                    ##time.sleep(5)
                                    if ( ("correct_map" not in loaded_json) and ("student_answers" not in loaded_json)):
                                        interactivo = True
                                        print("pregunta " + preg.codigo + " es un recurso interactivo")
                                        preg.respuesta += "No se entonctro respuesta en el json"
                                        ##time.sleep(3)

                                    #print ("id respuesta: " + idRespuesta + " " + str(loaded_json[idRespuesta]))
                                    print (" cantidad de respuestas " + str(len(loaded_json["student_answers"][idRespuesta])) )
                                    cantrespuestas = len(loaded_json["student_answers"][idRespuesta])
                                    print (" student answers: " + str(loaded_json["student_answers"][idRespuesta]))
                                    #print (" correctness: " + str(loaded_json[idRespuesta]))
                                    
                                    ##time.sleep(30)

                                    if( isinstance( loaded_json["student_answers"][idRespuesta] ,list ) ):
                                        print("hay mas de una respuesta")
                                        for resp in loaded_json["student_answers"][idRespuesta]:
                                            preg.respuestas.append(resp)
                                            if(addfromList and cantrespuestas >0):
                                                preg.respuesta = "_"
                                                preg.respuesta += str(resp) + "_"
                                    else:
                                        preg.respuestas.append(loaded_json["student_answers"][idRespuesta])
                                        if(addfromList and cantrespuestas >0):
                                            preg.respuesta = "_"
                                            resp1 = loaded_json["student_answers"][idRespuesta]
                                            preg.respuesta += str(resp1) + "_"

                                    
                                    if("correct_map" in loaded_json):
                                        contestado = True
                                        correcta = loaded_json["correct_map"][idRespuesta]["correctness"]
                                        print ("evaluar la respuesta " + str(correcta))

                                        if(correcta == "correct"):
                                            print (idRespuesta + " es correcta")
                                            preg.esCorrecta = True
                                            preg.score = 12
                                        else:
                                            print (idRespuesta + " NO esta correcta")
                                            preg.esCorrecta = False
                                            preg.score = 0
                                    else:
                                        print("no esta la respuesta, pero no debería entrar aquí")
                                        ##time.sleep(3)
                                        #contestado = False


                                    if ("has_saved_answers" in loaded_json):
                                        contestado = False 
                                                                       
                                else:
                                    print("no esta la respuesta " + row[3])
                                    preg.respuesta = "Respuesta no existente"
                                    contestado = False
                                    #if(preg.deTaller == False):
                                    #time.sleep(3)


                                for precheck in usr.preguntas:
                                    #print ("block key [" + preg.o
                                    if(precheck.blockKeyCompleto == preg.blockKeyCompleto and contestado):
                                        print ("esta repetido el block key [" + preg.blockKeyCompleto +"] para el usuario " + usr.nombre)
                                        ##time.sleep(1)
                                        print ("id respuesta [" + precheck.idrespuesta + "] VS [" + preg.idrespuesta +"]")
                                        if(preg.esCorrecta == False):
                                            precheck.esCorrecta = False
                                            precheck.nincorrectas +=1
                                        else:
                                            precheck.ncorrectas += 1
                                        
                                        precheck.cantidaddeveces +=1
                                        
                                        contestado = False
                                        ##time.sleep(5)


                            if(tipo == "vof+block"):
                                print("\n es un verdadero y falso")
                                cantresp = 0
                                resp = ""
                                contestado = True
                                print ("respuestas")
                                print (loaded_json["respuestas"])
                                for reskey, valkey in loaded_json["respuestas"].items():
                                    cantresp +=1
                                    print(reskey + ":" + str(valkey))
                                    preg.respuestas.append(valkey)
                                    resp += ( valkey +"_")
                                if(len(resp) > 1): resp = resp[:-1]
                                preg.respuesta = resp

                                if(loaded_json["score"] == 1):
                                    preg.esCorrecta = True
                                    preg.score = 12
                                else:
                                    preg.esCorrecta = False
                                    preg.score = 12 * float(loaded_json["score"])

                                preg.idrespuesta = find_id(preg.blockKey,'@')

                            if(tipo == "freetextresponse+block"):
                                print("\n es un free text")
                                contestado = True
                                preg.idrespuesta = find_id(preg.blockKey,'@')
                                #preg.intentos =  loaded_json["count_attempts"]
                                preg.respuesta = loaded_json["student_answer"]
                                if( isinstance( loaded_json["student_answer"] ,list ) ):
                                    print("hay mas de una respuesta")
                                    for resp in loaded_json["student_answer"]:
                                        preg.respuestas.append(resp)
                                    else:
                                        preg.respuestas.append = loaded_json["student_answer"]
                                preg.esCorrecta = None
                                if(loaded_json.get("count_attempts")):
                                    preg.intentos = loaded_json["count_attempts"]


                            if(tipo == "openassessment+block"):
                                print ("pregunta con ORA ORA ORA...")
                                contestado = True
                                preg.tipoOra = True
                                if("submission_uuid" in loaded_json):
                                    submiid = loaded_json["submission_uuid"]
                                    contestado = True
                                else:
                                    submiid  = "No existo"
                                    print("No encontre submission_uuid en el json de " + preg.blockKeyCompleto)
                                    ##time.sleep(20)
                                    contestado = False
                                print ("submisionid " + str(submiid))
                                for ora in listadeoras:
                                    if(str(ora.submissioid) == str(submiid)):
                                        print ("Lo encontre " + str(ora.submissioid) + " VS. " + str(submiid) )
                                        preg.oraCode = submiid
                                        contestado = True
                                        preg.score = ora.score
                                        preg.fechaRespuesta = ora.fecharespuesta
                                        preg.respuesta = ora.respuesta
                                        if(ora.score > 7):
                                            preg.esCorrecta = True
                                            #preg.score = 12
                                        else:
                                            preg.esCorrecta = False
                                            #preg.score = ora.score
                                            #preg.score = 0
                                        preg.score = max(0,min(12,ora.score))
                                        
                                ##time.sleep(10)

                            if(tipo == "dialogsquestionsxblock+block"):
                                print("\n es un dialogo question")
                                contestado = True
                                preg.idrespuesta = find_id(preg.blockKey,'@')
                                respu = ""
                                for resp in loaded_json["student_answers"]:                                        
                                    preg.respuestas.append(resp)
                                    respu += resp + "_"
                                if(len(respu) > 1): respu = respu[:-1]
                                preg.respuesta = respu

                                if(loaded_json["score"] == 1):
                                    preg.esCorrecta = True
                                else:
                                    preg.esCorrecta = False

                            if(tipo == "drag-and-drop-v2+block"):
                                print("\n es un drag and drop")
                                if(loaded_json.get("attempts")):
                                    preg.intentos = loaded_json["attempts"]
                                buena = True
                                respu = ""
                                for item in loaded_json["item_state"]:                                        
                                    preg.respuestas.append( loaded_json["item_state"][item]["zone"]  )
                                    respu += loaded_json["item_state"][item]["zone"] +"_"
                                    if (  loaded_json["item_state"][item]["correct"] == False ):
                                        buena = False
                                if(len(respu) > 1): respu = respu[:-1]    
                                preg.respuesta = respu
                                preg.esCorrecta = buena
                                preg.idrespuesta = find_id(preg.blockKey,'@')
                                        
                                ##time.sleep(5)

                            preg.contestado = contestado
                            
                            #if(preg.deTaller == False):
                                #print ( "usuario "+ usr.username +" pregunta " + preg.pagina + " .contestado: " + str(preg.contestado)  )
                                ##time.sleep(5)
                            

                            if preg.consentimiento and (enunciadopreg.find("NOMBRE") != -1 or enunciadopreg.find("RUT") != -1):
                                contestado = False

                            #if preg.consentimiento and contestado:
                                #print("estoy agregando cons: "+enunciadopreg)
                                #time.sleep(5)



                            #Revisar si el usuario de prueba tiene estta pregunta
                            blockkey_encontrado = False
                            for preguntaUprueba in usuarioPrueba.preguntas:
                                if preg.codigo == preguntaUprueba.codigo:
                                    blockkey_encontrado = True 

                            if(contestado and blockkey_encontrado == False and recap == False and preg.consentimiento == False and preg.reglamento == False ):
                                logErrores.append("la pregunta: " + preg.blockKeyCompleto + " de "+ preg.esunapreguntade + ", contestada por " + usuarioActual + " no se encontro en el usuario de prueba ") 

                            if(contestado and blockkey_encontrado and recap == False):


                                print("nombre del usuario " + str(usr.username))
                                if(preg.esDeEncuesta):
                                    print ("agrego la pregunta " + preg.codigo + " de encuesta a " + usr.username + " (encuesta " + str(preg.numeroEncuesta)  + ")")
                                    ##time.sleep(2)
                                if(preg.consentimiento):
                                    print ("agrego la pregunta " + preg.codigo + " de consentimiento a " + usr.username )
                                    ##time.sleep(2)
                                if(preg.reglamento):
                                    print ("agrego la pregunta " + preg.codigo + " de reglamento a " + usr.username )
                                    ##time.sleep(2)
                                
                                print ("Agragar pregunta " + preg.pagina + " " + preg.blockKey + " a " + usr.username)

                                usr.preguntas.append(preg) 

                                
                                if(tipo == "problem+block"): usr.totalPreguntasPBlock += 1
                                if(tipo == "vof+block"): usr.totalPreguntasVF += 1
                                if(tipo == "freetextresponse+block"): usr.totalPreguntasFreeResp += 1
                                if(tipo == "dialogsquestionsxblock+block"): usr.totalPreguntasDiaQues += 1
                                if(tipo == "drag-and-drop-v2+block"): usr.totalPreguntasDragDrop += 1
                                usr.totalContestadas += 1


                                if(preg.esCorrecta):
                                    usr.totalBuenas += 1
                                elif (preg.esCorrecta == False):
                                    usr.totalMalas += 1
                                else:
                                    usr.totalPreguntasFreeResp += 1

                                #guardare cuantas ha respondido por taller-actividad
                                if(preg.esDeControl == False and preg.preguntaEvaluada == False and preg.esDeEncuesta == False and preg.consentimiento == False and preg.reglamento == False and preg.pretest == False and preg.postest == False and preg.tallerNumero > 0 ):
                                    if preg.tallerNumero not in usr.totalPreguntasPorTaller:
                                        usr.totalPreguntasPorTaller[preg.tallerNumero] = {}
                                        usr.totalPreguntasCorrectasPorTaller[preg.tallerNumero] = {} # no contempla preguntas aviertas
                                        usr.totalPreguntasBuenasOMalasPorTaller[preg.tallerNumero] = {}
                                        usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero] = {} #contempla preguntas aviertas

                                    if preg.actividadNumero not in usr.totalPreguntasPorTaller[preg.tallerNumero]:
                                        usr.totalPreguntasPorTaller[preg.tallerNumero][preg.actividadNumero] = 1
                                        #usr.totalPreguntasCorrectasPorTaller[preg.tallerNumero][preg.actividadNumero]= 1
                                    else:
                                        usr.totalPreguntasPorTaller[preg.tallerNumero][preg.actividadNumero] += 1
                                        #usr.totalPreguntasCorrectasPorTaller[preg.tallerNumero][preg.actividadNumero]+= 1

                                    if(preg.esCorrecta == True):
                                        
                                        if( preg.actividadNumero not in usr.totalPreguntasCorrectasPorTaller[preg.tallerNumero]):
                                            usr.totalPreguntasCorrectasPorTaller[preg.tallerNumero][preg.actividadNumero] = 1
                                            usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero][preg.actividadNumero] = 1
                                        else:
                                            usr.totalPreguntasCorrectasPorTaller[preg.tallerNumero][preg.actividadNumero] += 1
                                            usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero][preg.actividadNumero] += 1

                                    if(preg.esCorrecta != None):
                                        if (preg.actividadNumero not in usr.totalPreguntasBuenasOMalasPorTaller[preg.tallerNumero]) :
                                            usr.totalPreguntasBuenasOMalasPorTaller[preg.tallerNumero][preg.actividadNumero] = 1
                                        else:
                                            usr.totalPreguntasBuenasOMalasPorTaller[preg.tallerNumero][preg.actividadNumero] += 1
                                    else:
                                        if (preg.actividadNumero not in usr.totalPreguntasBuenasOMalasPorTaller[preg.tallerNumero]) :
                                            usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero][preg.actividadNumero] = 1
                                        else:
                                            if (preg.actividadNumero not in usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero]) :
                                                usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero][preg.actividadNumero] = 1
                                            else:
                                                usr.totalPreguntasBuenasEnviadasTaller[preg.tallerNumero][preg.actividadNumero] += 1

                                    usr.totalConTaller += 1 
                             
                                #Parte sacada del usuario de prueba
                                
                                print ("Agragar pregunta " + preg.pagina + " " + preg.blockKey + " a " + usr.username)
                                #if(preg.deTaller == False and preg.esDeControl == False):
                                ##time.sleep(3)
                                """
                                #esto sirvio para ver cuando estaba guardando dos veces una preg en un usuario lo dejo por si acaso
                                for p in usr.preguntas:
                                    if p.codigo == preg.codigo:
                                        print("Estoy repitiendo "+preg.codigo+" para "+usr.username+" "+preg.blockKeyCompleto)
                                        print(row)
                                        time.sleep(5)
                                        break
                                """

                                if(preg.fechaRespuesta != None ):
                                    
                                    #print ( "fecha tomada " + str(preg.fechaRespuesta)) #
                                    #fechaTomada = datetime.strptime(preg.fechaRespuesta, '%Y-%m-%dT%H:%M:%SZ')
                                    fechaTomada = preg.fechaRespuesta 

                                    if(usr.ultimaconexion == None):
                                        usr.ultimaconexion = fechaTomada
                                    else:
                                        print (str(fechaTomada) + " -VS- " + str(usr.ultimaconexion) )
                                        #print("formato fecha tomada " + str(type(fechaTomada)) )
                                        #print("formato fecha ultima conexion " +  str(type(usr.ultimaconexion))  )
                                        if( fechaTomada > usr.ultimaconexion ):
                                            print("La fecha obtenida es MAYOR a la guardada")
                                            usr.ultimaconexion = fechaTomada
                                    
                                    ##time.sleep(10)
                                #if str(usr.username.lower()).find(nombreUsuarioPrueba.lower()) != -1 :
                                 #   usuarioPrueba.preguntas.append(preg)
                                  #  print("Añadrir pregunta al usuario de prueba " + preg.codigo)
                                    ##time.sleep(3)
                                ##time.sleep(12)


                                ##time.sleep(3)
                            

                        print("")
                    ##time.sleep(10)

            #print(" pasar a la siguiente fila")    
            ##time.sleep(2)

    #creteXLS (usuarios)
    print (" Recorrido ")


def RecorrerCSVParaEncuesta(fileLocation,usuarios): 
    print("Recorrer CVS en busca de preguntas de encuestas")
    ##time.sleep(3)

    File = fileLocation
    with open(File, newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        completo = False
        for row in reader:
            if(len(row) > 6 ):
                nombre = row[0]
                print ("Nombre tomado del csv: " + nombre)
                ##time.sleep(3)
                for usr in usuarios:
                    print ("Nombre en la lista de usuarios " + usr.username )
                    ##time.sleep(1)
                    if usr.username.lower().find(nombre.lower()) != -1: 
                        #usuarioActual = nombre
                        print ("Lo encontre " + nombre + " VS " + usr.username)
                        ##time.sleep(1)
                        titulos= row[2].split(">")

                        if( len(titulos) > 1 ):

                            #ubicaciones 
                            loc1 = 1
                            loc2 = 2
                            loc3 = 3
                            loc4 = 4

                            print("tamaño del arreglo titulos " + str(len(titulos)) )
                            if(len(titulos) == 5):
                                completo == True
                            elif(len(titulos) == 4):
                                loc1 = 0
                                loc2 = 1 
                                loc3 = 2
                                loc4 = 3
                                completo == False
                            elif(len(titulos) == 3):
                                loc1 = 0
                                loc2 = 0 
                                loc3 = 1
                                loc4 = 4

                            
                            subtitulo = str(titulos[loc2].lower())
                            if(subtitulo[0] == ' '):
                                print("Tiene un espaciio en blanco")
                                subtitulo = subtitulo[1:]
                                ##time.sleep(1)

                            if( subtitulo.startswith("encuesta") ):
                                print("Encontre la encuesta " + str(titulos[loc2]))
                                nombreencuesta = subtitulo

                                if(nombreencuesta in nombresEncuestas):
                                    pregencuesta = subpreguntaencuesta()

                                    indexEncuesta = nombresEncuestas.index(nombreencuesta)
                                    numeroEncuesta = indexEncuesta + 1
                            
                                    paginaNumero = 0
                                    if( str(titulos[loc3]).find("#") > 0):
                                        paginatext = titulos[loc3].split('#')
                                        if(len(paginatext) > 1):
                                            paginaNumero = int(paginatext[1])

                                    tipo = row[1]
                                    idrespuesta = row[3]
                                    nombrepregunta = row[4]
                                    respuesta = row[5]
                                    arrayresp = row[6]

                                    if(tipo == 'freetextresponse'):
                                        nombrepregunta = "p_abierta"
                                        loaded_json = json.loads(row[8])
                                        if("student_answer" in loaded_json):
                                            respuesta = loaded_json['student_answer']

                                    blockkeyArray = row[7].split('@')
                                    if(len(blockkeyArray) > 1 ):
                                        blockKeyCompleto = row[7]
                                        blockKey = blockkeyArray[0]
                                        blockTipo = blockkeyArray[1]
                                        blockkeyArray[2]

                                    #idrespuesta = find_id(blockKey,'@')


                                    if(str(idrespuesta).find("_") > 0):
                                        identificatorios =idrespuesta.split('_')
                                        print(str(identificatorios))
                                        ##time.sleep(5)
                                        pregencuesta.numeroidentificatorio1 = identificatorios[1]
                                        if(len(identificatorios) > 2):
                                            pregencuesta.numeroidentificatorio2 = identificatorios[2]

                                    pregencuesta.usuariorut = usr.RUT
                                    pregencuesta.nombre = usr.ApellidoP + " " + usr.nombre
                                    pregencuesta.nombreusuario = usr.username
                                    pregencuesta.nombreencuesta = nombreencuesta
                                    pregencuesta.nombrepregunta = nombrepregunta
                                    pregencuesta.numeroencuesta = numeroEncuesta
                                    pregencuesta.respuesta = str(respuesta)
                                    pregencuesta.arrayres = str(arrayresp)
                                    pregencuesta.blockkey = blockKeyCompleto
                                    pregencuesta.pagina = paginaNumero
                                    pregencuesta.tipo = tipo
                                    ListaPregutnasEncuestas.append(pregencuesta)
                
    print (" Recorrido ")
    
def is_json(myjson):
  try:
    json_object = json.loads(myjson)
  except ValueError as e:
    return False
  return True

'''
    print ("Totales")
    print ("total de preguntas " + str(totalPreguntas))
    print ("total preguntas PBlock " + str(totalPreguntasPBlock))
    print ("total preguntas VF " + str(totalPreguntasVF))
    print ("total preguntas FReeResp " + str(totalPreguntasFreeResp))
    print ("total preguntas DQues " + str(totalPreguntasDiaQues))
    print ("total preguntas DragAndDrop " + str(totalPreguntasDragDrop))
'''


# Aqui empieza ejecución
t = time.time()
cantidadTalleres = 0
logErrores = []
loaded_file = ""
location_to_save_report = ""

listaUsuarios = ""




try:
	loaded_file = sys.argv[1]
except IndexError:
    loaded_file = "C:\CMM-LabE\Seguimiento CSV\Entradas json\Entrada-IEP rev.json"


if len(sys.argv)>2:
	if sys.argv[2] == '-silent':
	    def print(*args):
	        pass


print("Archivo: " + loaded_file)

with open(loaded_file, 'r') as loaded_json_file:
    data=loaded_json_file.read()

JsonEntrada = json.loads(data)

#print (JsonEntrada)

listaUsuarios = JsonEntrada["listausuarios"]
orafile = JsonEntrada["ora"]
listaTalleres = JsonEntrada ["talleres"]
jsonUsuarioPrueba = JsonEntrada["jsonuserprueba"]

print ("Talleres " + str(listaTalleres))

# Descarga archivo Drive de lista usuarios del curso en específico
curso = JsonEntrada['listausuarios'].split('/')[-1]
codigosSec = curso.split('_')[1]
anoSec = curso.split('_')[2]
programa = codigosSec[3:8]          # 3 - 8 va el programa del curso ej: MEDIA, ELEARN, BASIC
siglasCurso = codigosSec[8:11]      # 8 - 11 van las siglas del curso ej IEP, DPA, SND, etc
sleCurso = codigosSec[11:14]        # 11 - 14 van las siglas del servicio local o territorio (o en su defecto instancia) ej RMP (region metropolitana), COA (costa araucania), CON (conce), PIL (piloto)
numeroSeccion = codigosSec[14:]     # 14 - final van los dos dígitos que indican el numero de la sección ej 01, 02, etc.
nombreArchivoDriveDatosUsuarios = 'DATOS_CPEIP_'+programa+'_'+siglasCurso+'_'+sleCurso+'_SEC'+numeroSeccion
#listaUsuariosFilepath = '/var/www/html/seguimiento/datosUsuarios.xlsx'


listaUsuariosFilepath = 'datosUsuarios.xlsx'
if os.path.isfile(listaUsuariosFilepath):
    os.remove(listaUsuariosFilepath)

codigodrive_listausuarios = JsonEntrada["codigodrive_listausuarios"]
driveapi.downloadFile(listaUsuariosFilepath,codigodrive_listausuarios,nombreArchivoDriveDatosUsuarios)

#########################################

# Descarga archivo Drive lista negra 
#listaNegraFilePath = '/var/www/html/seguimiento/listaNegra.xlsx'
listaNegraFilePath = 'listaNegra.xlsx'
codigodrive_listanegra = JsonEntrada["codigodrive_listanegra"]
driveapi.downloadFile(listaNegraFilePath,codigodrive_listanegra,'listaUsuariosEquipo')


# Descraga de version anterior del reporte 
codigoCarpetaPlanillasInput = JsonEntrada["codigodrive_reporte"]
prenombre = programa+'_'+siglasCurso+'_'+sleCurso+'_SEC'+numeroSeccion
nombreReporteDrive = 'REPORTE_'+ prenombre
#ReportePath = '/var/www/html/seguimiento/ReporteDescargado.xlsx'
ReportePath = 'ReporteDescargado.xlsx'

#Codigo para subir la encuesta a drive
codigoCarpetaEncuesta = JsonEntrada["codigodrive_encuesta"]
nombreEncuestaDrive = "ENCUESTA_" + prenombre
nombreAnomaliasDrive = "ANOMALIAS_"+ prenombre


######################################

logErrores.append("#--------# Registro de anomalias de: " + prenombre + " #--------#") #Añadir nombre del curso a la lista de errores

nombreUsuarioPrueba = "UsuarioPruebaCMMEDU"
#nombreUsuarioPrueba = "ricardof"

usuarios = []
listaoras = []
listanegra = []
ListaPregutnasEncuestas = []
nombresEncuestas =[]
controlesdisponibles = []
#conteoPreguntasEncuestas

print ("Execution")
LoadUsuarioDePrueba(jsonUsuarioPrueba)
crearListaXLS(listaUsuarios) #llenar la lista de usuarios
LoadDatosUsuario(listaUsuariosFilepath) #agregar datos adicionales
LoadListaNega(listaNegraFilePath) #Eliminar ususarios 
quitarUsuarioPruebadelaLista(nombreUsuarioPrueba)
crearListaOra(orafile)
for taller in listaTalleres: #Recorrer talleres 
    print("recorrer CSV Taller "+ taller)
    #time.sleep(1)
    RecorrerXLS(taller,usuarios,listaoras)
for taller2 in listaTalleres:
    print("recorrer CSV para encuesta en "+ taller)
    #time.sleep(1)
    RecorrerCSVParaEncuesta(taller,usuarios)


jsonFile = "nada"
if generarjson :
    jsonFile = createJson(usuarios,prenombre) #Crear achivo Json



#quitarUsuarioPruebadelaLista(nombreUsuarioPrueba)



archivoSalida = creteXLS (usuarios,prenombre)  #Crear archivo .xls
#creteDocumentoEncuesta(usuarios) #Creae archivo encuesta .xls
archivoSalidaEncuesta = creteDocumentoEncuesta2(ListaPregutnasEncuestas,prenombre)  #Creae archivo encuesta2 .xl

#Crear logs de errores
if len(logErrores) < 2 :
    logErrores.append("ESTE CURSO NO PRESENtA ANOMALIAS REGISTRADAS... Excelente")

logErrores.append("#-----------#Fin Anomalias#-----------#") #Añadir nombre del curso a la lista de errores

AnomaliasFile = CreateErrorFile(logErrores,prenombre)


print("Total preguntas por taller: " + str(totalPreguntasPorTaller) + " Cantidad de talleres " + str(cantidadTalleres))
print("Total preguntas por control " + str(totalPreguntasPorControl) + " cantidad de controles " + str(cantidadControles))
print("Total Encuestas " + str(totalEncuestas) + " cantidad de encuestas " + str(cantidadEncuestas))
print("Total preguntas evaluadas" + str(totalPreguntasEvaluadas) + " cantidad preguntas evaluadas " + str(cantidadTipoPreguntaEvaluada) ) 
print( "Preguntas que pueden ser buenas o malas: " + str(totalPreguntasBuenasOMalas) )
print("Encuestas " + str(nombresEncuestas) )
#print ("registro de anomalias " + str(logErrores))
#print("Lista preguntas encuestas " + str(ListaPregutnasEncuestas))
listausr ="Lista usuarios: "
for usrprint in usuarios:
    listausr += ","  + str(usrprint.username) + " "
print (listausr)



if(subirDrive):
    # Subir reporte a planillas input en Drive
    driveapi.uploadFile(archivoSalida, codigoCarpetaPlanillasInput, nombreReporteDrive)
    
    # Subir registro de anomalias a drive
    driveapi.uploadFile(AnomaliasFile, codigoCarpetaPlanillasInput, nombreAnomaliasDrive)

    if generarjson and jsonFile != "nada":
        driveapi.uploadFile(jsonFile, codigoCarpetaPlanillasInput, nombreAnomaliasDrive)


    # Subir documento de encuesta a la carpeta en drive
    driveapi.uploadFile(archivoSalidaEncuesta, codigoCarpetaEncuesta, nombreEncuestaDrive)
    #############################


print ("Finished")
if len(sys.argv)>2:
	if sys.argv[2]=='-silent':
		del print
print("tiempo empleado: "+str(round(time.time() - t,2))+" seg")

#autoabrir archivo excel
try:
	subprocess.check_output(["start", "C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL", archivoSalida],stderr=subprocess.STDOUT,shell=True)
except:
	pass